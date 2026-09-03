import os
import io
import re
import csv
import gzip
import json
import math
import time
import base64
import hmac
import secrets
import threading
import smtplib
import urllib.request
import urllib.error
from email.message import EmailMessage
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import quote as _url_quote, urlencode as _urlencode

import psycopg2
import psycopg2.extras
import psycopg2.pool
from jinja2.sandbox import SandboxedEnvironment
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, g, flash, jsonify, Response, send_from_directory
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

try:
    import mammoth
except Exception:
    # mammoth kurulu değilse (ör. requirements.txt henüz güncellenmediyse) uygulamanın
    # tamamı çökmesin diye — sadece Word'den tasarım yükleme özelliği devre dışı kalır,
    # geri kalan her şey normal çalışmaya devam eder.
    mammoth = None

try:
    from bs4 import BeautifulSoup
except Exception:
    # aynı mantık: kurulu değilse sadece kenarlık onarma adımı atlanır, uygulama çökmez.
    BeautifulSoup = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except Exception:
    # aynı mantık: kurulu değilse sadece Sayaç Durum Raporu'nun PDF indirme
    # seçeneği devre dışı kalır (yazdırılabilir HTML görünümü etkilenmez),
    # uygulamanın geri kalanı çökmez.
    A4 = mm = colors = SimpleDocTemplate = Table = TableStyle = None
    Paragraph = Spacer = PageBreak = RLImage = ParagraphStyle = pdfmetrics = TTFont = None

DATABASE_URL = os.environ.get("DATABASE_URL")
SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    # ESKİDEN burada GitHub'da herkese açık, sabit bir "yedek" anahtar vardı
    # ("gelistirme-icin-degistir"). SECRET_KEY oturum çerezlerini imzalamak için
    # kullanılıyor — bilinen/sabit bir değerle çalışırsa, bu değeri bilen biri
    # sahte bir oturum çerezi üretip HERHANGİ bir kullanıcı olarak giriş
    # yapabilirdi. Bu yüzden artık ortam değişkeni tanımlı değilse uygulama
    # sessizce zayıf bir anahtara düşmek yerine hiç başlamıyor.
    raise RuntimeError(
        "SECRET_KEY ortam değişkeni tanımlı değil. Güvenlik nedeniyle uygulama "
        "bilinen/sabit bir yedek anahtarla çalışmayı reddediyor. DigitalOcean "
        "panelinde bileşen ayarlarına rastgele, uzun bir SECRET_KEY ortam "
        "değişkeni ekleyip yeniden deploy edin."
    )

app = Flask(__name__)
app.secret_key = SECRET_KEY

# DigitalOcean App Platform, uygulamanın önünde kendi ters vekil (reverse
# proxy) sunucusunu çalıştırıyor — yani gerçek ziyaretçi IP'si doğrudan
# request.remote_addr'da değil, proxy'nin eklediği X-Forwarded-For başlığında
# gelir. ProxyFix, Werkzeug'a TEK bir güvenilir vekil katmanının arkasında
# olduğumuzu (x_for=1) söyleyip request.remote_addr'ı bu başlıktan doğru
# şekilde dolduruyor — aksi halde aşağıdaki giriş deneme sınırlaması (IP'ye
# göre) herkesi aynı (proxy'nin) IP'siymiş gibi görüp ya hep birlikte kilitler
# ya da hiç işe yaramaz.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1)

# Oturum çerezi (session cookie) güvenlik ayarları — önceden hiç ayarlanmamıştı,
# yani Flask'ın varsayılanları geçerliydi (SESSION_COOKIE_SECURE=False). Bunları
# açıkça ayarlıyoruz:
#   - HTTPONLY: JavaScript'in çerezi okumasını engeller (zaten Flask varsayılanı
#     True'dur, ama açıkça belirtmek daha güvenli/okunaklı).
#   - SAMESITE=Lax: çerezin başka sitelerden yapılan (form gönderimi gibi)
#     isteklerle otomatik gönderilmesini engeller — CSRF riskini azaltır.
#   - SECURE: çerezin sadece HTTPS üzerinden gönderilmesini zorunlu kılar.
#     Geliştirme (FLASK_DEBUG=1, yerel http://localhost) sırasında kapalı
#     kalır, aksi halde yerel testte oturum hiç açılmaz; DigitalOcean'daki
#     canlı ortam zaten HTTPS olduğu için üretimde bu her zaman True olur.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_DEBUG", "0") != "1"

# --- CSRF (siteler arası istek sahteciliği) koruması ------------------------
# Üçüncü parti bir pakete bağımlı olmadan, oturuma (session) bağlı bir "eşleşen
# token" deseni kullanılıyor: her oturum için rastgele, tahmin edilemez bir
# token üretilip session'da saklanır (generate_csrf/csrf_token). Veri
# değiştiren (POST/PUT/PATCH/DELETE) her istekte, formdan (ya da JSON
# isteklerde X-CSRFToken başlığından) gelen token bununla karşılaştırılır;
# eşleşmezse istek reddedilir. Bu, başka bir sitedeki kötü niyetli bir sayfanın
# ya da bir bağlantının, oturumu açık bir kullanıcının tarayıcısını kullanarak
# onun adına habersizce POST isteği göndermesini (CSRF) engeller — ör. giriş
# yapmış birine gönderilen sahte bir linke tıklanması sonucu istemeden veri
# silinmesi/değiştirilmesi gibi.


def generate_csrf():
    """Oturum için CSRF token'ı döner; oturumda henüz yoksa yeni, rastgele bir
    tane üretip session'a kaydeder. Şablonlarda {{ csrf_token() }} olarak
    çağrılabilmesi için aşağıda Jinja global'i olarak da kaydediliyor."""
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]


app.jinja_env.globals["csrf_token"] = generate_csrf


@app.before_request
def _csrf_dogrula():
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return None
    beklenen = session.get("csrf_token")
    gelen = (
        request.form.get("csrf_token")
        or request.headers.get("X-CSRFToken")
        or request.headers.get("X-CSRF-Token")
    )
    if not beklenen or not gelen or not hmac.compare_digest(beklenen, gelen):
        return (
            "Güvenlik doğrulaması başarısız oldu (sayfa çok uzun süre açık "
            "kalmış ya da oturum yenilenmiş olabilir). Lütfen sayfayı "
            "yenileyip işlemi tekrar deneyin.",
            400,
        )
    return None


# Yanlışlıkla (veya kötü niyetle) çok büyük dosya yüklenip sunucunun
# yorulmasını önlemek için tüm istekler için üst sınır (fotoğraf/video
# yükleme özelliği eklenince gerekli oldu). Birden fazla video aynı anda
# yüklenebildiği için tekil dosya sınırından (60 MB) daha geniş tutuluyor.
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024


@app.template_filter('tl')
def tl_format(deger):
    try:
        deger = float(deger or 0)
    except (TypeError, ValueError):
        deger = 0.0
    s = f"{deger:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return s


@app.template_filter('trsaat')
def tr_saat(deger):
    """Veritabanında NOW() ile (sunucu/DB saati UTC olduğu için UTC olarak)
    oluşan created_at/updated_at gibi 'bu ne zaman oldu' zaman damgalarını,
    ekranda gösterilmeden önce Türkiye saatine (UTC+3, yaz saati uygulaması
    yok) çevirir. Kullanıcının kendi seçtiği takvim tarihi alanları
    (fatura_tarihi, montaj_tarihi, gonderim_tarihi vb.) bu dönüşümü GEREKTİRMEZ
    — onlar zaten saat taşımayan salt tarihlerdir, bu filtre onlara
    uygulanmamalıdır."""
    if deger is None:
        return None
    try:
        return deger + timedelta(hours=3)
    except TypeError:
        return deger


def _fatura_no_temizle(deger):
    s = str(deger or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


_TURKCE_KUCUK_HARF_TABLOSU = str.maketrans({
    "İ": "i", "I": "i", "ı": "i", "i": "i",
    "Ç": "c", "ç": "c",
    "Ğ": "g", "ğ": "g",
    "Ö": "o", "ö": "o",
    "Ş": "s", "ş": "s",
    "Ü": "u", "ü": "u",
})


def _turkce_normallestir(deger):
    """Arama kutusuna büyük/küçük harf ya da Türkçe'ye özgü İ/I/ı/i, Ç/Ş/Ğ/Ö/Ü
    farkı gözetmeden yazılabilmesi için kullanılıyor. Postgres'in kendi
    LOWER()/ILIKE'ı veritabanının "locale" ayarına bağlı çalışıyor — birçok
    yönetilen (managed) Postgres kurulumu varsayılan olarak "C"/"C.UTF-8"
    locale kullanıyor, bu da Türkçe'ye özgü harfleri (özellikle İ/ı ve
    Ç/Ş/Ğ/Ö/Ü) tanımadığı için küçük harfe çevirmiyor — sonuçta "İBRAHİM"
    yazılı bir kayıt "ibrahim" ile aranınca bulunamıyor. Bunun yerine hem
    aranan kelimeyi (burada) hem veritabanı tarafındaki kolonu (bkz.
    _turkce_esle_kosul) SABİT, locale'den bağımsız bir çeviriyle küçük harfe
    çeviriyoruz ki ikisi her zaman aynı şekilde eşleşsin."""
    return (deger or "").translate(_TURKCE_KUCUK_HARF_TABLOSU).lower()


def _turkce_esle_kosul(kolon_ifadesi):
    """`_turkce_normallestir` ile birebir aynı çeviriyi SQL tarafında (bir
    kolon/ifade için) uygulayan, `LIKE %s` ile birlikte kullanılacak ifadeyi
    döndürür. `kolon_ifadesi` her zaman kod içinde sabit (kullanıcıdan
    gelmiyor), bu yüzden doğrudan string birleştirme güvenlidir."""
    return (
        f"LOWER(TRANSLATE({kolon_ifadesi}, "
        "'İIıiÇĞÖŞÜçğöşü', 'iiiicgosucgosu'))"
    )


def _telefon_formatla(deger):
    rakamlar = "".join(ch for ch in str(deger or "") if ch.isdigit())
    if not rakamlar:
        return ""
    if len(rakamlar) == 10:
        rakamlar = "0" + rakamlar
    if len(rakamlar) != 11:
        return str(deger).strip()
    return f"{rakamlar[0]} {rakamlar[1:4]} {rakamlar[4:7]} {rakamlar[7:9]} {rakamlar[9:11]}"


def _kimlik_no_temizle(deger):
    """Fatura kesme (Hızlı Bilişim e-Connect) için abone/arıza kaydına girilen
    TC Kimlik No (11 hane, bireysel) ya da Vergi No (10 hane, kurumsal)
    alanından rakam olmayan karakterleri (boşluk, tire vb.) temizler.
    Uzunluğuna bakarak hangisi olduğunu ayırt etmek _fatura_turu_belirle()
    içinde yapılır."""
    return "".join(ch for ch in str(deger or "") if ch.isdigit())


def _sayi_veya(deger, varsayilan=0):
    """Form alanından gelen bir metni (virgüllü/noktalı, boş olabilir)
    float'a çevirir; çevrilemezse varsayilan değeri döner (None de olabilir,
    ör. opsiyonel bir alan boş bırakıldıysa)."""
    if deger is None or str(deger).strip() == "":
        return varsayilan
    try:
        return float(str(deger).strip().replace(",", "."))
    except (TypeError, ValueError):
        return varsayilan


BIRLER = ["", "Bir", "İki", "Üç", "Dört", "Beş", "Altı", "Yedi", "Sekiz", "Dokuz"]
ONLAR = ["", "On", "Yirmi", "Otuz", "Kırk", "Elli", "Altmış", "Yetmiş", "Seksen", "Doksan"]
BASAMAK = ["", "Bin", "Milyon", "Milyar", "Trilyon"]


def _uc_basamak_oku(sayi):
    yuzler = sayi // 100
    kalan = sayi % 100
    onlar = kalan // 10
    birler = kalan % 10
    metin = ""
    if yuzler > 0:
        if yuzler == 1:
            metin += "Yüz "
        else:
            metin += BIRLER[yuzler] + " Yüz "
    if onlar > 0:
        metin += ONLAR[onlar] + " "
    if birler > 0:
        metin += BIRLER[birler] + " "
    return metin.strip()


def _sayi_yaziya_cevir(sayi):
    sayi = int(sayi)
    if sayi == 0:
        return "Sıfır"
    parcalar = []
    grup_no = 0
    while sayi > 0:
        grup = sayi % 1000
        if grup > 0:
            grup_metni = _uc_basamak_oku(grup)
            if grup_no == 1 and grup == 1:
                grup_metni = "Bin"
            elif grup_no > 0:
                grup_metni = (grup_metni + " " + BASAMAK[grup_no]).strip()
            parcalar.insert(0, grup_metni)
        sayi //= 1000
        grup_no += 1
    return " ".join(parcalar).strip()


def _tutar_yaziya_cevir(tutar):
    tutar = round(float(tutar or 0), 2)
    tl_kismi = int(tutar)
    kurus_kismi = int(round((tutar - tl_kismi) * 100))
    metin = _sayi_yaziya_cevir(tl_kismi) + " Türk Lirası"
    if kurus_kismi > 0:
        metin += " " + _sayi_yaziya_cevir(kurus_kismi) + " Kuruş"
    return metin


def _odeme_sekli_esle(metin):
    m = (metin or "").strip().upper()
    if any(k in m for k in ["NAKİT", "NAKIT", "ELDEN"]):
        return "nakit"
    if any(k in m for k in ["HAVALE", "BANKA", "EFT"]):
        return "havale"
    if "KART" in m:
        return "kredi_karti"
    if any(k in m for k in ["ÇEK", "CEK"]):
        return "cek"
    return None


DISPLAY_KOLONLARI = [
    ("s_no", "S.No"),
    ("koy_adi", "Köy"),
    ("adi", "Adı"),
    ("soyadi", "Soyadı"),
    ("sayac_no", "Sayaç No"),
    ("senet_tutari", "Senet Tutarı"),
    ("sayac_tutari", "Sayaç Tutarı"),
    ("alinan_tutar", "Alınan"),
    ("sayac_kalan", "Sayaç Kalan"),
    ("malzeme_tutari", "Malzeme Tutarı"),
    ("malzeme_alinan", "Malzeme Alınan"),
    ("malzeme_kalan", "Malzeme Kalan"),
    ("toplam_kalan", "Toplam Kalan"),
    ("senet_no", "Senet No"),
    ("senet_sahibi_adi", "Senet Sahibi Adı"),
    ("senet_sahibi_soyadi", "Senet Sahibi Soyadı"),
    ("telefon", "Telefon"),
    ("telefon2", "Telefon 2"),
    ("baba_adi", "Baba Adı"),
    ("montaj_tarihi", "Montaj Tarihi"),
    ("montaj_personeli", "Montaj Personeli"),
    ("odeme_tarihi", "Ödeme Tarihi"),
    ("odeme_sekli", "Ödeme Şekli"),
    ("odeme_gun_sozu", "Ödeme Gün Sözü"),
    ("odemeyi_gonderen", "Ödemeyi Gönderen"),
    ("aciklama", "Açıklama"),
    ("muhtara_odenecek", "Muhtara Ödenecek"),
    ("muhtara_odenen", "Muhtara Ödenen"),
    ("muhtara_kalan", "Muhtara Kalan"),
    ("fatura_no", "Fatura No"),
]

KOLON_BILGI = {
    "s_no": ("s_no", "sayi"),
    "koy_adi": ("koy_adi", "metin"),
    "adi": ("adi", "metin"),
    "soyadi": ("soyadi", "metin"),
    "sayac_no": ("sayac_no", "metin"),
    "senet_tutari": ("senet_tutari", "sayi"),
    "sayac_tutari": ("sayac_tutari", "sayi"),
    "alinan_tutar": ("alinan_tutar", "sayi"),
    "sayac_kalan": ("(sayac_tutari - alinan_tutar)", "sayi"),
    "malzeme_tutari": ("malzeme_tutari", "sayi"),
    "malzeme_alinan": ("malzeme_alinan", "sayi"),
    "malzeme_kalan": ("(malzeme_tutari - malzeme_alinan)", "sayi"),
    "toplam_kalan": ("(sayac_tutari + malzeme_tutari - alinan_tutar - malzeme_alinan)", "sayi"),
    "senet_no": ("senet_no", "metin"),
    "senet_sahibi_adi": ("senet_sahibi_adi", "metin"),
    "senet_sahibi_soyadi": ("senet_sahibi_soyadi", "metin"),
    "telefon": ("telefon", "metin"),
    "telefon2": ("telefon2", "metin"),
    "baba_adi": ("baba_adi", "metin"),
    "montaj_tarihi": ("montaj_tarihi", "tarih"),
    "montaj_personeli": ("montaj_personeli", "metin"),
    "odeme_tarihi": ("odeme_tarihi", "tarih"),
    "odeme_sekli": ("odeme_sekli", "metin"),
    "odeme_gun_sozu": ("odeme_gun_sozu", "tarih"),
    "odemeyi_gonderen": ("odemeyi_gonderen", "metin"),
    "aciklama": ("aciklama", "metin"),
    "muhtara_odenecek": ("muhtara_odenecek", "sayi"),
    "muhtara_odenen": ("muhtara_odenen", "sayi"),
    "muhtara_kalan": ("(muhtara_odenecek - muhtara_odenen)", "sayi"),
    "fatura_no": ("fatura_no", "metin"),
}

SAYISAL_KOLONLAR = {k for k, (_, tur) in KOLON_BILGI.items() if tur == "sayi"}
RENK_KOLONLARI = {"sayac_kalan", "malzeme_kalan", "toplam_kalan", "muhtara_kalan"}

ARIZA_DISPLAY_KOLONLARI = [
    ("s_no", "S.No"),
    ("ozel_s_no", "Özel S.No"),
    ("koy_adi", "Köy Adı"),
    ("yeni_seri_no", "Yeni Seri No"),
    ("seri_no", "Seri No"),
    ("adi", "Adı"),
    ("soyadi", "Soyadı"),
    ("telefon", "Telefon"),
    ("telefon2", "Telefon 2"),
    ("ariza_ucret", "Arıza Ücret"),
    ("alinan_ucret", "Alınan Ücret"),
    ("kalan_ucret", "Kalan Ücret"),
    ("gelis_tarihi", "Geliş Tarihi"),
    ("takilan_tarih", "Takılan Tarih"),
    ("teslim_tarihi", "Teslim Edilen Tarih"),
    ("sayac_kredisi", "Sayaç Kredisi"),
    ("tespit_edilen_ariza", "Tespit Edilen Arıza"),
    ("tespit_aciklama", "Tespit Açıklama"),
    ("yapilan_islemler", "Yapılan İşlemler"),
    ("islem_aciklama", "İşlem Açıklama"),
]

ARIZA_KOLON_BILGI = {
    "s_no": ("s_no", "sayi"),
    "ozel_s_no": ("ozel_s_no", "metin"),
    "koy_adi": ("koy_adi", "metin"),
    "yeni_seri_no": ("yeni_seri_no", "metin"),
    "seri_no": ("seri_no", "metin"),
    "adi": ("adi", "metin"),
    "soyadi": ("soyadi", "metin"),
    "telefon": ("telefon", "metin"),
    "telefon2": ("telefon2", "metin"),
    "ariza_ucret": ("ariza_ucret", "sayi"),
    "alinan_ucret": ("alinan_ucret", "sayi"),
    "kalan_ucret": ("(ariza_ucret - alinan_ucret)", "sayi"),
    "gelis_tarihi": ("gelis_tarihi", "tarih"),
    "takilan_tarih": ("takilan_tarih", "tarih"),
    "teslim_tarihi": ("teslim_tarihi", "tarih"),
    "sayac_kredisi": ("sayac_kredisi", "metin"),
    "tespit_edilen_ariza": ("tespit_edilen_ariza", "metin"),
    "tespit_aciklama": ("tespit_aciklama", "metin"),
    "yapilan_islemler": ("yapilan_islemler", "metin"),
    "islem_aciklama": ("islem_aciklama", "metin"),
}

ARIZA_SAYISAL_KOLONLAR = {k for k, (_, tur) in ARIZA_KOLON_BILGI.items() if tur == "sayi"}

# Fabrika / Tamir listesi — Abone Listesi / Arıza Takip'teki aynı sütun
# filtreleme sistemi (üst arama kutusu + "hangi alanlarda aransın" onay
# kutuları + her sütunun kendi "Tümü ▾" çoklu seçim kutusu) buraya da
# uygulanıyor. "ft." takısı KULLANILMIYOR — bu ifadeler hem fabrika_listesi()
# içindeki (alias'lı) sorguda hem de /api/kolon-secenekleri'nin (alias'sız,
# tek tablo üstünde çalışan) sorgusunda aynen kullanılabilsin diye kasıtlı.
FABRIKA_DISPLAY_KOLONLARI = [
    ("sira_no", "Sıra No"),
    ("seri_no", "Seri No"),
    ("abone_adi", "Abone Adı"),
    ("koy_adi", "Köy Adı"),
    ("telefon", "Telefon"),
    ("uretim_yili", "Üretim Yılı"),
    ("tespit_edilen_ariza", "Tespit Edilen Arıza"),
    ("yerine_sayac_takildi", "Yerine Sayaç"),
    ("abone_karti", "Abone Kartı"),
    ("durum", "Durum"),
    ("gonderim_tarihi", "Gönderim Tarihi"),
    ("donus_tarihi", "Dönüş Tarihi"),
    ("tamir_ucreti", "Tamir Ücreti"),
]

FABRIKA_KOLON_BILGI = {
    "sira_no": ("sira_no", "sayi"),
    "seri_no": ("seri_no", "metin"),
    "abone_adi": ("abone_adi", "metin"),
    "koy_adi": ("koy_adi", "metin"),
    "telefon": ("telefon", "metin"),
    "uretim_yili": ("uretim_yili", "metin"),
    "tespit_edilen_ariza": ("tespit_edilen_ariza", "metin"),
    "yerine_sayac_takildi": ("CASE WHEN yerine_sayac_takildi THEN 'Takıldı' ELSE 'Takılmadı' END", "metin"),
    "abone_karti": ("CASE WHEN abone_karti = 'alindi' THEN 'ALINDI' ELSE 'ALINMADI' END", "metin"),
    "durum": ("durum", "metin"),
    "gonderim_tarihi": ("gonderim_tarihi", "tarih"),
    "donus_tarihi": ("donus_tarihi", "tarih"),
    "tamir_ucreti": ("tamir_ucreti", "sayi"),
}

FABRIKA_SAYISAL_KOLONLAR = {k for k, (_, tur) in FABRIKA_KOLON_BILGI.items() if tur == "sayi"}

_FABRIKA_ALAN_TANIMLARI = [
    ("seri_no", "Seri No", "seri_no", False),
    ("abone_adi", "Abone Adı", "abone_adi", False),
    ("koy_adi", "Köy Adı", "koy_adi", False),
    ("telefon", "Telefon", "telefon", False),
    ("uretim_yili", "Üretim Yılı", "uretim_yili", False),
    ("tespit_edilen_ariza", "Tespit Edilen Arıza", "tespit_edilen_ariza", False),
    ("takilan_sayac_serisi", "Takılan Sayaç Serisi", "takilan_sayac_serisi", False),
    ("odeyen", "Ödeyen", "odeyen", False),
    ("tamir_ucreti", "Tamir Ücreti", "tamir_ucreti", True),
    ("parca_maliyeti", "Parça Maliyeti", "parca_maliyeti", True),
    ("sira_no", "Sıra No", "sira_no", True),
]
_FABRIKA_ALAN_HARITASI = {k: (kolon, sayisal) for k, _, kolon, sayisal in _FABRIKA_ALAN_TANIMLARI}


def _fabrika_kolon_takimi(db=None):
    """Fabrika/Tamir listesi için (kolon_listesi, kolon_bilgi, sayisal_kolonlar,
    ozel_alanlar) döndürür — abone/arıza'daki özel alan sistemi burada yok,
    ozel_alanlar her zaman boş liste."""
    return FABRIKA_DISPLAY_KOLONLARI, FABRIKA_KOLON_BILGI, FABRIKA_SAYISAL_KOLONLAR, []


def _fabrika_filtre_kosulu_olustur(disari_anahtar, kolon_bilgi):
    """fabrika_listesi() sayfasında o an uygulanmış olan TÜM filtreleri (durum
    sekmesi, köy, arama ve diğer sütun filtreleri) tek bir SQL koşuluna
    çevirir — /api/kolon-secenekleri bir sütunun seçeneklerini hesaplarken
    diğer filtrelerle daraltılmış hâli görsün diye kullanılır."""
    kosul = "silindi_mi IS NOT TRUE"
    params = []
    durum_filtre = request.args.get("durum", "").strip()
    koy = request.args.get("koy", "").strip()
    q = request.args.get("q", "").strip()
    alanlar_secili = request.args.getlist("alan")
    if durum_filtre and durum_filtre in FABRIKA_DURUM_ETIKETLERI:
        kosul += " AND durum = %s"
        params.append(durum_filtre)
    if koy:
        kosul += " AND koy_adi = %s"
        params.append(koy)
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in _FABRIKA_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in _FABRIKA_ALAN_HARITASI:
                kolon, sayisal = _FABRIKA_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            kosul += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params
    for anahtar in kolon_bilgi.keys():
        if anahtar == disari_anahtar:
            continue
        alt_kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if alt_kosul:
            kosul += f" AND {alt_kosul}"
            params += param_listesi
    return kosul, params


def _fabrika_satir_sozlugu(k):
    """Fabrika/Tamir listesindeki bir satırı, tabloda gösterileceği hazır
    (biçimlendirilmiş) hâle çevirir."""
    return {
        "id": k["id"],
        "sira_no": k["sira_no"],
        "seri_no": k["seri_no"],
        "abone_adi": k["abone_adi"] or "",
        "koy_adi": k["koy_adi"] or "",
        "telefon": _telefon_formatla(k["telefon"]),
        "uretim_yili": k["uretim_yili"] or "",
        "tespit_edilen_ariza": k["tespit_edilen_ariza"] or "",
        "yerine_sayac_takildi": k["yerine_sayac_takildi"],
        "takilan_sayac_serisi": k["takilan_sayac_serisi"] or "",
        "abone_karti": "ALINDI" if k["abone_karti"] == "alindi" else "ALINMADI",
        "durum": k["durum"],
        "gonderim_tarihi": _gg_aa_yyyy(str(k["gonderim_tarihi"])) if k["gonderim_tarihi"] else "",
        "donus_tarihi": _gg_aa_yyyy(str(k["donus_tarihi"])) if k["donus_tarihi"] else "",
        "tamir_ucreti": tl_format(k["tamir_ucreti"]),
        "koli_no": k["koli_no"],
        "koli_gonderim_id": k["koli_gonderim_id"],
    }

ARIZA_ALAN_TANIMLARI = [
    ("adi", "Adı", "adi", False),
    ("alinan_ucret", "Alınan Ücret", "alinan_ucret", True),
    ("ariza_ucret", "Arıza Ücret", "ariza_ucret", True),
    ("gelis_tarihi", "Geliş Tarihi", "gelis_tarihi", False),
    ("islem_aciklama", "İşlem Açıklama", "islem_aciklama", False),
    ("kalan_ucret", "Kalan Ücret", "(ariza_ucret - alinan_ucret)", True),
    ("koy_adi", "Köy Adı", "koy_adi", False),
    ("ozel_s_no", "Özel S.No", "ozel_s_no", False),
    ("s_no", "S.No", "s_no", True),
    ("sayac_kredisi", "Sayaç Kredisi", "sayac_kredisi", False),
    ("seri_no", "Seri No", "seri_no", False),
    ("soyadi", "Soyadı", "soyadi", False),
    ("takilan_tarih", "Takılan Tarih", "takilan_tarih", False),
    ("telefon", "Telefon", "telefon", False),
    ("telefon2", "Telefon 2", "telefon2", False),
    ("teslim_tarihi", "Teslim Edilen Tarih", "teslim_tarihi", False),
    ("tespit_aciklama", "Tespit Açıklama", "tespit_aciklama", False),
    ("tespit_edilen_ariza", "Tespit Edilen Arıza", "tespit_edilen_ariza", False),
    ("yapilan_islemler", "Yapılan İşlemler", "yapilan_islemler", False),
    ("yeni_seri_no", "Yeni Seri No", "yeni_seri_no", False),
]

_ABONE_ALFABETIK_SIRA = [
    "aciklama", "adi", "alinan_tutar", "baba_adi", "fatura_no", "koy_adi",
    "malzeme_alinan", "malzeme_kalan", "malzeme_tutari", "montaj_personeli",
    "montaj_tarihi",
    "muhtara_kalan", "muhtara_odenecek", "muhtara_odenen", "odeme_gun_sozu",
    "odeme_sekli", "odeme_tarihi", "odemeyi_gonderen", "s_no", "sayac_kalan",
    "sayac_no", "sayac_tutari", "senet_no", "senet_sahibi_adi",
    "senet_sahibi_soyadi", "senet_tutari", "soyadi", "telefon", "telefon2",
    "toplam_kalan",
]
_DISPLAY_KOLON_HARITASI = dict(DISPLAY_KOLONLARI)
DISPLAY_KOLONLARI_ALFABETIK = [(k, _DISPLAY_KOLON_HARITASI[k]) for k in _ABONE_ALFABETIK_SIRA]

_OZEL_ALAN_TUR_PG = {"metin": "TEXT", "tarih": "TEXT", "sayi": "REAL"}
_OZEL_ALAN_TUR_ETIKETLERI = {"metin": "Metin", "tarih": "Tarih", "sayi": "Sayı"}

ABONE_FORM_ALAN_SIRASI = [
    ("koy_adi", "Köy Adı"), ("adi", "Adı"), ("soyadi", "Soyadı"), ("sayac_no", "Sayaç No"),
    ("baba_adi", "Baba Adı"), ("telefon", "Telefon"), ("telefon2", "Telefon 2"),
    ("sayac_tutari", "Sayaç Tutarı"), ("alinan_tutar", "Alınan Tutar"),
    ("senet_sahibi_adi", "Senet Sahibi Adı"), ("senet_sahibi_soyadi", "Senet Sahibi Soyadı"),
    ("malzeme_tutari", "Malzeme Tutarı"), ("malzeme_alinan", "Malzeme Alınan"),
    ("montaj_tarihi", "Montaj Tarihi"), ("montaj_personeli", "Montaj Personeli"),
    ("odeme_tarihi", "Ödeme Tarihi"), ("odeme_sekli", "Ödeme Şekli"),
    ("odemeyi_gonderen", "Ödemeyi Gönderen"), ("odeme_gun_sozu", "Ödeme Gün Sözü"),
    ("fatura_no", "Fatura No"), ("muhtara_odenecek", "Muhtara Ödenecek"),
    ("muhtara_odenen", "Muhtara Ödenen"), ("aciklama", "Açıklama"),
]
ARIZA_FORM_ALAN_SIRASI = [
    ("ozel_s_no", "Özel S.No"), ("seri_no", "Seri No"), ("koy_adi", "Köy Adı"),
    ("yeni_seri_no", "Yeni Seri No"), ("adi", "Adı"), ("soyadi", "Soyadı"),
    ("telefon", "Telefon"), ("telefon2", "Telefon 2"),
    ("ariza_ucret", "Arıza Ücret"), ("alinan_ucret", "Alınan Ücret"),
    ("gelis_tarihi", "Geliş Tarihi"), ("takilan_tarih", "Takılan Tarih"),
    ("teslim_tarihi", "Teslim Edilen Tarih"), ("sayac_kredisi", "Sayaç Kredisi"),
    ("tespit_aciklama", "Tespit Edilen Arıza - Açıklama"),
    ("islem_aciklama", "Yapılan İşlemler - Açıklama"),
]


def _ozel_alanlari_getir(db, tablo):
    """Bir tablonun ('abone' ya da 'ariza') aktif özel alanlarını, gösterim
    sırasına göre döndürür (id, kolon_adi, etiket, tur, sonra_gelen_alan
    alanlarını içeren sözlük listesi)."""
    cur = db.cursor()
    cur.execute(
        "SELECT id, kolon_adi, etiket, tur, sonra_gelen_alan FROM ozel_alan "
        "WHERE tablo = %s AND aktif = TRUE ORDER BY sira, id",
        (tablo,),
    )
    satirlar = cur.fetchall()
    cur.close()
    return satirlar


def _ozel_alan_harita(ozel_alanlar):
    """Özel alanları, formda hangi sabit alandan HEMEN SONRA görüneceklerine
    göre gruplar (sonra_gelen_alan -> o alana bağlı özel alanlar listesi,
    kendi aralarında sira/id sırasıyla). '' anahtarı, formun en sonundaki
    "Özel Alanlar" kutusuna düşen (henüz özel olarak yerleştirilmemiş)
    alanları tutar."""
    harita = {}
    for oa in ozel_alanlar:
        anahtar = oa["sonra_gelen_alan"] or ""
        harita.setdefault(anahtar, []).append(oa)
    return harita


def _form_onizleme_sirasi(sabit_alan_sirasi, ozel_alanlar):
    """Özel Alan Ayarları sayfasındaki sürükle-bırak önizlemesi için, sabit
    (koddan gelen) alanlarla özel alanları TEK bir listede, formda göründükleri
    gerçek sırayla birleştirir."""
    harita = _ozel_alan_harita(ozel_alanlar)
    sonuc = []
    for anahtar, etiket in sabit_alan_sirasi:
        sonuc.append({"tip": "sabit", "anahtar": anahtar, "etiket": etiket})
        for oa in harita.get(anahtar, []):
            sonuc.append({
                "tip": "ozel", "anahtar": oa["kolon_adi"], "etiket": oa["etiket"],
                "tur": oa["tur"], "id": oa["id"],
            })
    for oa in harita.get("", []):
        sonuc.append({
            "tip": "ozel", "anahtar": oa["kolon_adi"], "etiket": oa["etiket"],
            "tur": oa["tur"], "id": oa["id"],
        })
    return sonuc


def _abone_kolon_takimi(db):
    """DISPLAY_KOLONLARI/KOLON_BILGI/SAYISAL_KOLONLAR'ı, kullanıcının Özel Alan
    Ayarları'ndan eklediği abone alanlarıyla birleştirip döndürür:
    (kolon_listesi, kolon_bilgi, sayisal_kolonlar, ozel_alanlar)."""
    ozel = _ozel_alanlari_getir(db, "abone")
    kolon_listesi = DISPLAY_KOLONLARI + [(o["kolon_adi"], o["etiket"]) for o in ozel]
    kolon_bilgi = dict(KOLON_BILGI)
    for o in ozel:
        kolon_bilgi[o["kolon_adi"]] = (o["kolon_adi"], o["tur"])
    sayisal = SAYISAL_KOLONLAR | {o["kolon_adi"] for o in ozel if o["tur"] == "sayi"}
    return kolon_listesi, kolon_bilgi, sayisal, ozel


def _ariza_kolon_takimi(db):
    """ARIZA_DISPLAY_KOLONLARI/ARIZA_KOLON_BILGI/ARIZA_SAYISAL_KOLONLAR'ı,
    kullanıcının Özel Alan Ayarları'ndan eklediği arıza alanlarıyla birleştirip
    döndürür: (kolon_listesi, kolon_bilgi, sayisal_kolonlar, ozel_alanlar)."""
    ozel = _ozel_alanlari_getir(db, "ariza")
    kolon_listesi = ARIZA_DISPLAY_KOLONLARI + [(o["kolon_adi"], o["etiket"]) for o in ozel]
    kolon_bilgi = dict(ARIZA_KOLON_BILGI)
    for o in ozel:
        kolon_bilgi[o["kolon_adi"]] = (o["kolon_adi"], o["tur"])
    sayisal = ARIZA_SAYISAL_KOLONLAR | {o["kolon_adi"] for o in ozel if o["tur"] == "sayi"}
    return kolon_listesi, kolon_bilgi, sayisal, ozel


def _ozel_alan_deger_formatla(deger, tur):
    """Bir özel alan değerini, listede gösterilirken diğer alanlarla aynı
    bicimde (para/tarih/metin) göstermek için biçimlendirir."""
    if deger is None:
        return ""
    if tur == "sayi":
        return tl_format(deger)
    if tur == "tarih":
        return _gg_aa_yyyy(str(deger))
    return deger


def _ozel_alan_ekle(db, tablo, etiket, tur):
    """Yeni bir özel alan tanımlar VE ilgili tabloya gerçek bir sütun ekler."""
    if tablo not in ("abone", "ariza"):
        raise ValueError("geçersiz tablo")
    if tur not in _OZEL_ALAN_TUR_PG:
        raise ValueError("geçersiz tür")
    cur = db.cursor()
    cur.execute("SELECT nextval(pg_get_serial_sequence('ozel_alan', 'id')) AS id")
    yeni_id = cur.fetchone()["id"]
    kolon_adi = f"ozel_{yeni_id}"
    cur.execute(
        "INSERT INTO ozel_alan (id, tablo, kolon_adi, etiket, tur, sira) VALUES "
        "(%s, %s, %s, %s, %s, (SELECT COALESCE(MAX(sira), -1) + 1 FROM ozel_alan WHERE tablo = %s))",
        (yeni_id, tablo, kolon_adi, etiket, tur, tablo),
    )
    pg_tur = _OZEL_ALAN_TUR_PG[tur]
    cur.execute(f"ALTER TABLE {tablo} ADD COLUMN {kolon_adi} {pg_tur}")
    db.commit()
    cur.close()


def _ozel_alan_sil(db, ozel_alan_id):
    """Özel alanı formdan/listeden gizler (aktif=FALSE). Veri kaybı olmaması
    için ilgili veritabanı sütunu SİLİNMEZ, sadece pasif hale getirilir."""
    cur = db.cursor()
    cur.execute("UPDATE ozel_alan SET aktif = FALSE WHERE id = %s", (ozel_alan_id,))
    db.commit()
    cur.close()


def _ozel_alan_sirala(db, tablo, sira_listesi):
    """Özel Alan Ayarları sayfasındaki sürükle-bırak önizlemesinden gelen YENİ
    tam sırayı (sabit alan anahtarları + özel alanların kolon_adi'leri, formda
    göründükleri sırayla karışık) kaydeder."""
    if tablo not in ("abone", "ariza"):
        raise ValueError("geçersiz tablo")
    sabit_anahtarlar = {k for k, _ in (ABONE_FORM_ALAN_SIRASI if tablo == "abone" else ARIZA_FORM_ALAN_SIRASI)}
    ozel_kolon_id = {oa["kolon_adi"]: oa["id"] for oa in _ozel_alanlari_getir(db, tablo)}

    cur = db.cursor()
    son_sabit = ""
    sira_sayaci = 0
    for anahtar in sira_listesi:
        if anahtar in sabit_anahtarlar:
            son_sabit = anahtar
            continue
        if anahtar not in ozel_kolon_id:
            continue
        cur.execute(
            "UPDATE ozel_alan SET sonra_gelen_alan = %s, sira = %s WHERE id = %s AND tablo = %s",
            (son_sabit, sira_sayaci, ozel_kolon_id[anahtar], tablo),
        )
        sira_sayaci += 1
    db.commit()
    cur.close()


_ABONE_ALAN_TANIMLARI = [
    ("aciklama", "Açıklama", "aciklama", False),
    ("adi", "Adı", "adi", False),
    ("alinan_tutar", "Alınan", "alinan_tutar", True),
    ("baba_adi", "Baba Adı", "baba_adi", False),
    ("fatura_no", "Fatura No", "fatura_no", False),
    ("koy", "Köy", "koy_adi", False),
    ("malzeme_alinan", "Malzeme Alınan", "malzeme_alinan", True),
    ("malzeme_kalan", "Malzeme Kalan", "(malzeme_tutari - malzeme_alinan)", True),
    ("malzeme_tutari", "Malzeme Tutarı", "malzeme_tutari", True),
    ("montaj_personeli", "Montaj Personeli", "montaj_personeli", False),
    ("montaj_tarihi", "Montaj Tarihi", "montaj_tarihi", False),
    ("muhtara_kalan", "Muhtara Kalan", "(muhtara_odenecek - muhtara_odenen)", True),
    ("muhtara_odenecek", "Muhtara Ödenecek", "muhtara_odenecek", True),
    ("muhtara_odenen", "Muhtara Ödenen", "muhtara_odenen", True),
    ("odeme_gun_sozu", "Ödeme Gün Sözü", "odeme_gun_sozu", False),
    ("odeme_sekli", "Ödeme Şekli", "odeme_sekli", False),
    ("odeme_tarihi", "Ödeme Tarihi", "odeme_tarihi", False),
    ("odemeyi_gonderen", "Ödemeyi Gönderen", "odemeyi_gonderen", False),
    ("s_no", "S.No", "s_no", True),
    ("sayac_kalan", "Sayaç Kalan", "(sayac_tutari - alinan_tutar)", True),
    ("sayac_no", "Sayaç No", "sayac_no", False),
    ("sayac_tutari", "Sayaç Tutarı", "sayac_tutari", True),
    ("senet_no", "Senet No", "senet_no", False),
    ("senet_sahibi_adi", "Senet Sahibi Adı", "senet_sahibi_adi", False),
    ("senet_sahibi_soyadi", "Senet Sahibi Soyadı", "senet_sahibi_soyadi", False),
    ("senet_tutari", "Senet Tutarı", "senet_tutari", True),
    ("soyadi", "Soyadı", "soyadi", False),
    ("telefon", "Telefon", "telefon", False),
    ("telefon2", "Telefon 2", "telefon2", False),
    ("toplam_kalan", "Toplam Kalan", "(sayac_tutari + malzeme_tutari - alinan_tutar - malzeme_alinan)", True),
]
_ABONE_ALAN_HARITASI = {k: (kolon, sayisal) for k, _, kolon, sayisal in _ABONE_ALAN_TANIMLARI}


def _abone_filtreli_kayitlari_getir(db):
    """Abone Listesi sayfasındaki (q / koy / alan / deger_*) filtrelerinin aynısını
    mevcut request.args'a göre uygulayarak, filtrelenmiş ham abone satırlarını döndürür."""
    q = request.args.get("q", "").strip()
    koy = request.args.get("koy", "").strip()
    alanlar_secili = request.args.getlist("alan")

    sql = "SELECT * FROM abone WHERE 1=1"
    params = []
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in _ABONE_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in _ABONE_ALAN_HARITASI:
                kolon, sayisal = _ABONE_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            sql += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params
    if koy:
        sql += " AND koy_adi = %s"
        params.append(koy)

    kolon_listesi, kolon_bilgi, _sayisal, _ozel = _abone_kolon_takimi(db)
    for anahtar, _ in kolon_listesi:
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi

    sql += f" ORDER BY s_no {'DESC' if _sira_yonu_al() == 'desc' else 'ASC'}"
    cur = db.cursor()
    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()
    cur.close()
    return kayitlar_ham

_ARIZA_ALFABETIK_SIRA = [
    "adi", "alinan_ucret", "ariza_ucret", "gelis_tarihi", "islem_aciklama",
    "kalan_ucret", "koy_adi", "ozel_s_no", "s_no", "sayac_kredisi", "seri_no",
    "soyadi", "takilan_tarih", "telefon", "telefon2", "teslim_tarihi", "tespit_aciklama",
    "tespit_edilen_ariza", "yapilan_islemler", "yeni_seri_no",
]
_ARIZA_DISPLAY_KOLON_HARITASI = dict(ARIZA_DISPLAY_KOLONLARI)
ARIZA_DISPLAY_KOLONLARI_ALFABETIK = [(k, _ARIZA_DISPLAY_KOLON_HARITASI[k]) for k in _ARIZA_ALFABETIK_SIRA]


def _izgara_satir(n, sutun=4):
    """columns x N onay kutusu ızgarasında kaç satır gerektiğini hesaplar."""
    if n <= 0:
        return 1
    return math.ceil(n / sutun)


def _sira_yonu_al():
    """Şu anki istekten (?sira=asc|desc) sıralama yönünü okur."""
    return "desc" if request.args.get("sira") == "desc" else "asc"


def _sira_toggle_qs():
    """Şu anki tüm filtre/arama parametrelerini (çoklu seçim kutuları dahil)
    koruyarak sadece 'sira' parametresini ters çevrilmiş haliyle döndürür."""
    args = request.args.to_dict(flat=False)
    yeni = "asc" if _sira_yonu_al() == "desc" else "desc"
    args["sira"] = [yeni]
    ciftler = []
    for anahtar, degerler in args.items():
        for deger in degerler:
            ciftler.append((anahtar, deger))
    return _urlencode(ciftler)


SAYFA_BOYUTU = 100


def _sayfa_no_al():
    """Şu anki istekten (?sayfa=N) sayfa numarasını okur."""
    try:
        sayfa = int(request.args.get("sayfa", "1"))
    except (TypeError, ValueError):
        sayfa = 1
    return max(1, sayfa)


def _sayfalama_qs(sayfa):
    """Şu anki tüm filtre/arama/sıralama parametrelerini koruyarak sadece
    'sayfa' parametresini verilen değere ayarlanmış haliyle döndürür."""
    args = request.args.to_dict(flat=False)
    args["sayfa"] = [str(sayfa)]
    ciftler = []
    for anahtar, degerler in args.items():
        for deger in degerler:
            ciftler.append((anahtar, deger))
    return _urlencode(ciftler)


def _sayfala(satirlar):
    """Zaten filtrelenip sıralanmış TAM satır listesini alır; şu anki istekten
    okunan sayfa numarasına göre sadece o sayfaya denk gelen dilimi döndürür."""
    toplam_bulunan = len(satirlar)
    if request.args.get("tumu") == "1":
        return satirlar, toplam_bulunan, 1, 1
    toplam_sayfa = max(1, math.ceil(toplam_bulunan / SAYFA_BOYUTU))
    sayfa = min(_sayfa_no_al(), toplam_sayfa)
    baslangic = (sayfa - 1) * SAYFA_BOYUTU
    return satirlar[baslangic:baslangic + SAYFA_BOYUTU], toplam_bulunan, sayfa, toplam_sayfa


def _tumunu_goster_qs():
    """Şu anki tüm filtre/arama parametrelerini koruyarak 'tumu=1' ekler."""
    args = request.args.to_dict(flat=False)
    args.pop("sayfa", None)
    args["tumu"] = ["1"]
    ciftler = []
    for anahtar, degerler in args.items():
        for deger in degerler:
            ciftler.append((anahtar, deger))
    return _urlencode(ciftler)


FORM_SECENEK_GRUPLARI = {
    "tespit_edilen_ariza": "Tespit Edilen Arıza",
    "yapilan_islemler": "Yapılan İşlemler",
}

_TESPIT_EDILEN_ARIZA_VARSAYILAN = [
    "Arıza Simgesi", "Data", "Dijital Su Almış", "Ekran Yok",
    "Error 1", "Error 2", "Error 3", "Error 4", "Error 5",
    "Harcama Uyuşmuyor", "Harcama Yapmıyor",
    "Kondansatör Devre Dışı", "Kondansatör Yok",
    "Küre Dönmüyor", "Küre Paslı", "Küre Zor Dönüyor",
    "Magnet", "Mekanik Patlak", "Motor Oksitli", "Motor Switch Arızalı",
    "Pil Bitik", "Pil Zayıf", "Sıkıntı Yok",
]

_YAPILAN_ISLEMLER_VARSAYILAN = [
    "Formatlandı",
    "Kart Değişti", "Kart Ekran Değişti", "Kart Okuyucu Değişti", "Kart Temizlendi",
    "Kondansatör Devreye Alındı", "Kondansatör Takıldı",
    "Küre Değişti", "Küre Temizlendi",
    "Mekanik Değişti", "Mekanik Patlak Tamir", "Mekanik Pervane Değişti",
    "Motor Değişti", "Motor Switch Değişti", "Motor Tamir Edildi",
    "Pil Takıldı", "Resetlendi", "Sayım Aparatı Değişti",
]

FORM_SECENEK_VARSAYILANLARI = {
    "tespit_edilen_ariza": _TESPIT_EDILEN_ARIZA_VARSAYILAN,
    "yapilan_islemler": _YAPILAN_ISLEMLER_VARSAYILAN,
}


def _form_secenekleri_getir(db, grup):
    """Bir seçenek grubunun güncel listesini, kayıtlı gösterim sırasına göre
    veritabanından okur."""
    cur = db.cursor()
    cur.execute("SELECT deger FROM form_secenegi WHERE grup = %s ORDER BY sira, id", (grup,))
    satirlar = cur.fetchall()
    cur.close()
    return [s["deger"] for s in satirlar]

GRUP_RENK_PALETI = [
    "#c0392b", "#1f6fb2", "#8e44ad", "#0e8a6d", "#c2740c",
    "#2c3e50", "#c2185b", "#00796b", "#8a6d00", "#5b3a29",
    "#d35400", "#1a7a3c",
]

YEDEKLENECEK_TABLOLAR = [
    "abone", "tahsilat", "ariza", "ariza_tahsilat", "kullanici",
    "fabrika_gonderim", "fabrika_koli", "fabrika_tamir",
]


def _gg_aa_yyyy(t):
    if not t:
        return ""
    t = str(t).strip()
    if len(t) >= 10 and t[4:5] == "-" and t[7:8] == "-":
        return t[8:10] + "." + t[5:7] + "." + t[0:4]
    if len(t) >= 10 and t[2:3] == "." and t[5:6] == ".":
        return t[0:10]
    return t


def _iso_tarih_mi(t):
    t = (t or "").strip()
    return len(t) == 10 and t[4:5] == "-" and t[7:8] == "-"


def _ddmmyyyy_to_iso(t):
    t = (t or "").strip()
    if len(t) == 10 and t[2:3] == "." and t[5:6] == ".":
        return t[6:10] + "-" + t[3:5] + "-" + t[0:2]
    return None


def _tarih_iso_hale_getir(t):
    """<input type=date> alanına doğrudan basılabilecek YYYY-AA-GG biçimini döndürür."""
    t = (t or "").strip()
    if not t:
        return ""
    if _iso_tarih_mi(t):
        return t
    return _ddmmyyyy_to_iso(t) or ""


_KOLON_BOS_DEGER = "__BOS__"


def _kolon_secenekleri(db, anahtar, tablo, bilgi_sozlugu, ekstra_kosul=None, ekstra_params=None):
    """Bir sütun filtre kutusunun seçenek listesini üretir."""
    ifade, tur = bilgi_sozlugu[anahtar]
    ekstra_kosul_sql = f" AND {ekstra_kosul}" if ekstra_kosul else ""
    ekstra_params = list(ekstra_params or [])
    cur = db.cursor()
    if tur == "sayi":
        cur.execute(
            f"SELECT DISTINCT deger FROM (SELECT ROUND(CAST({ifade} AS NUMERIC), 2) AS deger FROM {tablo} WHERE 1=1{ekstra_kosul_sql}) t WHERE deger IS NOT NULL ORDER BY deger",
            ekstra_params,
        )
    else:
        cur.execute(
            f"SELECT DISTINCT {ifade} AS deger FROM {tablo} WHERE {ifade} IS NOT NULL AND {ifade} != ''{ekstra_kosul_sql} ORDER BY deger",
            ekstra_params,
        )
    satirlar = cur.fetchall()

    if tur == "sayi":
        cur.execute(f"SELECT EXISTS (SELECT 1 FROM {tablo} WHERE {ifade} IS NULL{ekstra_kosul_sql}) AS var", ekstra_params)
    else:
        cur.execute(f"SELECT EXISTS (SELECT 1 FROM {tablo} WHERE ({ifade} IS NULL OR {ifade} = ''){ekstra_kosul_sql}) AS var", ekstra_params)
    bos_var = cur.fetchone()["var"]
    cur.close()

    secenekler = []
    if bos_var:
        secenekler.append((_KOLON_BOS_DEGER, "(Boş)"))
    for s in satirlar:
        ham = s["deger"]
        if ham is None:
            continue
        if tur == "sayi":
            metin = tl_format(ham)
        elif tur == "tarih":
            metin = _gg_aa_yyyy(str(ham))
        else:
            metin = str(ham)
        secenekler.append((str(ham), metin))
    return secenekler


def _kolon_kosul_coklu(anahtar, deger_listesi, bilgi_sozlugu):
    ifade, tur = bilgi_sozlugu[anahtar]
    bos_secili = _KOLON_BOS_DEGER in deger_listesi
    diger_degerler = [d for d in deger_listesi if d != _KOLON_BOS_DEGER]
    parcalar = []
    params = []
    if tur == "sayi":
        sayilar = []
        for d in diger_degerler:
            try:
                sayilar.append(round(float(str(d).replace(",", ".")), 2))
            except ValueError:
                pass
        if sayilar:
            yer_tutucular = ", ".join(["%s"] * len(sayilar))
            parcalar.append(f"ROUND(CAST({ifade} AS NUMERIC), 2) IN ({yer_tutucular})")
            params += sayilar
        if bos_secili:
            parcalar.append(f"{ifade} IS NULL")
    else:
        if diger_degerler:
            yer_tutucular = ", ".join(["%s"] * len(diger_degerler))
            parcalar.append(f"{ifade} IN ({yer_tutucular})")
            params += diger_degerler
        if bos_secili:
            parcalar.append(f"({ifade} IS NULL OR {ifade} = '')")
    if not parcalar:
        return None, []
    return "(" + " OR ".join(parcalar) + ")", params


def _kolon_kosul_haric(anahtar, haric_deger_listesi, bilgi_sozlugu):
    """'İşaretli olmayanlar' küçük bir küme olduğunda tüm binlerce işaretli
    değeri adrese eklemek yerine sadece HARİÇ değerleri gönderip NOT IN'e çeviririz."""
    ifade, tur = bilgi_sozlugu[anahtar]
    bos_haric = _KOLON_BOS_DEGER in haric_deger_listesi
    diger_degerler = [d for d in haric_deger_listesi if d != _KOLON_BOS_DEGER]
    params = []
    if tur == "sayi":
        sayilar = []
        for d in diger_degerler:
            try:
                sayilar.append(round(float(str(d).replace(",", ".")), 2))
            except ValueError:
                pass
        diger_degerler = sayilar
        deger_ifadesi = f"ROUND(CAST({ifade} AS NUMERIC), 2)"
        bos_kosulu = f"{ifade} IS NULL"
    else:
        deger_ifadesi = ifade
        bos_kosulu = f"({ifade} IS NULL OR {ifade} = '')"
    if bos_haric:
        kosul = f"NOT {bos_kosulu}"
        if diger_degerler:
            yer_tutucular = ", ".join(["%s"] * len(diger_degerler))
            kosul += f" AND {deger_ifadesi} NOT IN ({yer_tutucular})"
            params += diger_degerler
        return f"({kosul})", params
    if not diger_degerler:
        return None, []
    yer_tutucular = ", ".join(["%s"] * len(diger_degerler))
    return f"({bos_kosulu} OR {deger_ifadesi} NOT IN ({yer_tutucular}))", diger_degerler


def _kolon_secim_kosulu(anahtar, kolon_bilgi):
    """Tek bir sütun için, o an istekte gelen deger_<anahtar> / haric_<anahtar>
    parametrelerinden SQL koşulu ve parametre listesi üretir."""
    haric_secilenler = request.args.getlist(f"haric_{anahtar}")
    if haric_secilenler:
        return _kolon_kosul_haric(anahtar, haric_secilenler, kolon_bilgi)
    secilenler = request.args.getlist(f"deger_{anahtar}")
    if secilenler:
        return _kolon_kosul_coklu(anahtar, secilenler, kolon_bilgi)
    return None, []


def _abone_filtre_kosulu_olustur(disari_anahtar, kolon_bilgi):
    """abone_listesi() sayfasında o an uygulanmış olan TÜM filtreleri tek bir
    SQL koşuluna çevirir."""
    kosul = "1=1"
    params = []
    q = request.args.get("q", "").strip()
    koy = request.args.get("koy", "").strip()
    alanlar_secili = request.args.getlist("alan")
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in _ABONE_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in _ABONE_ALAN_HARITASI:
                kolon, sayisal = _ABONE_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            kosul += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params
    if koy:
        kosul += " AND koy_adi = %s"
        params.append(koy)
    for anahtar in kolon_bilgi.keys():
        if anahtar == disari_anahtar:
            continue
        alt_kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if alt_kosul:
            kosul += f" AND {alt_kosul}"
            params += param_listesi
    return kosul, params


def _ariza_filtre_kosulu_olustur(disari_anahtar, kolon_bilgi):
    """ariza_listesi() / ariza_ciktisi() sayfalarında o an uygulanmış olan
    filtreleri tek bir SQL koşuluna çevirir."""
    kosul = "1=1"
    params = []
    ARIZA_ALAN_HARITASI = {k: (kolon, sayisal) for k, _, kolon, sayisal in ARIZA_ALAN_TANIMLARI}
    q = request.args.get("q", "").strip()
    alanlar_secili = request.args.getlist("alan")
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in ARIZA_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in ARIZA_ALAN_HARITASI:
                kolon, sayisal = ARIZA_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            kosul += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params
    for anahtar in kolon_bilgi.keys():
        if anahtar == disari_anahtar:
            continue
        alt_kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if alt_kosul:
            kosul += f" AND {alt_kosul}"
            params += param_listesi
    return kosul, params


def _abone_satir_sozlugu(k, ozel_alanlar=None):
    sayac_kalan = (k["sayac_tutari"] or 0) - (k["alinan_tutar"] or 0)
    malzeme_kalan = (k["malzeme_tutari"] or 0) - (k["malzeme_alinan"] or 0)
    toplam_kalan = sayac_kalan + malzeme_kalan
    muhtara_kalan = (k["muhtara_odenecek"] or 0) - (k["muhtara_odenen"] or 0)
    ozel_alanlar = ozel_alanlar or []
    renk = {anahtar: '' for anahtar, _ in DISPLAY_KOLONLARI}
    for oa in ozel_alanlar:
        renk[oa["kolon_adi"]] = ''
    renk["sayac_kalan"] = 'kirmizi' if sayac_kalan > 0 else 'yesil'
    renk["malzeme_kalan"] = 'kirmizi' if malzeme_kalan > 0 else 'yesil'
    renk["toplam_kalan"] = 'kirmizi' if toplam_kalan > 0 else 'yesil'
    renk["muhtara_kalan"] = 'kirmizi' if muhtara_kalan > 0 else 'yesil'
    satir = {
        "id": k["id"],
        "s_no": k["s_no"],
        "koy_adi": k["koy_adi"],
        "adi": k["adi"],
        "soyadi": k["soyadi"],
        "sayac_no": k["sayac_no"],
        "senet_tutari": tl_format(k["senet_tutari"]),
        "sayac_tutari": tl_format(k["sayac_tutari"]),
        "alinan_tutar": tl_format(k["alinan_tutar"]),
        "sayac_kalan": tl_format(sayac_kalan),
        "malzeme_tutari": tl_format(k["malzeme_tutari"]),
        "malzeme_alinan": tl_format(k["malzeme_alinan"]),
        "malzeme_kalan": tl_format(malzeme_kalan),
        "toplam_kalan": tl_format(toplam_kalan),
        "senet_no": k["senet_no"],
        "senet_sahibi_adi": k["senet_sahibi_adi"],
        "senet_sahibi_soyadi": k["senet_sahibi_soyadi"],
        "telefon": _telefon_formatla(k["telefon"]),
        "telefon2": _telefon_formatla(k["telefon2"]),
        "baba_adi": k["baba_adi"],
        "montaj_tarihi": _gg_aa_yyyy(k["montaj_tarihi"]),
        "montaj_personeli": k["montaj_personeli"],
        "odeme_tarihi": _gg_aa_yyyy(k["odeme_tarihi"]),
        "odeme_sekli": k["odeme_sekli"],
        "odeme_gun_sozu": _gg_aa_yyyy(k["odeme_gun_sozu"]),
        "odemeyi_gonderen": k["odemeyi_gonderen"],
        "aciklama": k["aciklama"],
        "muhtara_odenecek": tl_format(k["muhtara_odenecek"]),
        "muhtara_odenen": tl_format(k["muhtara_odenen"]),
        "muhtara_kalan": tl_format(muhtara_kalan),
        "fatura_no": _fatura_no_temizle(k["fatura_no"]),
        "_renk": renk,
    }
    for oa in ozel_alanlar:
        satir[oa["kolon_adi"]] = _ozel_alan_deger_formatla(k[oa["kolon_adi"]], oa["tur"])
    return satir


def _ariza_satir_sozlugu(k, ozel_alanlar=None):
    kalan_ucret = (k["ariza_ucret"] or 0) - (k["alinan_ucret"] or 0)
    ozel_alanlar = ozel_alanlar or []
    renk = {anahtar: '' for anahtar, _ in ARIZA_DISPLAY_KOLONLARI}
    for oa in ozel_alanlar:
        renk[oa["kolon_adi"]] = ''
    renk["kalan_ucret"] = 'kirmizi' if kalan_ucret > 0 else 'yesil'
    satir = {
        "id": k["id"],
        "s_no": k["s_no"],
        "ozel_s_no": k["ozel_s_no"],
        "koy_adi": k["koy_adi"],
        "yeni_seri_no": k["yeni_seri_no"],
        "seri_no": k["seri_no"],
        "adi": k["adi"],
        "soyadi": k["soyadi"],
        "telefon": _telefon_formatla(k["telefon"]),
        "telefon2": _telefon_formatla(k["telefon2"]),
        "ariza_ucret": tl_format(k["ariza_ucret"]),
        "alinan_ucret": tl_format(k["alinan_ucret"]),
        "kalan_ucret": tl_format(kalan_ucret),
        "gelis_tarihi": _gg_aa_yyyy(k["gelis_tarihi"]),
        "takilan_tarih": _gg_aa_yyyy(k["takilan_tarih"]),
        "teslim_tarihi": _gg_aa_yyyy(k["teslim_tarihi"]),
        "sayac_kredisi": k["sayac_kredisi"],
        "tespit_edilen_ariza": k["tespit_edilen_ariza"],
        "tespit_aciklama": k["tespit_aciklama"],
        "yapilan_islemler": k["yapilan_islemler"],
        "islem_aciklama": k["islem_aciklama"],
        "_renk": renk,
    }
    for oa in ozel_alanlar:
        satir[oa["kolon_adi"]] = _ozel_alan_deger_formatla(k[oa["kolon_adi"]], oa["tur"])
    return satir


_MONTAJ_FORMU_VARSAYILAN_SABLON = """<div class="montaj-formu-kopya" style="border:1px solid #999; border-radius:6px; padding:14px 18px; margin-bottom:14px; font-size:12.5px; line-height:1.5; color:#1f2a24;">
    <img src="{{ url_for('static', filename='montaj_formu_header.png') }}" alt="ALGI - ELEKTROMED" style="max-width:560px; width:100%; display:block; margin-bottom:8px;">

    <div style="display:flex; justify-content:space-between; margin-bottom:8px; font-weight:700;">
        <div>KÖY ADI&nbsp;&nbsp;:&nbsp;&nbsp;{{ koy_adi }}</div>
        <div>{{ montaj_tarihi }}</div>
    </div>

    <table style="width:100%; border-collapse:collapse; margin-bottom:6px;">
        <tr><td colspan="4" style="border:1px solid #333; background:#eee; text-align:center; font-weight:700; padding:4px;">GARANTİ BELGESİ - ABONE BİLGİLERİ</td></tr>
        <tr>
            <td style="border:1px solid #333; padding:5px; width:18%; font-weight:700; background:#f7f7f7;">Adı Soyadı</td>
            <td style="border:1px solid #333; padding:5px; width:32%;">{{ adi }} {{ soyadi }}</td>
            <td style="border:1px solid #333; padding:5px; width:18%; font-weight:700; background:#f7f7f7;">Sayaç No</td>
            <td style="border:1px solid #333; padding:5px; width:32%;">{{ sayac_no }}</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:5px; font-weight:700; background:#f7f7f7;">İletişim</td>
            <td style="border:1px solid #333; padding:5px;">{{ telefon }}{% if telefon2 %} / {{ telefon2 }}{% endif %}</td>
            <td style="border:1px solid #333; padding:5px; font-weight:700; background:#f7f7f7;">Mekanik Sayaç Alındı</td>
            <td style="border:1px solid #333; padding:5px;">&nbsp;</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:5px; font-weight:700; background:#f7f7f7;">Sayaç Ücreti</td>
            <td style="border:1px solid #333; padding:5px;">{{ sayac_ucreti }}</td>
            <td style="border:1px solid #333; padding:5px; font-weight:700; background:#f7f7f7;">Tesisat Ücreti</td>
            <td style="border:1px solid #333; padding:5px;">{{ tesisat_ucreti }}</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:5px; font-weight:700; background:#f7f7f7;">Alınan</td>
            <td style="border:1px solid #333; padding:5px;">{{ sayac_ucreti_alinan }}</td>
            <td style="border:1px solid #333; padding:5px; font-weight:700; background:#f7f7f7;">Alınan</td>
            <td style="border:1px solid #333; padding:5px;">{{ tesisat_ucreti_alinan }}</td>
        </tr>
    </table>

    <table style="width:100%; border-collapse:collapse; margin-bottom:6px; font-size:11px;">
        <tr><td colspan="6" style="border:1px solid #333; background:#eee; text-align:center; font-weight:700; padding:3px;">MONTAJDA KULLANILAN MALZEMELER</td></tr>
        <tr style="font-weight:700; background:#f7f7f7;">
            <td style="border:1px solid #333; padding:3px;">MALZEME</td><td style="border:1px solid #333; padding:3px; width:8%;">ADET</td>
            <td style="border:1px solid #333; padding:3px;">MALZEME</td><td style="border:1px solid #333; padding:3px; width:8%;">ADET</td>
            <td style="border:1px solid #333; padding:3px;">MALZEME</td><td style="border:1px solid #333; padding:3px; width:8%;">ADET</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:3px;">REKOR 20x20 İÇ DİŞLİ OYNAR BAŞLI</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">MAŞON 32x32 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">MAŞON 20x20 KAPLİN</td><td style="border:1px solid #333;">&nbsp;</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:3px;">VANA 20x20 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">TE 20x20x20 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">REDÜKSİYON 25x20 KAPLİN</td><td style="border:1px solid #333;">&nbsp;</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:3px;">DİRSEK 20x20 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">REDÜKSİYON 25x20 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">REDÜKSİYON 32x20 KAPLİN</td><td style="border:1px solid #333;">&nbsp;</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:3px;">MAŞON 20x20 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">REDÜKSİYON 32x20 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">DİRSEK 20x20 KAPLİN</td><td style="border:1px solid #333;">&nbsp;</td>
        </tr>
        <tr>
            <td style="border:1px solid #333; padding:3px;">MAŞON 25x25 PPRC</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">KORUMA KABI</td><td style="border:1px solid #333;">&nbsp;</td>
            <td style="border:1px solid #333; padding:3px;">&nbsp;</td><td style="border:1px solid #333;">&nbsp;</td>
        </tr>
    </table>

    <div style="margin:6px 0;">
        <strong style="color:#c0392b;">ÖDEME BİLGİLERİ</strong><br>
        HESAP SAHİBİ : SEVDİYE ÇOBAN<br>
        HALK BANKASI IBAN: TR95 0001 2009 6680 0009 0053 87
    </div>

    <div style="margin:6px 0; font-size:11.5px;">
        {{ montaj_tarihi }} Tarihinde sayacım firma tarafından bilgim dahilinde değiştirildi. Sayacın garanti şartları aşağıda belirtilmiştir.
    </div>

    <div style="font-size:11px;">
        <strong>Garanti Şartları</strong><br>
        1 - Yüklenici sayaçları 10 (On) yıl boyunca garanti kapsamı dışında bedeli karşılığı teknik servis hizmeti sağlayacaktır.<br>
        2 - Üretici tarafından üretimi yapılan sayaçlar TSE standartlarında olup, garanti süresi 2 (iki) yıl olacaktır.<br>
        3 - Su gibi etkenlere direkt ve devamlı maruz kaldığında garanti dışı işlem yapılacaktır.<br>
        4 - Montajı yapılan sayaçlar abone tarafından belirlenmiştir. Montaj yerinde sonradan doğabilecek sorunlar aboneye aittir.
    </div>

    <div style="display:flex; justify-content:space-between; margin-top:22px; font-size:11.5px; text-align:center;">
        <div style="flex:1;">
            <div style="height:46px;"></div>
            Kurum Personeli
        </div>
        <div style="flex:1;">
            <div style="height:46px; display:flex; align-items:flex-end; justify-content:center;">{{ montaj_imza }}</div>
            {{ montaj_personeli }}<br>ELEKTROMED Yetkili Personeli
        </div>
        <div style="flex:1;">
            <div style="height:46px; display:flex; align-items:flex-end; justify-content:center;">{{ abone_imza }}</div>
            {{ adi }} {{ soyadi }}<br>Abone Veya Vekili
        </div>
    </div>
</div>"""


def _word_tasarimini_onar(html):
    """Word'den (.docx) mammoth ile HTML'e çevrilen bir tasarımda, mammoth'un
    KASITLI olarak atladığı/eklemediği bazı şeyleri onarır."""
    if BeautifulSoup is None or not html:
        return html

    soup = BeautifulSoup(html, "html.parser")

    for p in soup.find_all("p"):
        hucre_icinde = p.find_parent(["td", "th"]) is not None
        p["style"] = "margin:0;" if hucre_icinde else "margin:0 0 2px 0;"

    for p in soup.find_all("p"):
        icerik_cocuklari = [c for c in p.contents if not (isinstance(c, str) and not c.strip())]
        if len(icerik_cocuklari) == 1 and getattr(icerik_cocuklari[0], "name", None) == "img":
            icerik_cocuklari[0]["style"] = "width:100%; height:auto; display:block;"

    if "<table" not in html:
        return str(soup)

    for tablo in soup.find_all("table"):
        satirlar = tablo.find_all("tr")
        if len(satirlar) < 2:
            tek_satir = satirlar[0] if satirlar else None
            if tek_satir is not None:
                tablo["style"] = "width:100%; border-collapse:collapse;"
                hucreler = tek_satir.find_all(["td", "th"], recursive=False)
                resimli = tek_satir.find("img") is not None
                if not resimli:
                    if len(hucreler) == 2:
                        hucreler[-1]["style"] = "text-align:right;"
                    elif len(hucreler) == 3:
                        for h in hucreler:
                            h["style"] = "text-align:center;"
            continue

        tablo["style"] = "width:100%; border-collapse:collapse; table-layout:fixed; margin-bottom:3px;"

        kolon_sayisi = 0
        kolon_uzunluklari = []
        for satir in satirlar:
            hucreler = satir.find_all(["td", "th"], recursive=False)
            if not hucreler or any(h.get("colspan") for h in hucreler):
                continue
            if kolon_sayisi == 0:
                kolon_sayisi = len(hucreler)
                kolon_uzunluklari = [0] * kolon_sayisi
            if len(hucreler) != kolon_sayisi:
                continue
            for i, h in enumerate(hucreler):
                uzunluk = len(h.get_text(strip=True))
                if uzunluk > kolon_uzunluklari[i]:
                    kolon_uzunluklari[i] = uzunluk

        if kolon_sayisi:
            mevcut_colgroup = tablo.find("colgroup")
            if mevcut_colgroup:
                mevcut_colgroup.decompose()
            TABAN_YUZDE = 8
            toplam_uzunluk = sum(kolon_uzunluklari) or kolon_sayisi
            kalan_yuzde = 100 - TABAN_YUZDE * kolon_sayisi
            colgroup = soup.new_tag("colgroup")
            genislikler = []
            for uzunluk in kolon_uzunluklari:
                pay = TABAN_YUZDE + (kalan_yuzde * uzunluk / toplam_uzunluk if kalan_yuzde > 0 else 0)
                genislikler.append(pay)
            fark = 100 - sum(genislikler)
            genislikler[-1] += fark
            for genislik in genislikler:
                col = soup.new_tag("col")
                col["style"] = f"width:{round(genislik, 4)}%;"
                colgroup.append(col)
            tablo.insert(0, colgroup)

        for hucre in tablo.find_all(["td", "th"]):
            stil = ("border:1px solid #333; padding:2px 4px; overflow-wrap:break-word; "
                    "white-space:normal;")
            if hucre.get("colspan"):
                stil += " background:#eee; text-align:center;"
            else:
                icerik_cocuklari = [c for c in hucre.contents if not (isinstance(c, str) and not c.strip())]
                tek_strong = (
                    len(icerik_cocuklari) == 1
                    and getattr(icerik_cocuklari[0], "name", None) == "strong"
                ) or (
                    len(icerik_cocuklari) == 1
                    and getattr(icerik_cocuklari[0], "name", None) == "p"
                    and len(icerik_cocuklari[0].contents) == 1
                    and getattr(icerik_cocuklari[0].contents[0], "name", None) == "strong"
                )
                if tek_strong:
                    stil += " background:#f7f7f7;"
            hucre["style"] = stil
    return str(soup)


_ORNEK_IMZA_YER_TUTUCU = (
    "data:image/svg+xml;utf8,"
    "<svg xmlns='http://www.w3.org/2000/svg' width='150' height='40'>"
    "<text x='75' y='26' font-family='cursive' font-size='15' font-style='italic' "
    "fill='%23999' text-anchor='middle'>(Örnek İmza)</text></svg>"
)


def _imza_veri_url_img_etiketi(veri_url):
    """Bir imza data: URL'ini, Montaj Formu'na basılmaya hazır bir <img> etiketine çevirir."""
    if not veri_url:
        return ""
    return (
        f'<img src="{veri_url}" alt="İmza" '
        f'style="max-height:44px; max-width:150px; display:inline-block;">'
    )


def _montaj_formu_veri(satir):
    """`_abone_satir_sozlugu()` çıktısından Montaj Formu şablonu için birleştirme
    (mail-merge) verisi hazırlar."""
    return {
        "adi": satir.get("adi") or "",
        "soyadi": satir.get("soyadi") or "",
        "koy_adi": satir.get("koy_adi") or "",
        "sayac_no": satir.get("sayac_no") or "",
        "telefon": satir.get("telefon") or "",
        "telefon2": satir.get("telefon2") or "",
        "montaj_tarihi": (satir.get("montaj_tarihi") or "").replace(".", "/"),
        "sayac_ucreti": satir.get("sayac_tutari") or "0,00",
        "sayac_ucreti_alinan": satir.get("alinan_tutar") or "0,00",
        "tesisat_ucreti": satir.get("malzeme_tutari") or "0,00",
        "tesisat_ucreti_alinan": satir.get("malzeme_alinan") or "0,00",
        "montaj_personeli": satir.get("montaj_personeli") or "",
        "montaj_imza": _imza_veri_url_img_etiketi(satir.get("_montaj_imza_veri_url")),
        "abone_imza": _imza_veri_url_img_etiketi(satir.get("_abone_imza_veri_url")),
    }


_MONTAJ_FORMU_SANDBOX = SandboxedEnvironment()


def _montaj_formu_render_tek(sablon_icerik, satir):
    """Şablonu tek bir abone verisiyle render eder."""
    veri = _montaj_formu_veri(satir)
    try:
        return _MONTAJ_FORMU_SANDBOX.from_string(sablon_icerik).render(**veri), None
    except Exception as e:
        return None, str(e)


_MONTAJ_FORMU_TEST_VERISI = {
    "adi": "TEST", "soyadi": "TEST", "koy_adi": "TEST", "sayac_no": "0",
    "telefon": "", "telefon2": "", "montaj_tarihi": "01.01.2026",
    "sayac_tutari": "0,00", "alinan_tutar": "0,00",
    "malzeme_tutari": "0,00", "malzeme_alinan": "0,00", "montaj_personeli": "",
}


def _montaj_formu_sablonlar_listele(db):
    """Kayıtlı TÜM Montaj Formu tasarımlarının (id, ad) listesini döner."""
    cur = db.cursor()
    cur.execute("SELECT id, ad FROM montaj_formu_sablon ORDER BY id")
    satirlar = cur.fetchall()
    cur.close()
    return satirlar


def _montaj_formu_sablon_getir(db, sablon_id=None):
    """Belirli bir tasarımı ya da hiç verilmemişse kayıtlı İLK tasarımı döner."""
    cur = db.cursor()
    if sablon_id:
        cur.execute("SELECT id, ad, icerik FROM montaj_formu_sablon WHERE id = %s", (sablon_id,))
    else:
        cur.execute("SELECT id, ad, icerik FROM montaj_formu_sablon ORDER BY id LIMIT 1")
    satir = cur.fetchone()
    cur.close()
    if satir:
        return satir
    return {"id": None, "ad": "Varsayılan", "icerik": _MONTAJ_FORMU_VARSAYILAN_SABLON}


def _montaj_formu_sablon_kaydet(db, sablon_id, yeni_icerik):
    """Var olan bir Montaj Formu tasarımının içeriğini günceller."""
    cur = db.cursor()
    cur.execute(
        "UPDATE montaj_formu_sablon SET icerik = %s, guncelleme_tarihi = NOW() WHERE id = %s",
        (yeni_icerik, sablon_id),
    )
    db.commit()
    cur.close()


def _montaj_formu_sablon_olustur(db, ad, icerik):
    """Yeni, isimli bir Montaj Formu tasarımı oluşturur ve yeni kaydın id'sini döner."""
    cur = db.cursor()
    cur.execute(
        "INSERT INTO montaj_formu_sablon (ad, icerik) VALUES (%s, %s) RETURNING id",
        (ad or "Yeni Tasarım", icerik),
    )
    yeni_id = cur.fetchone()["id"]
    db.commit()
    cur.close()
    return yeni_id


def _montaj_formu_sablon_yeniden_adlandir(db, sablon_id, yeni_ad):
    cur = db.cursor()
    cur.execute("UPDATE montaj_formu_sablon SET ad = %s WHERE id = %s", (yeni_ad, sablon_id))
    db.commit()
    cur.close()


def _montaj_formu_sablon_sil(db, sablon_id):
    """Bir tasarımı siler."""
    cur = db.cursor()
    cur.execute("DELETE FROM montaj_formu_sablon WHERE id = %s", (sablon_id,))
    db.commit()
    cur.close()


def ensure_db():
    schema_yolu = os.path.join(os.path.dirname(os.path.abspath(__file__)), "schema.sql")
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    with open(schema_yolu, "r", encoding="utf-8") as f:
        cur.execute(f.read())
    conn.commit()

    admin_kullanici = os.environ.get("ADMIN_KULLANICI")
    admin_sifre = os.environ.get("ADMIN_SIFRE")
    if admin_kullanici and admin_sifre:
        cur.execute(
            "SELECT id FROM kullanici WHERE kullanici_adi = %s", (admin_kullanici,)
        )
        var_mi = cur.fetchone()
        if not var_mi:
            cur.execute(
                "INSERT INTO kullanici (kullanici_adi, sifre_hash) VALUES (%s, %s)",
                (admin_kullanici, generate_password_hash(admin_sifre)),
            )
            conn.commit()

    cur.execute("SELECT COUNT(*) FROM montaj_formu_sablon")
    if cur.fetchone()[0] == 0:
        cur.execute(
            "INSERT INTO montaj_formu_sablon (icerik) VALUES (%s)",
            (_MONTAJ_FORMU_VARSAYILAN_SABLON,),
        )
        conn.commit()

    for grup, varsayilan_liste in FORM_SECENEK_VARSAYILANLARI.items():
        cur.execute("SELECT COUNT(*) FROM form_secenegi WHERE grup = %s", (grup,))
        if cur.fetchone()[0] == 0:
            for sira, deger in enumerate(varsayilan_liste):
                cur.execute(
                    "INSERT INTO form_secenegi (grup, deger, sira) VALUES (%s, %s, %s)",
                    (grup, deger, sira),
                )
            conn.commit()

    # Gönderimler listesinin sırası artık id yerine gönderim tarihine göre
    # hesaplanıyor (bkz. _fabrika_gonderim_sira_numaralarini_yenile) — burada
    # ilk kurulumda/deploy'da mevcut kayıtlar için bir kereliğine dolduruluyor,
    # sonrasında her ekleme/silme onu güncel tutuyor.
    cur.execute(
        """
        UPDATE fabrika_gonderim fg
        SET sira_no = t.yeni_sira
        FROM (
            SELECT id, ROW_NUMBER() OVER (ORDER BY gonderim_tarihi ASC NULLS LAST, id ASC) AS yeni_sira
            FROM fabrika_gonderim
        ) t
        WHERE fg.id = t.id AND (fg.sira_no IS DISTINCT FROM t.yeni_sira)
        """
    )
    conn.commit()

    # Fabrika/Tamir listesindeki "Sıra No" da aynı mantıkla (bkz.
    # _fabrika_tamir_sira_numaralarini_yenile) kayıt oluşturuluş sırasına göre
    # 1'den başlayarak boşluksuz tutuluyor; silinmiş (silindi_mi) kayıtlar
    # sayılmıyor. Burada ilk kurulumda/deploy'da mevcut kayıtlar için bir
    # kereliğine dolduruluyor.
    cur.execute(
        """
        UPDATE fabrika_tamir ft
        SET sira_no = t.yeni_sira
        FROM (
            SELECT id, ROW_NUMBER() OVER (ORDER BY created_at ASC NULLS LAST, id ASC) AS yeni_sira
            FROM fabrika_tamir
            WHERE silindi_mi IS NOT TRUE
        ) t
        WHERE ft.id = t.id AND (ft.sira_no IS DISTINCT FROM t.yeni_sira)
        """
    )
    conn.commit()

    cur.close()
    conn.close()


ensure_db()

_DB_HAVUZU = psycopg2.pool.ThreadedConnectionPool(
    1, 8, DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor
)


def get_db():
    if "db" not in g:
        g.db = _DB_HAVUZU.getconn()
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        try:
            db.rollback()
        except Exception:
            pass
        _DB_HAVUZU.putconn(db)


def _filtre_durumu_uygula(route_adi):
    """Liste sayfalarında uygulanan filtreleri oturumda hatırlar."""
    anahtar = f"filtre_{route_adi}"
    if request.args.get("bos") == "1":
        session.pop(anahtar, None)
        return redirect(url_for(route_adi))
    qs = request.query_string.decode()
    if qs:
        session[anahtar] = qs
        return None
    kayitli = session.get(anahtar)
    if kayitli:
        return redirect(url_for(route_adi) + "?" + kayitli)
    return None


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


_GIRIS_DENEME_KILIDI = threading.Lock()
_GIRIS_BASARISIZ_DENEMELER = {}
_GIRIS_MAKS_DENEME = 5
_GIRIS_KILIT_SURESI_SN = 5 * 60


def _giris_kilitli_mi(ip):
    with _GIRIS_DENEME_KILIDI:
        kayit = _GIRIS_BASARISIZ_DENEMELER.get(ip)
        if not kayit:
            return False, 0
        sayi, son_zaman = kayit
        if sayi < _GIRIS_MAKS_DENEME:
            return False, 0
        kalan = _GIRIS_KILIT_SURESI_SN - (time.time() - son_zaman)
        if kalan <= 0:
            del _GIRIS_BASARISIZ_DENEMELER[ip]
            return False, 0
        return True, kalan


def _giris_basarisiz_kaydet(ip):
    with _GIRIS_DENEME_KILIDI:
        sayi, _ = _GIRIS_BASARISIZ_DENEMELER.get(ip, (0, 0))
        _GIRIS_BASARISIZ_DENEMELER[ip] = (sayi + 1, time.time())


def _giris_basarili_temizle(ip):
    with _GIRIS_DENEME_KILIDI:
        _GIRIS_BASARISIZ_DENEMELER.pop(ip, None)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        istemci_ip = request.remote_addr or "bilinmiyor"
        kilitli, kalan_sn = _giris_kilitli_mi(istemci_ip)
        if kilitli:
            flash(
                f"Çok fazla başarısız giriş denemesi yapıldı. Lütfen "
                f"{math.ceil(kalan_sn / 60)} dakika sonra tekrar deneyin."
            )
            return render_template("login.html")

        kullanici_adi = request.form.get("kullanici_adi", "").strip()
        sifre = request.form.get("sifre", "")

        db = get_db()
        cur = db.cursor()
        cur.execute("SELECT * FROM kullanici WHERE kullanici_adi = %s", (kullanici_adi,))
        user = cur.fetchone()
        cur.close()

        if user and check_password_hash(user["sifre_hash"], sifre):
            _giris_basarili_temizle(istemci_ip)
            session.clear()
            session["user_id"] = user["id"]
            session["kullanici_adi"] = user["kullanici_adi"]
            return redirect(url_for("abone_listesi"))

        _giris_basarisiz_kaydet(istemci_ip)
        flash("Kullanıcı adı veya şifre hatalı.")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/hesap-ayarlari", methods=["GET", "POST"])
@login_required
def hesap_ayarlari():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM kullanici WHERE id = %s", (session["user_id"],))
    kullanici = cur.fetchone()

    if request.method == "POST":
        mevcut_sifre = request.form.get("mevcut_sifre", "")
        yeni_kullanici_adi = request.form.get("yeni_kullanici_adi", "").strip()
        yeni_sifre = request.form.get("yeni_sifre", "")
        yeni_sifre_tekrar = request.form.get("yeni_sifre_tekrar", "")

        if not kullanici or not check_password_hash(kullanici["sifre_hash"], mevcut_sifre):
            flash("Mevcut şifre yanlış, değişiklik yapılmadı.")
        elif not yeni_kullanici_adi:
            flash("Kullanıcı adı boş olamaz.")
        elif yeni_sifre and yeni_sifre != yeni_sifre_tekrar:
            flash("Yeni şifre ile tekrarı birbirini tutmuyor.")
        elif yeni_sifre and len(yeni_sifre) < 8:
            flash("Yeni şifre en az 8 karakter olmalı.")
        else:
            cur.execute(
                "SELECT id FROM kullanici WHERE kullanici_adi = %s AND id != %s",
                (yeni_kullanici_adi, kullanici["id"]),
            )
            cakisma = cur.fetchone()
            if cakisma:
                flash("Bu kullanıcı adı zaten başka bir hesapta kullanılıyor.")
            else:
                if yeni_sifre:
                    cur.execute(
                        "UPDATE kullanici SET kullanici_adi = %s, sifre_hash = %s WHERE id = %s",
                        (yeni_kullanici_adi, generate_password_hash(yeni_sifre), kullanici["id"]),
                    )
                else:
                    cur.execute(
                        "UPDATE kullanici SET kullanici_adi = %s WHERE id = %s",
                        (yeni_kullanici_adi, kullanici["id"]),
                    )
                db.commit()
                session["kullanici_adi"] = yeni_kullanici_adi
                cur.close()
                flash("Hesap bilgileriniz güncellendi.")
                return redirect(url_for("hesap_ayarlari"))

    cur.close()
    return render_template(
        "hesap_ayarlari.html", kullanici=kullanici,
        ofis_enlem=_ayar_getir(db, "ofis_enlem"),
        ofis_boylam=_ayar_getir(db, "ofis_boylam"),
        yedek_alici_eposta=_ayar_getir(db, "yedek_alici_eposta"),
        fatura_satici=_hizli_satici_bilgisi(db),
        hizli_ayarli_mi=_hizli_ayarli_mi(),
        hizli_ortam=_hizli_ortam(),
    )


@app.route("/hesap-ayarlari/fatura-ayarlari", methods=["POST"])
@login_required
def hesap_ayarlari_fatura_ayarlari():
    """'Fatura Ayarları' kutusunun kaydet düğmesi."""
    db = get_db()
    for anahtar in _HIZLI_SATICI_AYAR_ANAHTARLARI:
        deger = request.form.get(anahtar, "").strip()
        if anahtar == "fatura_satici_kimlik_no":
            deger = "".join(ch for ch in deger if ch.isdigit())
        _ayar_kaydet(db, anahtar, deger)
    flash("Fatura ayarları (satıcı bilgileri) kaydedildi.")
    return redirect(url_for("hesap_ayarlari"))


@app.route("/hesap-ayarlari/fatura-baglanti-testi", methods=["POST"])
@login_required
def hesap_ayarlari_fatura_baglanti_testi():
    """'Bağlantıyı Test Et' butonu."""
    db = get_db()
    if not _hizli_ayarli_mi():
        flash("API ayarları (HIZLI_SECRET_KEY / HIZLI_API_KEY / HIZLI_KULLANICI_ADI / HIZLI_SIFRE ortam değişkenleri) tanımlanmamış.")
        return redirect(url_for("hesap_ayarlari"))
    token, hata = _hizli_token_al(db, zorla_yenile=True)
    if hata:
        flash(f"Bağlantı testi BAŞARISIZ ({_hizli_ortam()} ortamı): {hata}")
    else:
        flash(f"Bağlantı testi BAŞARILI ({_hizli_ortam()} ortamı) — giriş yapılabildi, token alındı.")
    return redirect(url_for("hesap_ayarlari"))


@app.route("/hesap-ayarlari/ofis-konumu", methods=["POST"])
@login_required
def hesap_ayarlari_ofis_konumu():
    """'Ofis Konumu' kutusunun kaydet düğmesi."""
    enlem = _konum_sayilastir(request.form.get("ofis_enlem"))
    boylam = _konum_sayilastir(request.form.get("ofis_boylam"))
    if enlem is None or boylam is None or not (-90 <= enlem <= 90) or not (-180 <= boylam <= 180):
        flash("Geçersiz koordinat — enlem -90 ile 90, boylam -180 ile 180 arasında bir sayı olmalı.")
        return redirect(url_for("hesap_ayarlari"))
    db = get_db()
    _ayar_kaydet(db, "ofis_enlem", str(enlem))
    _ayar_kaydet(db, "ofis_boylam", str(boylam))
    flash("Ofis konumu kaydedildi.")
    return redirect(url_for("hesap_ayarlari"))


@app.route("/hesap-ayarlari/yedek-eposta", methods=["POST"])
@login_required
def hesap_ayarlari_yedek_eposta():
    """'Yedek Al E-postası' kutusunun kaydet düğmesi."""
    eposta = request.form.get("yedek_alici_eposta", "").strip()
    if eposta and ("@" not in eposta or "." not in eposta.split("@")[-1]):
        flash("Geçerli bir e-posta adresi girin.")
        return redirect(url_for("hesap_ayarlari"))
    db = get_db()
    _ayar_kaydet(db, "yedek_alici_eposta", eposta)
    if eposta:
        flash("Yedek al e-postası kaydedildi — bundan sonra 'Yedek Al' basıldığında dosya bu adrese de gönderilecek.")
    else:
        flash("Yedek al e-postası kaldırıldı — yedekler artık sadece indirme olarak alınacak.")
    return redirect(url_for("hesap_ayarlari"))


@app.route("/api/ofis-konumu")
@login_required
def ofis_konumu_api():
    """abone_form.html / ariza_form.html'deki "Konum Al" butonu."""
    db = get_db()
    enlem = _ayar_getir(db, "ofis_enlem")
    boylam = _ayar_getir(db, "ofis_boylam")
    return jsonify({
        "enlem": float(enlem) if enlem else None,
        "boylam": float(boylam) if boylam else None,
    })


@app.route("/")
@login_required
def index():
    return redirect(url_for("abone_listesi"))


@app.route("/sw.js")
def service_worker():
    yanit = send_from_directory(app.static_folder, "sw.js")
    yanit.headers["Service-Worker-Allowed"] = "/"
    yanit.headers["Content-Type"] = "application/javascript"
    return yanit


def _yedek_eposta_gonder(icerik_bytes, dosya_adi, alici_eposta):
    """Yedek dosyasını SMTP üzerinden alıcıya e-posta eki olarak gönderir."""
    smtp_sunucu = os.environ.get("SMTP_HOST")
    smtp_kullanici = os.environ.get("SMTP_KULLANICI")
    smtp_sifre = os.environ.get("SMTP_SIFRE")
    smtp_gonderen = os.environ.get("SMTP_GONDEREN") or smtp_kullanici

    if not smtp_sunucu or not smtp_kullanici or not smtp_sifre:
        return False, (
            "SMTP ayarları (SMTP_HOST / SMTP_KULLANICI / SMTP_SIFRE ortam "
            "değişkenleri) henüz tanımlanmamış, bu yüzden yedek e-posta ile "
            "gönderilemedi."
        )

    try:
        smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    except (TypeError, ValueError):
        smtp_port = 587

    mesaj = EmailMessage()
    mesaj["Subject"] = f"ALGI BİLİŞİM Yedek - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    mesaj["From"] = smtp_gonderen
    mesaj["To"] = alici_eposta
    mesaj.set_content(
        "ALGI BİLİŞİM sisteminin veritabanı yedeği ektedir.\n\n"
        "Bu e-posta, sistemdeki 'Yedek Al' butonuna basıldığında otomatik "
        "olarak gönderilmiştir."
    )
    mesaj.add_attachment(icerik_bytes, maintype="text", subtype="plain", filename=dosya_adi)

    try:
        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_sunucu, smtp_port, timeout=20) as sunucu:
                sunucu.login(smtp_kullanici, smtp_sifre)
                sunucu.send_message(mesaj)
        else:
            with smtplib.SMTP(smtp_sunucu, smtp_port, timeout=20) as sunucu:
                sunucu.ehlo()
                sunucu.starttls()
                sunucu.ehlo()
                sunucu.login(smtp_kullanici, smtp_sifre)
                sunucu.send_message(mesaj)
        return True, None
    except Exception as e:
        return False, str(e)


_HIZLI_TABAN_URL = {
    "test": "https://econnecttest.hizliteknoloji.com.tr",
    "prod": "https://econnect.hizliteknoloji.com.tr",
}

_HIZLI_APP_TYPE = {"efatura": 1, "earsiv": 2}
_HIZLI_PROFILE_ID = {"efatura": "TICARIFATURA", "earsiv": "EARSIVFATURA"}
_HIZLI_INVOICE_TYPE_CODE = "SATIS"
_HIZLI_KDV_TAX_CODE = "0015"
_HIZLI_PARA_BIRIMI = "TRY"


def _hizli_ortam():
    return "prod" if os.environ.get("HIZLI_ORTAM", "test").strip().lower() == "prod" else "test"


def _hizli_taban_url():
    return _HIZLI_TABAN_URL[_hizli_ortam()]


def _hizli_ayarli_mi():
    return all([
        os.environ.get("HIZLI_SECRET_KEY"),
        os.environ.get("HIZLI_API_KEY"),
        os.environ.get("HIZLI_KULLANICI_ADI"),
        os.environ.get("HIZLI_SIFRE"),
    ])


def _hizli_istek(yol, govde=None, token=None, metod="POST", zaman_asimi=20):
    """Hızlı Bilişim e-Connect API'sine JSON istek gönderir."""
    url = _hizli_taban_url() + yol
    veri_bytes = json.dumps(govde).encode("utf-8") if govde is not None else None
    istek = urllib.request.Request(url, data=veri_bytes, method=metod)
    istek.add_header("Content-Type", "application/json")
    istek.add_header("Accept", "application/json")
    if token:
        istek.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(istek, timeout=zaman_asimi) as yanit:
            gövde_metni = yanit.read().decode("utf-8")
            return True, (json.loads(gövde_metni) if gövde_metni else {})
    except urllib.error.HTTPError as e:
        hata_govdesi = ""
        try:
            hata_govdesi = e.read().decode("utf-8")
        except Exception:
            pass
        return False, f"HTTP {e.code}: {hata_govdesi or e.reason}"
    except urllib.error.URLError as e:
        return False, f"Bağlantı hatası: {e.reason}"
    except Exception as e:
        return False, str(e)


def _hizli_sifreli_kimlik(db):
    """UtilEncrypt sonucunu önbellekten döner; yoksa Hızlı Bilişim'den bir kez alıp saklar."""
    onbellek_kullanici = _ayar_getir(db, "hizli_sifreli_kullanici")
    onbellek_sifre = _ayar_getir(db, "hizli_sifreli_sifre")
    if onbellek_kullanici and onbellek_sifre:
        return (onbellek_kullanici, onbellek_sifre), None

    basarili, veri = _hizli_istek(
        "/HizliApi/RestApi/UtilEncrypt",
        {
            "secretKey": os.environ.get("HIZLI_SECRET_KEY", ""),
            "username": os.environ.get("HIZLI_KULLANICI_ADI", ""),
            "password": os.environ.get("HIZLI_SIFRE", ""),
        },
    )
    if not basarili:
        return None, f"Kullanıcı bilgileri şifrelenemedi: {veri}"
    veri_obj = veri[0] if isinstance(veri, list) and veri else veri
    sifreli_kullanici = veri_obj.get("username") if isinstance(veri_obj, dict) else None
    sifreli_sifre = veri_obj.get("password") if isinstance(veri_obj, dict) else None
    if not sifreli_kullanici or not sifreli_sifre:
        return None, f"Şifreleme yanıtı beklenmedik biçimde geldi: {veri}"
    _ayar_kaydet(db, "hizli_sifreli_kullanici", sifreli_kullanici)
    _ayar_kaydet(db, "hizli_sifreli_sifre", sifreli_sifre)
    return (sifreli_kullanici, sifreli_sifre), None


def _hizli_token_al(db, zorla_yenile=False):
    """Login token'ını önbellekten döner; yoksa ya da süresi dolmuşsa yeniden login olur."""
    if not zorla_yenile:
        onbellek_token = _ayar_getir(db, "hizli_token")
        onbellek_zaman = _ayar_getir(db, "hizli_token_zaman")
        if onbellek_token and onbellek_zaman:
            try:
                alinma_zamani = datetime.fromisoformat(onbellek_zaman)
                if (datetime.now() - alinma_zamani).total_seconds() < 20 * 3600:
                    return onbellek_token, None
            except ValueError:
                pass

    kimlik, hata = _hizli_sifreli_kimlik(db)
    if hata:
        return None, hata
    sifreli_kullanici, sifreli_sifre = kimlik

    basarili, veri = _hizli_istek(
        "/HizliApi/RestApi/Login",
        {
            "apiKey": os.environ.get("HIZLI_API_KEY", ""),
            "username": sifreli_kullanici,
            "password": sifreli_sifre,
        },
    )
    if not basarili:
        return None, f"Hızlı Bilişim'e giriş yapılamadı: {veri}"
    veri_obj = veri[0] if isinstance(veri, list) and veri else veri
    token = None
    if isinstance(veri_obj, dict):
        if veri_obj.get("IsSucceeded") is False:
            return None, f"Hızlı Bilişim'e giriş başarısız: {veri_obj.get('Message') or veri_obj}"
        token = veri_obj.get("Token") or veri_obj.get("token") or veri_obj.get("accessToken")
    elif isinstance(veri_obj, str):
        token = veri_obj
    if not token:
        return None, f"Giriş yanıtında bir token bulunamadı: {veri}"
    _ayar_kaydet(db, "hizli_token", token)
    _ayar_kaydet(db, "hizli_token_zaman", datetime.now().isoformat())
    return token, None


def _hizli_kdv_ayir(tutar_kdv_dahil, kdv_orani=20):
    """KDV dahil bir tutardan KDV hariç taban ve KDV tutarını ayırır."""
    tutar_kdv_dahil = float(tutar_kdv_dahil or 0)
    kdv_haric = round(tutar_kdv_dahil / (1 + kdv_orani / 100), 2)
    kdv_tutari = round(tutar_kdv_dahil - kdv_haric, 2)
    return kdv_haric, kdv_tutari


def _fatura_turu_belirle(kimlik_no):
    """kimlik_no'nun uzunluğuna göre fatura türünü ÖNERİR."""
    rakamlar = "".join(ch for ch in str(kimlik_no or "") if ch.isdigit())
    if len(rakamlar) == 11:
        return "earsiv"
    if len(rakamlar) == 10:
        return "efatura"
    return None


_HIZLI_SATICI_AYAR_ANAHTARLARI = [
    "fatura_satici_kimlik_no", "fatura_satici_vergi_dairesi", "fatura_satici_unvan",
    "fatura_satici_ad", "fatura_satici_soyad", "fatura_satici_adres",
    "fatura_satici_eposta", "fatura_satici_telefon",
]


def _hizli_satici_bilgisi(db):
    return {anahtar: (_ayar_getir(db, anahtar) or "") for anahtar in _HIZLI_SATICI_AYAR_ANAHTARLARI}


def _hizli_satici_eksik_mi(satici):
    zorunlu = ["fatura_satici_kimlik_no", "fatura_satici_unvan",
               "fatura_satici_vergi_dairesi", "fatura_satici_adres"]
    return any(not satici.get(a) for a in zorunlu)


def _kalem_hesapla(kalem):
    """Bir fatura kalemi için ara/iskonto/KDV/toplam tutarlarını hesaplar."""
    miktar = float(kalem.get("miktar") or 0)
    birim_fiyat = float(kalem.get("birim_fiyat") or 0)
    iskonto_orani = float(kalem.get("iskonto_orani") or 0)
    kdv_orani = float(kalem.get("kdv_orani") or 0)
    diger_vergi = float(kalem.get("diger_vergi") or 0)

    ara_tutar = round(miktar * birim_fiyat, 2)
    iskonto_tutari = round(ara_tutar * iskonto_orani / 100, 2)
    mal_hizmet_tutari = round(ara_tutar - iskonto_tutari, 2)
    kdv_tutari = round(mal_hizmet_tutari * kdv_orani / 100, 2)
    satir_toplam = round(mal_hizmet_tutari + kdv_tutari + diger_vergi, 2)

    sonuc = dict(kalem)
    sonuc.update({
        "miktar": miktar, "birim_fiyat": birim_fiyat, "iskonto_orani": iskonto_orani,
        "kdv_orani": kdv_orani, "diger_vergi": diger_vergi,
        "ara_tutar": ara_tutar, "iskonto_tutari": iskonto_tutari,
        "mal_hizmet_tutari": mal_hizmet_tutari, "kdv_tutari": kdv_tutari,
        "satir_toplam": satir_toplam,
    })
    return sonuc


def _fatura_kalemlerini_formdan_oku(form):
    """Fatura Kes ekranındaki kalem tablosunun POST verisinden hesaplanmış kalem sözlükleri listesi üretir."""
    aciklamalar = form.getlist("kalem_aciklama[]")
    miktarlar = form.getlist("kalem_miktar[]")
    birim_fiyatlar = form.getlist("kalem_birim_fiyat[]")
    iskonto_oranlari = form.getlist("kalem_iskonto_orani[]")
    iskonto_nedenleri = form.getlist("kalem_iskonto_nedeni[]")
    kdv_oranlari = form.getlist("kalem_kdv_orani[]")
    diger_vergiler = form.getlist("kalem_diger_vergi[]")

    def _sayi(liste, i, varsayilan=0.0):
        if i >= len(liste):
            return varsayilan
        try:
            return float(str(liste[i]).replace(",", "."))
        except (TypeError, ValueError):
            return varsayilan

    kalemler = []
    for i in range(len(aciklamalar)):
        aciklama = (aciklamalar[i] or "").strip()
        if not aciklama:
            continue
        miktar = _sayi(miktarlar, i, 0)
        birim_fiyat = _sayi(birim_fiyatlar, i, 0)
        if miktar <= 0 or birim_fiyat < 0:
            continue
        ham = {
            "aciklama": aciklama,
            "miktar": miktar,
            "birim_fiyat": birim_fiyat,
            "iskonto_orani": _sayi(iskonto_oranlari, i, 0),
            "iskonto_nedeni": (iskonto_nedenleri[i].strip() if i < len(iskonto_nedenleri) else ""),
            "kdv_orani": _sayi(kdv_oranlari, i, 20),
            "diger_vergi": _sayi(diger_vergiler, i, 0),
        }
        kalemler.append(_kalem_hesapla(ham))
    return kalemler


def _hizli_invoice_model_olustur(satici, alici, kalemler, fatura_turu, fatura_tarihi=None):
    """SendInvoiceModel'in beklediği InvoiceModel nesnesini oluşturur."""
    simdi = datetime.now()
    if fatura_tarihi:
        if isinstance(fatura_tarihi, str):
            tarih_str = fatura_tarihi
        else:
            tarih_str = fatura_tarihi.strftime("%Y-%m-%d")
    else:
        tarih_str = simdi.strftime("%Y-%m-%d")
    invoice_lines = []
    toplam_kdv_haric = 0.0
    toplam_kdv = 0.0
    toplam_diger_vergi = 0.0
    for i, kalem in enumerate(kalemler, start=1):
        satir = {
            "ID": i,
            "Item_Name": kalem["aciklama"],
            "Quantity_Amount": kalem["miktar"],
            "Quantity_Unit_User": "ADET",
            "Price_Amount": kalem["birim_fiyat"],
            "Price_Total": kalem["mal_hizmet_tutari"],
            "lineTaxes": [{
                "Tax_Code": _HIZLI_KDV_TAX_CODE,
                "Tax_Name": "KDV",
                "Tax_Base": kalem["mal_hizmet_tutari"],
                "Tax_Perc": kalem["kdv_orani"],
                "Tax_Amnt": kalem["kdv_tutari"],
            }],
        }
        if kalem.get("iskonto_tutari"):
            satir["Discount_Percent"] = kalem["iskonto_orani"]
            satir["Discount_Amount"] = kalem["iskonto_tutari"]
            if kalem.get("iskonto_nedeni"):
                satir["Discount_Reason"] = kalem["iskonto_nedeni"]
        invoice_lines.append(satir)
        toplam_kdv_haric += kalem["mal_hizmet_tutari"]
        toplam_kdv += kalem["kdv_tutari"]
        toplam_diger_vergi += kalem.get("diger_vergi") or 0

    toplam_kdv_haric = round(toplam_kdv_haric, 2)
    toplam_kdv = round(toplam_kdv, 2)
    toplam_diger_vergi = round(toplam_diger_vergi, 2)
    toplam_kdv_dahil = round(toplam_kdv_haric + toplam_kdv + toplam_diger_vergi, 2)

    alici_kimlik_no = "".join(ch for ch in str(alici.get("kimlik_no") or "") if ch.isdigit())
    customer = {
        "IdentificationID": alici_kimlik_no,
        "PartyName": f"{alici.get('adi', '')} {alici.get('soyadi', '')}".strip(),
        "Person_FirstName": alici.get("adi", ""),
        "Person_FamilyName": alici.get("soyadi", ""),
        "StreetName": alici.get("adres", ""),
        "CityName": alici.get("koy_adi", ""),
        "CountryName": "Türkiye",
        "ElectronicMail": alici.get("eposta", ""),
        "Telephone": alici.get("telefon", ""),
    }
    if alici.get("vergi_dairesi"):
        customer["TaxSchemeName"] = alici["vergi_dairesi"]

    satici_ad_soyad = f"{satici.get('fatura_satici_ad', '')} {satici.get('fatura_satici_soyad', '')}".strip()
    supplier = {
        "IdentificationID": "".join(ch for ch in str(satici.get("fatura_satici_kimlik_no") or "") if ch.isdigit()),
        "PartyName": satici.get("fatura_satici_unvan", ""),
        "Person_FirstName": satici.get("fatura_satici_ad", ""),
        "Person_FamilyName": satici.get("fatura_satici_soyad", ""),
        "TaxSchemeName": satici.get("fatura_satici_vergi_dairesi", ""),
        "StreetName": satici.get("fatura_satici_adres", ""),
        "CountryName": "Türkiye",
        "ElectronicMail": satici.get("fatura_satici_eposta", ""),
        "Telephone": satici.get("fatura_satici_telefon", ""),
    }

    invoice_model = {
        "invoiceheader": {
            "ProfileID": _HIZLI_PROFILE_ID[fatura_turu],
            "InvoiceTypeCode": _HIZLI_INVOICE_TYPE_CODE,
            "IssueDate": tarih_str,
            "IssueTime": simdi.strftime("%H:%M:%S"),
            "DocumentCurrencyCode": _HIZLI_PARA_BIRIMI,
            "LineExtensionAmount": toplam_kdv_haric,
            "TaxInclusiveAmount": toplam_kdv_dahil,
            "PayableAmount": toplam_kdv_dahil,
        },
        "customer": customer,
        "supplier": supplier,
        "invoiceLines": invoice_lines,
    }
    return invoice_model, toplam_kdv_dahil, toplam_kdv_haric, toplam_kdv


def _hizli_fatura_gonder(db, kaynak_tur, kaynak_id, alici, kalemler, fatura_turu, olusturan_kullanici, fatura_tarihi=None):
    """Bir abone/arıza kaydı için e-Fatura/e-Arşiv Fatura keser."""
    satici = _hizli_satici_bilgisi(db)
    kalem_ozet = ", ".join(f"{k['aciklama']}: {tl_format(k['satir_toplam'])} TL" for k in kalemler)
    yerel_id = f"{kaynak_tur}-{kaynak_id}-{secrets.token_hex(4)}"
    fatura_tarihi = fatura_tarihi or datetime.now().strftime("%Y-%m-%d")

    cur = db.cursor()

    if not _hizli_ayarli_mi():
        cur.execute(
            "INSERT INTO fatura (kaynak_tur, kaynak_id, fatura_turu, yerel_id, durum, hata_mesaji, kalemler, olusturan_kullanici, fatura_tarihi) "
            "VALUES (%s, %s, %s, %s, 'hata', %s, %s, %s, %s) RETURNING id",
            (kaynak_tur, kaynak_id, fatura_turu, yerel_id,
             "Hızlı Bilişim API ayarları (ortam değişkenleri) henüz tanımlanmamış.",
             kalem_ozet, olusturan_kullanici, fatura_tarihi),
        )
        fatura_id = cur.fetchone()["id"]
        db.commit()
        cur.close()
        return fatura_id

    invoice_model, toplam_dahil, toplam_haric, toplam_kdv = _hizli_invoice_model_olustur(
        satici, alici, kalemler, fatura_turu, fatura_tarihi
    )

    cur.execute(
        "INSERT INTO fatura (kaynak_tur, kaynak_id, fatura_turu, yerel_id, durum, "
        "tutar_kdv_dahil, tutar_kdv_haric, kdv_tutari, kalemler, olusturan_kullanici, fatura_tarihi) "
        "VALUES (%s, %s, %s, %s, 'beklemede', %s, %s, %s, %s, %s, %s) RETURNING id",
        (kaynak_tur, kaynak_id, fatura_turu, yerel_id,
         toplam_dahil, toplam_haric, toplam_kdv, kalem_ozet, olusturan_kullanici, fatura_tarihi),
    )
    fatura_id = cur.fetchone()["id"]
    db.commit()

    token, hata = _hizli_token_al(db)
    if hata:
        cur.execute("UPDATE fatura SET durum = 'hata', hata_mesaji = %s WHERE id = %s", (hata, fatura_id))
        db.commit()
        cur.close()
        return fatura_id

    govde = {
        "inputDocument": [{
            "AppType": _HIZLI_APP_TYPE[fatura_turu],
            "SourceUrn": satici.get("fatura_satici_kimlik_no", ""),
            "DestinationIdentifier": "".join(ch for ch in str(alici.get("kimlik_no") or "") if ch.isdigit()),
            "DestinationUrn": "".join(ch for ch in str(alici.get("kimlik_no") or "") if ch.isdigit()),
            "InvoiceModel": invoice_model,
            "LocalId": yerel_id,
            "IsDraft": False,
        }]
    }
    basarili, veri = _hizli_istek("/HizliApi/RestApi/SendInvoiceModel", govde, token=token)

    if not basarili:
        cur.execute("UPDATE fatura SET durum = 'hata', hata_mesaji = %s WHERE id = %s", (veri, fatura_id))
        db.commit()
        cur.close()
        return fatura_id

    fatura_uuid = None
    if isinstance(veri, dict):
        fatura_uuid = veri.get("DocumentUUID") or veri.get("uuid") or veri.get("Uuid")
    elif isinstance(veri, list) and veri:
        ilk = veri[0]
        if isinstance(ilk, dict):
            fatura_uuid = ilk.get("DocumentUUID") or ilk.get("uuid") or ilk.get("Uuid")

    cur.execute(
        "UPDATE fatura SET durum = 'basarili', fatura_uuid = %s WHERE id = %s",
        (fatura_uuid, fatura_id),
    )
    db.commit()
    cur.close()
    return fatura_id


def _hizli_fatura_pdf_al(db, fatura_id):
    """Daha önce başarıyla kesilmiş bir faturanın PDF içeriğini çekip 'fatura' tablosuna kaydeder."""
    cur = db.cursor()
    cur.execute("SELECT * FROM fatura WHERE id = %s", (fatura_id,))
    kayit = cur.fetchone()
    if not kayit or kayit["durum"] != "basarili" or not kayit["fatura_uuid"]:
        cur.close()
        return False, "Bu fatura henüz başarıyla kesilmemiş, PDF alınamaz."

    token, hata = _hizli_token_al(db)
    if hata:
        cur.close()
        return False, hata

    yol = (
        f"/HizliApi/RestApi/GetDocumentFile"
        f"?AppType={_HIZLI_APP_TYPE[kayit['fatura_turu']]}&Uuid={kayit['fatura_uuid']}&Tur=pdf&IsDraft=false"
    )
    basarili, veri = _hizli_istek(yol, govde=None, token=token, metod="GET")
    if not basarili:
        cur.close()
        return False, veri

    veri_obj = veri[0] if isinstance(veri, list) and veri else veri
    pdf_base64 = None
    if isinstance(veri_obj, dict):
        pdf_base64 = veri_obj.get("FileContent") or veri_obj.get("Content") or veri_obj.get("data")
    elif isinstance(veri_obj, str):
        pdf_base64 = veri_obj
    if not pdf_base64:
        cur.close()
        return False, f"PDF yanıtında dosya içeriği bulunamadı: {veri}"

    try:
        pdf_bytes = base64.b64decode(pdf_base64)
    except Exception as e:
        cur.close()
        return False, f"PDF içeriği çözümlenemedi: {e}"

    cur.execute("UPDATE fatura SET pdf_icerik = %s WHERE id = %s", (psycopg2.Binary(pdf_bytes), fatura_id))
    db.commit()
    cur.close()
    return True, None


def _netgsm_ayarli_mi():
    return all([os.environ.get("NETGSM_USERCODE"), os.environ.get("NETGSM_PASSWORD")])


def _telefon_e164_dene(telefon):
    """Bir Türkiye telefon numarasını E.164 formatına çevirmeyi dener."""
    if not telefon:
        return None
    rakamlar = "".join(ch for ch in str(telefon) if ch.isdigit())
    if len(rakamlar) == 10 and rakamlar.startswith("5"):
        return "90" + rakamlar
    if len(rakamlar) == 11 and rakamlar.startswith("05"):
        return "90" + rakamlar[1:]
    if len(rakamlar) == 12 and rakamlar.startswith("90"):
        return rakamlar
    if len(rakamlar) == 13 and rakamlar.startswith("090"):
        return "90" + rakamlar[3:]
    if len(rakamlar) == 14 and rakamlar.startswith("0090"):
        return rakamlar[2:]
    return None


def _netgsm_whatsapp_gonder(telefon, icerik):
    """Netgsm üzerinden WhatsApp mesajı gönderir. HENÜZ GERÇEK BİR API ÇAĞRISI YAPMIYOR."""
    if not _netgsm_ayarli_mi():
        return False, (
            "Netgsm API ayarları (NETGSM_USERCODE / NETGSM_PASSWORD ortam "
            "değişkenleri) henüz tanımlanmamış."
        )
    e164 = _telefon_e164_dene(telefon)
    if not e164:
        return False, f"Telefon numarası ({telefon or 'boş'}) WhatsApp gönderimi için geçerli bir formata çevrilemedi."
    return False, (
        "Netgsm WhatsApp API entegrasyonu henüz tamamlanmadı (API dokümanı "
        "bekleniyor). Alt yapı hazır, sadece gönderim çağrısı eksik."
    )


_NETGSM_SMS_HATA_KODLARI = {
    "20": "Mesaj metninde bir sorun var ya da standart maksimum karakter sayısı aşıldı.",
    "30": "Kullanıcı adı/şifre hatalı ya da API erişim izni yok (IP kısıtlaması olabilir).",
    "40": "Mesaj başlığı (gönderici adı) Netgsm panelinde tanımlı/onaylı değil.",
    "50": "Bu abonelik hesabıyla İYS kontrollü gönderim yapılamıyor.",
    "51": "Aboneliğiniz için İYS Marka bilgisi bulunamadı.",
    "70": "İstek geçersiz — parametrelerden biri hatalı ya da eksik.",
    "80": "Gönderim limiti aşıldı.",
    "85": "Aynı numaraya 1 dakika içinde 20'den fazla gönderim isteği oluşturulamaz.",
}


def _netgsm_sms_ayarli_mi():
    return _netgsm_ayarli_mi() and bool(os.environ.get("NETGSM_MSGHEADER"))


def _telefon_yerel_format_dene(telefon):
    """Bir Türkiye telefon numarasını Netgsm SMS API'sinin beklediği 10 haneli yerel formata çevirmeyi dener."""
    e164 = _telefon_e164_dene(telefon)
    if not e164:
        return None
    return e164[2:]


def _netgsm_sms_gonder(telefon, icerik):
    """Netgsm SMS REST API'si üzerinden SMS gönderir."""
    if not _netgsm_ayarli_mi():
        return False, (
            "Netgsm API ayarları (NETGSM_USERCODE / NETGSM_PASSWORD ortam "
            "değişkenleri) henüz tanımlanmamış."
        )
    msgheader = os.environ.get("NETGSM_MSGHEADER")
    if not msgheader:
        return False, (
            "Netgsm SMS mesaj başlığı (NETGSM_MSGHEADER ortam değişkeni) "
            "henüz tanımlanmamış — Netgsm panelinden onaylı bir başlık "
            "alıp bu isimle tanımlamanız gerekiyor."
        )
    yerel_no = _telefon_yerel_format_dene(telefon)
    if not yerel_no:
        return False, f"Telefon numarası ({telefon or 'boş'}) SMS gönderimi için geçerli bir formata çevrilemedi."

    govde = {
        "msgheader": msgheader,
        "messages": [{"msg": icerik, "no": yerel_no}],
        "encoding": "TR",
    }
    veri_bytes = json.dumps(govde).encode("utf-8")
    kimlik_str = f"{os.environ.get('NETGSM_USERCODE')}:{os.environ.get('NETGSM_PASSWORD')}"
    kimlik_b64 = base64.b64encode(kimlik_str.encode("ascii")).decode("ascii")
    istek = urllib.request.Request(
        "https://api.netgsm.com.tr/sms/rest/v2/send", data=veri_bytes, method="POST"
    )
    istek.add_header("Content-Type", "application/json")
    istek.add_header("Accept", "application/json")
    istek.add_header("Authorization", f"Basic {kimlik_b64}")
    try:
        with urllib.request.urlopen(istek, timeout=20) as yanit:
            yanit.read()
        return True, None
    except urllib.error.HTTPError as e:
        kod, mesaj = None, None
        try:
            hata_json = json.loads(e.read().decode("utf-8"))
            kod = hata_json.get("code")
            mesaj = hata_json.get("message")
        except Exception:
            pass
        aciklama = _NETGSM_SMS_HATA_KODLARI.get(kod)
        if aciklama:
            return False, f"Netgsm hatası ({kod}): {aciklama}"
        return False, f"Netgsm hatası: HTTP {e.code}" + (f" - {mesaj}" if mesaj else "")
    except urllib.error.URLError as e:
        return False, f"Netgsm'e bağlanılamadı: {e.reason}"
    except Exception as e:
        return False, f"SMS gönderiminde beklenmeyen hata: {e}"


def _eposta_ayarli_mi():
    return all([
        os.environ.get("SMTP_HOST"),
        os.environ.get("SMTP_KULLANICI"),
        os.environ.get("SMTP_SIFRE"),
    ])


def _mesaj_eposta_gonder(alici_eposta, icerik):
    """Aboneye SMTP üzerinden düz metin e-posta gönderir."""
    if not alici_eposta or "@" not in alici_eposta:
        return False, f"Geçerli bir e-posta adresi yok ({alici_eposta or 'boş'})."
    if not _eposta_ayarli_mi():
        return False, (
            "SMTP ayarları (SMTP_HOST / SMTP_KULLANICI / SMTP_SIFRE ortam "
            "değişkenleri) henüz tanımlanmamış, bu yüzden e-posta gönderilemedi."
        )
    smtp_sunucu = os.environ.get("SMTP_HOST")
    smtp_kullanici = os.environ.get("SMTP_KULLANICI")
    smtp_sifre = os.environ.get("SMTP_SIFRE")
    smtp_gonderen = os.environ.get("SMTP_GONDEREN") or smtp_kullanici
    try:
        smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    except (TypeError, ValueError):
        smtp_port = 587

    mesaj = EmailMessage()
    mesaj["Subject"] = "ALGI BİLİŞİM"
    mesaj["From"] = smtp_gonderen
    mesaj["To"] = alici_eposta
    mesaj.set_content(icerik)

    try:
        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_sunucu, smtp_port, timeout=20) as sunucu:
                sunucu.login(smtp_kullanici, smtp_sifre)
                sunucu.send_message(mesaj)
        else:
            with smtplib.SMTP(smtp_sunucu, smtp_port, timeout=20) as sunucu:
                sunucu.ehlo()
                sunucu.starttls()
                sunucu.ehlo()
                sunucu.login(smtp_kullanici, smtp_sifre)
                sunucu.send_message(mesaj)
        return True, None
    except Exception as e:
        return False, str(e)


def _mesaj_gonder(db, kaynak_tur, kaynak_id, alici_adi, alici_telefon, kanal, icerik, olusturan_kullanici, alici_eposta=None):
    """Tek bir alıcıya, seçilen kanaldan mesaj gönderir; sonucu 'mesaj' tablosuna kaydeder."""
    cur = db.cursor()
    cur.execute(
        "INSERT INTO mesaj (kaynak_tur, kaynak_id, kanal, alici_adi, alici_telefon, alici_eposta, icerik, durum, olusturan_kullanici) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, 'beklemede', %s) RETURNING id",
        (kaynak_tur, kaynak_id, kanal, alici_adi, alici_telefon, alici_eposta, icerik, olusturan_kullanici),
    )
    mesaj_id = cur.fetchone()["id"]
    db.commit()

    if kanal == "whatsapp":
        basarili, hata = _netgsm_whatsapp_gonder(alici_telefon, icerik)
    elif kanal == "sms":
        basarili, hata = _netgsm_sms_gonder(alici_telefon, icerik)
    elif kanal == "eposta":
        basarili, hata = _mesaj_eposta_gonder(alici_eposta, icerik)
    else:
        basarili, hata = False, f"'{kanal}' kanalı henüz desteklenmiyor."

    if basarili:
        cur.execute("UPDATE mesaj SET durum = 'basarili' WHERE id = %s", (mesaj_id,))
    else:
        cur.execute("UPDATE mesaj SET durum = 'hata', hata_mesaji = %s WHERE id = %s", (hata, mesaj_id))
    db.commit()
    cur.close()
    return mesaj_id


@app.route("/mesajlar")
@login_required
def mesaj_listesi():
    """'Mesajlarım' sayfası."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM mesaj ORDER BY created_at DESC LIMIT 500")
    mesajlar = cur.fetchall()
    cur.close()
    return render_template("mesaj_listesi.html", mesajlar=mesajlar)


@app.route("/abone/<int:abone_id>/mesaj-gonder", methods=["GET", "POST"])
@login_required
def abone_mesaj_gonder(abone_id):
    """Tek bir aboneye WhatsApp/SMS/E-posta mesajı gönderir."""
    db = get_db()
    geri = request.args.get("geri", "") or request.form.get("geri", "")
    cur = db.cursor()
    cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
    abone = cur.fetchone()
    cur.close()
    if abone is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("abone_listesi"))

    if request.method == "POST":
        kanal = request.form.get("kanal", "whatsapp")
        icerik = request.form.get("icerik", "").strip()
        if not icerik:
            flash("Mesaj içeriği boş olamaz.")
            return redirect(url_for("abone_mesaj_gonder", abone_id=abone_id, geri=geri))
        if kanal == "eposta":
            eposta = abone["eposta"]
            if not eposta:
                flash("Bu abonenin kayıtlı bir e-posta adresi yok, mesaj gönderilemedi.")
                return redirect(url_for("abone_duzenle", abone_id=abone_id))
            _mesaj_gonder(
                db, "abone", abone_id, f"{abone['adi']} {abone['soyadi']}", None, kanal, icerik,
                session.get("kullanici_adi", ""), alici_eposta=eposta,
            )
        else:
            telefon = abone["telefon"] or abone["telefon2"]
            if not telefon:
                flash("Bu abonenin kayıtlı bir telefon numarası yok, mesaj gönderilemedi.")
                return redirect(url_for("abone_duzenle", abone_id=abone_id))
            _mesaj_gonder(
                db, "abone", abone_id, f"{abone['adi']} {abone['soyadi']}", telefon, kanal, icerik,
                session.get("kullanici_adi", ""),
            )
        flash("Mesaj gönderim isteği oluşturuldu, sonucunu 'Mesajlarım' sayfasından görebilirsiniz.")
        return redirect(url_for("mesaj_listesi"))

    hedefler = [{
        "id": abone["id"], "ad": f"{abone['adi']} {abone['soyadi']}",
        "telefon": abone["telefon"] or abone["telefon2"] or "",
        "eposta": abone["eposta"] or "",
    }]
    return render_template(
        "mesaj_gonder.html", hedefler=hedefler, tekli=True,
        baslik=f"Mesaj Gönder - {abone['adi']} {abone['soyadi']}",
        gonder_url=url_for("abone_mesaj_gonder", abone_id=abone_id),
        geri_url=url_for("abone_duzenle", abone_id=abone_id) + (f"?geri={_url_quote(geri, safe='')}" if geri else ""),
        liste_url=url_for("abone_listesi"), netgsm_ayarli_mi=_netgsm_ayarli_mi(),
        sms_hazir=_netgsm_sms_ayarli_mi(), eposta_hazir=_eposta_ayarli_mi(), geri=geri,
    )


@app.route("/abone/toplu-mesaj", methods=["GET", "POST"])
@login_required
def abone_toplu_mesaj():
    """Abone Listesi'nde o an görülen filtrelenmiş kayıt kümesinin tamamına mesaj gönderir."""
    db = get_db()
    sorgu_dizesi = request.form.get("sorgu", "") if request.method == "POST" else request.query_string.decode()

    if request.method == "POST":
        kanal = request.form.get("kanal", "whatsapp")
        icerik = request.form.get("icerik", "").strip()
        abone_idler = [int(x) for x in request.form.getlist("abone_id[]") if x.strip().isdigit()]
        if not icerik:
            flash("Mesaj içeriği boş olamaz.")
            return redirect(url_for("abone_toplu_mesaj") + (f"?{sorgu_dizesi}" if sorgu_dizesi else ""))
        if not abone_idler:
            flash("Gönderilecek alıcı bulunamadı.")
            return redirect(url_for("abone_listesi"))

        cur = db.cursor()
        gonderilen = 0
        telefonsuz = 0
        epostasiz = 0
        for abone_id in abone_idler:
            cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
            abone = cur.fetchone()
            if not abone:
                continue
            if kanal == "eposta":
                eposta = abone["eposta"]
                if not eposta:
                    epostasiz += 1
                    continue
                _mesaj_gonder(
                    db, "abone", abone_id, f"{abone['adi']} {abone['soyadi']}", None, kanal, icerik,
                    session.get("kullanici_adi", ""), alici_eposta=eposta,
                )
            else:
                telefon = abone["telefon"] or abone["telefon2"]
                if not telefon:
                    telefonsuz += 1
                    continue
                _mesaj_gonder(
                    db, "abone", abone_id, f"{abone['adi']} {abone['soyadi']}", telefon, kanal, icerik,
                    session.get("kullanici_adi", ""),
                )
            gonderilen += 1
        cur.close()

        ozet = f"{gonderilen} alıcıya mesaj gönderim isteği oluşturuldu."
        if telefonsuz:
            ozet += f" {telefonsuz} kaydın telefon numarası olmadığı için atlandı."
        if epostasiz:
            ozet += f" {epostasiz} kaydın e-posta adresi olmadığı için atlandı."
        ozet += " Sonuçlarını 'Mesajlarım' sayfasından görebilirsiniz."
        flash(ozet)
        return redirect(url_for("mesaj_listesi"))

    kayitlar_ham = _abone_filtreli_kayitlari_getir(db)
    if not kayitlar_ham:
        flash("Filtreye uyan abone bulunamadı.")
        return redirect(url_for("abone_listesi"))

    hedefler = [{
        "id": k["id"], "ad": f"{k['adi']} {k['soyadi']}", "telefon": k["telefon"] or k["telefon2"] or "",
        "eposta": k["eposta"] or "",
    } for k in kayitlar_ham]

    return render_template(
        "mesaj_gonder.html", hedefler=hedefler, tekli=False,
        baslik=f"Toplu Mesaj Gönder ({len(hedefler)} kayıt)",
        gonder_url=url_for("abone_toplu_mesaj"), sorgu=sorgu_dizesi,
        geri_url=url_for("abone_listesi") + (f"?{sorgu_dizesi}" if sorgu_dizesi else ""),
        liste_url=url_for("abone_listesi"), netgsm_ayarli_mi=_netgsm_ayarli_mi(),
        sms_hazir=_netgsm_sms_ayarli_mi(), eposta_hazir=_eposta_ayarli_mi(), geri="",
    )


# "fatura" tablosunun kendisinde faturanın KİME kesildiğine dair bilgi
# (ad/soyad/köy/telefon/adres) tutulmuyor — bunlar kaynak_tur'a göre ('abone'
# veya 'ariza') abone/ariza tablosunda duruyor. Bu yüzden Faturalarım
# listesinde ve sütun filtrelerinde bir GERÇEK tablo yerine, bu bilgiyi
# LEFT JOIN ile önceden birleştiren bir alt sorguyu "tablo" gibi kullanıyoruz
# — _kolon_secenekleri gibi genel fonksiyonlar zaten "FROM {tablo}" şeklinde
# çalıştığı için bu alt sorgu metnini oldukları gibi kabul edip çalışıyorlar.
_FATURA_ALT_SORGU = (
    "(SELECT f.id, f.kaynak_tur, f.kaynak_id, f.fatura_turu, f.durum, "
    "f.tutar_kdv_dahil, f.kalemler, f.fatura_tarihi, f.created_at, "
    "CASE WHEN f.kaynak_tur = 'abone' THEN a.adi ELSE r.adi END AS alici_adi, "
    "CASE WHEN f.kaynak_tur = 'abone' THEN a.soyadi ELSE r.soyadi END AS alici_soyadi, "
    "CASE WHEN f.kaynak_tur = 'abone' THEN a.koy_adi ELSE r.koy_adi END AS alici_koy, "
    "CASE WHEN f.kaynak_tur = 'abone' THEN a.telefon ELSE r.telefon END AS alici_telefon, "
    "CASE WHEN f.kaynak_tur = 'abone' THEN a.adres ELSE r.adres END AS alici_adres "
    "FROM fatura f "
    "LEFT JOIN abone a ON f.kaynak_tur = 'abone' AND a.id = f.kaynak_id "
    "LEFT JOIN ariza r ON f.kaynak_tur = 'ariza' AND r.id = f.kaynak_id) fv"
)

FATURA_DISPLAY_KOLONLARI = [
    ("fatura_tarihi", "Tarih"),
    ("fatura_turu", "Fatura Türü"),
    ("alici_ad_soyad", "Alıcı"),
    ("alici_koy", "Alıcı Köy"),
    ("alici_telefon", "Alıcı Telefon"),
    ("kalemler", "Kalemler"),
    ("tutar_kdv_dahil", "KDV Dahil Tutar"),
    ("durum", "Durum"),
]

FATURA_KOLON_BILGI = {
    "fatura_tarihi": ("fatura_tarihi", "tarih"),
    "fatura_turu": ("CASE WHEN fatura_turu = 'earsiv' THEN 'e-Arşiv' ELSE 'e-Fatura' END", "metin"),
    "alici_ad_soyad": ("(COALESCE(alici_adi, '') || ' ' || COALESCE(alici_soyadi, ''))", "metin"),
    "alici_koy": ("alici_koy", "metin"),
    "alici_telefon": ("alici_telefon", "metin"),
    "kalemler": ("kalemler", "metin"),
    "tutar_kdv_dahil": ("tutar_kdv_dahil", "sayi"),
    "durum": ("CASE WHEN durum = 'basarili' THEN 'Başarılı' WHEN durum = 'hata' THEN 'Hata' ELSE 'Beklemede' END", "metin"),
}

FATURA_SAYISAL_KOLONLAR = {k for k, (_, tur) in FATURA_KOLON_BILGI.items() if tur == "sayi"}

_FATURA_ALAN_TANIMLARI = [
    ("alici_ad_soyad", "Alıcı Adı Soyadı", "(COALESCE(alici_adi, '') || ' ' || COALESCE(alici_soyadi, ''))", False),
    ("alici_koy", "Alıcı Köy", "alici_koy", False),
    ("alici_telefon", "Alıcı Telefon", "alici_telefon", False),
    ("alici_adres", "Alıcı Adres", "alici_adres", False),
    ("kalemler", "Kalemler", "kalemler", False),
    ("tutar_kdv_dahil", "KDV Dahil Tutar", "tutar_kdv_dahil", True),
]
_FATURA_ALAN_HARITASI = {k: (kolon, sayisal) for k, _, kolon, sayisal in _FATURA_ALAN_TANIMLARI}


def _fatura_kolon_takimi(db=None):
    return FATURA_DISPLAY_KOLONLARI, FATURA_KOLON_BILGI, FATURA_SAYISAL_KOLONLAR, []


def _fatura_filtre_kosulu_olustur(disari_anahtar, kolon_bilgi):
    kosul = "1=1"
    params = []
    q = request.args.get("q", "").strip()
    alanlar_secili = request.args.getlist("alan")
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in _FATURA_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in _FATURA_ALAN_HARITASI:
                kolon, sayisal = _FATURA_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            kosul += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params
    for anahtar in kolon_bilgi.keys():
        if anahtar == disari_anahtar:
            continue
        alt_kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if alt_kosul:
            kosul += f" AND {alt_kosul}"
            params += param_listesi
    return kosul, params


def _fatura_satir_sozlugu(k):
    tarih_kaynagi = k["fatura_tarihi"] or (str(k["created_at"])[:10] if k["created_at"] else None)
    return {
        "id": k["id"],
        "tarih": _gg_aa_yyyy(str(tarih_kaynagi)) if tarih_kaynagi else "",
        "kaynak_tur": k["kaynak_tur"],
        "kaynak_id": k["kaynak_id"],
        "fatura_turu": "e-Arşiv" if k["fatura_turu"] == "earsiv" else "e-Fatura",
        "alici_ad_soyad": ((k["alici_adi"] or "") + " " + (k["alici_soyadi"] or "")).strip(),
        "alici_koy": k["alici_koy"] or "",
        "alici_telefon": _telefon_formatla(k["alici_telefon"]),
        "kalemler": k["kalemler"] or "",
        "tutar_kdv_dahil": tl_format(k["tutar_kdv_dahil"]) if k["tutar_kdv_dahil"] is not None else "-",
        "durum": k["durum"],
    }


@app.route("/faturalar")
@login_required
def fatura_listesi():
    """'Faturalarım' sayfası — artık faturanın KİME kesildiğine dair bilgiler
    (alıcı adı/köy/telefon) ve Abone Listesi'ndekiyle aynı arama/sütun
    filtreleme sistemi de burada."""
    yonlendirme = _filtre_durumu_uygula("fatura_listesi")
    if yonlendirme:
        return yonlendirme

    db = get_db()
    q = request.args.get("q", "").strip()
    alanlar_secili = request.args.getlist("alan")
    alan_listesi = [(k, etiket) for k, etiket, _, _ in _FATURA_ALAN_TANIMLARI]

    sql = f"SELECT * FROM {_FATURA_ALT_SORGU} WHERE 1=1"
    params = []
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in _FATURA_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in _FATURA_ALAN_HARITASI:
                kolon, sayisal = _FATURA_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            sql += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params

    kolon_listesi, kolon_bilgi, sayisal_kolonlar, _ozel = _fatura_kolon_takimi()
    deger_secili = {}
    haric_secili = {}
    for anahtar, _e in kolon_listesi:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")
        haric_secili[anahtar] = request.args.getlist(f"haric_{anahtar}")
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi

    sql += f" ORDER BY id {'ASC' if _sira_yonu_al() == 'asc' else 'DESC'} LIMIT 500"

    cur = db.cursor()
    cur.execute(sql, params)
    faturalar_ham = cur.fetchall()
    cur.close()

    satirlar = [_fatura_satir_sozlugu(k) for k in faturalar_ham]
    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "fatura_listesi.html", satirlar=satirlar,
        q=q, secili_alanlar=alanlar_secili, alan_listesi=alan_listesi,
        kolon_listesi=kolon_listesi, deger_secili=deger_secili, haric_secili=haric_secili,
        sayisal_kolonlar=sayisal_kolonlar,
        arama_satir=_izgara_satir(len(alan_listesi)),
        arama_satir_2=_izgara_satir(len(alan_listesi), 2),
        filtreli_kayit=filtreli_kayit,
        sira=_sira_yonu_al(), sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
    )


@app.route("/fatura/<int:fatura_id>")
@login_required
def fatura_goruntule(fatura_id):
    """Tek bir faturanın sonucunu gösterir."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM fatura WHERE id = %s", (fatura_id,))
    fatura = cur.fetchone()
    cur.close()
    if fatura is None:
        flash("Fatura kaydı bulunamadı.")
        return redirect(url_for("fatura_listesi"))
    return render_template("fatura_goster.html", fatura=fatura)


@app.route("/fatura/<int:fatura_id>/pdf")
@login_required
def fatura_pdf_goster(fatura_id):
    """Faturanın PDF'ini gösterir."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM fatura WHERE id = %s", (fatura_id,))
    fatura = cur.fetchone()
    cur.close()
    if fatura is None:
        flash("Fatura kaydı bulunamadı.")
        return redirect(url_for("fatura_listesi"))

    if not fatura["pdf_icerik"]:
        basarili, hata = _hizli_fatura_pdf_al(db, fatura_id)
        if not basarili:
            flash(f"Fatura PDF'i alınamadı: {hata}")
            return redirect(url_for("fatura_goruntule", fatura_id=fatura_id))
        cur = db.cursor()
        cur.execute("SELECT pdf_icerik FROM fatura WHERE id = %s", (fatura_id,))
        fatura = cur.fetchone()
        cur.close()

    return Response(bytes(fatura["pdf_icerik"]), mimetype="application/pdf")


@app.route("/stok")
@login_required
def stok_listesi():
    """Ürün/malzeme kataloğunu, güncel stok miktarlarını ve düşük stok uyarılarını gösterir."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM stok_urun WHERE aktif = TRUE ORDER BY urun_adi")
    urunler = cur.fetchall()
    cur.close()
    dusuk_stok_sayisi = sum(1 for u in urunler if float(u["stok_miktari"] or 0) <= float(u["min_stok_seviyesi"] or 0))
    return render_template("stok_listesi.html", urunler=urunler, dusuk_stok_sayisi=dusuk_stok_sayisi)


@app.route("/stok/yeni", methods=["GET", "POST"])
@login_required
def stok_yeni():
    """Yeni bir ürün/malzeme kaydı (stok kartı) oluşturur."""
    if request.method == "POST":
        urun_adi = request.form.get("urun_adi", "").strip()
        if not urun_adi:
            flash("Ürün adı zorunludur.")
            return redirect(url_for("stok_yeni"))
        db = get_db()
        cur = db.cursor()
        cur.execute(
            "INSERT INTO stok_urun (urun_adi, birim, birim_fiyat, kdv_orani, stok_miktari, min_stok_seviyesi, aciklama) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (
                urun_adi,
                request.form.get("birim", "ADET").strip() or "ADET",
                _sayi_veya(request.form.get("birim_fiyat"), 0),
                _sayi_veya(request.form.get("kdv_orani"), 20),
                _sayi_veya(request.form.get("stok_miktari"), 0),
                _sayi_veya(request.form.get("min_stok_seviyesi"), 0),
                request.form.get("aciklama", "").strip(),
            ),
        )
        db.commit()
        cur.close()
        flash("Ürün eklendi.")
        return redirect(url_for("stok_listesi"))
    return render_template("stok_form.html", kayit=None)


@app.route("/stok/<int:urun_id>/duzenle", methods=["GET", "POST"])
@login_required
def stok_duzenle(urun_id):
    """Bir ürün/malzeme kaydını düzenler."""
    db = get_db()
    if request.method == "POST":
        urun_adi = request.form.get("urun_adi", "").strip()
        if not urun_adi:
            flash("Ürün adı zorunludur.")
            return redirect(url_for("stok_duzenle", urun_id=urun_id))
        cur = db.cursor()
        cur.execute(
            "UPDATE stok_urun SET urun_adi=%s, birim=%s, birim_fiyat=%s, kdv_orani=%s, "
            "stok_miktari=%s, min_stok_seviyesi=%s, aciklama=%s WHERE id=%s",
            (
                urun_adi,
                request.form.get("birim", "ADET").strip() or "ADET",
                _sayi_veya(request.form.get("birim_fiyat"), 0),
                _sayi_veya(request.form.get("kdv_orani"), 20),
                _sayi_veya(request.form.get("stok_miktari"), 0),
                _sayi_veya(request.form.get("min_stok_seviyesi"), 0),
                request.form.get("aciklama", "").strip(),
                urun_id,
            ),
        )
        db.commit()
        cur.close()
        flash("Ürün güncellendi.")
        return redirect(url_for("stok_listesi"))
    cur = db.cursor()
    cur.execute("SELECT * FROM stok_urun WHERE id = %s", (urun_id,))
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Ürün bulunamadı.")
        return redirect(url_for("stok_listesi"))
    return render_template("stok_form.html", kayit=kayit)


@app.route("/stok/<int:urun_id>/sil", methods=["POST"])
@login_required
def stok_sil(urun_id):
    """Bir ürünü (ve tüm hareket geçmişini) siler."""
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM stok_urun WHERE id = %s", (urun_id,))
    db.commit()
    cur.close()
    flash("Ürün silindi.")
    return redirect(url_for("stok_listesi"))


@app.route("/stok/<int:urun_id>/hareket", methods=["GET", "POST"])
@login_required
def stok_hareket_ekle(urun_id):
    """Bir ürüne giriş ya da çıkış hareketi ekler ve stok_miktari'nı günceller."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM stok_urun WHERE id = %s", (urun_id,))
    urun = cur.fetchone()
    cur.close()
    if urun is None:
        flash("Ürün bulunamadı.")
        return redirect(url_for("stok_listesi"))

    if request.method == "POST":
        hareket_turu = request.form.get("hareket_turu")
        if hareket_turu not in ("giris", "cikis"):
            flash("Geçerli bir hareket türü seçin.")
            return redirect(url_for("stok_hareket_ekle", urun_id=urun_id))
        miktar = _sayi_veya(request.form.get("miktar"), 0)
        if miktar <= 0:
            flash("Miktar sıfırdan büyük olmalı.")
            return redirect(url_for("stok_hareket_ekle", urun_id=urun_id))
        tarih = request.form.get("tarih", "").strip() or datetime.now().strftime("%Y-%m-%d")

        cur = db.cursor()
        cur.execute(
            "INSERT INTO stok_hareket (urun_id, hareket_turu, miktar, tarih, birim_fiyat, tedarikci_adi, aciklama, olusturan_kullanici) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
            (
                urun_id, hareket_turu, miktar, tarih,
                _sayi_veya(request.form.get("birim_fiyat"), None) if hareket_turu == "giris" else None,
                request.form.get("tedarikci_adi", "").strip() if hareket_turu == "giris" else "",
                request.form.get("aciklama", "").strip(),
                session.get("kullanici_adi", ""),
            ),
        )
        degisim = miktar if hareket_turu == "giris" else -miktar
        cur.execute("UPDATE stok_urun SET stok_miktari = stok_miktari + %s WHERE id = %s", (degisim, urun_id))
        db.commit()
        cur.close()
        flash("Stok hareketi kaydedildi.")
        return redirect(url_for("stok_hareketleri", urun_id=urun_id))

    return render_template(
        "stok_hareket_form.html", urun=urun, bugun=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/stok/<int:urun_id>/hareketler")
@login_required
def stok_hareketleri(urun_id):
    """Bir ürünün tüm giriş/çıkış hareket geçmişini gösterir."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM stok_urun WHERE id = %s", (urun_id,))
    urun = cur.fetchone()
    cur.close()
    if urun is None:
        flash("Ürün bulunamadı.")
        return redirect(url_for("stok_listesi"))
    cur = db.cursor()
    cur.execute(
        "SELECT * FROM stok_hareket WHERE urun_id = %s ORDER BY tarih DESC, id DESC",
        (urun_id,),
    )
    hareketler = cur.fetchall()
    cur.close()
    return render_template("stok_hareketleri.html", urun=urun, hareketler=hareketler)


@app.route("/stok/hareket/<int:hareket_id>/sil", methods=["POST"])
@login_required
def stok_hareket_sil(hareket_id):
    """Bir stok hareketini siler ve stok_miktari'na olan etkisini geri alır."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT urun_id, hareket_turu, miktar FROM stok_hareket WHERE id = %s", (hareket_id,))
    hareket = cur.fetchone()
    if hareket is None:
        cur.close()
        flash("Hareket kaydı bulunamadı.")
        return redirect(url_for("stok_listesi"))
    geri_alma = -hareket["miktar"] if hareket["hareket_turu"] == "giris" else hareket["miktar"]
    cur.execute("UPDATE stok_urun SET stok_miktari = stok_miktari + %s WHERE id = %s", (geri_alma, hareket["urun_id"]))
    cur.execute("DELETE FROM stok_hareket WHERE id = %s", (hareket_id,))
    db.commit()
    urun_id = hareket["urun_id"]
    cur.close()
    flash("Hareket silindi, stok miktarı geri alındı.")
    return redirect(url_for("stok_hareketleri", urun_id=urun_id))


# ---------------------------------------------------------------------------
# FABRİKA / TAMİR MODÜLÜ: arızalı sayaçların üreticiye/fabrikaya tamire
# gönderilip geri gelme sürecinin takibi, kargo/koli (8'erli kutu) yönetimi
# ve "Sayaç Durum Raporu" (ARIZALI SAYAÇ BİLGİ FORMU) çıktısı.
# ---------------------------------------------------------------------------

FABRIKA_KOLI_KAPASITESI = 8

FABRIKA_DURUM_ETIKETLERI = {
    "beklemede": "Beklemede (Gönderilmedi)",
    "gonderildi": "Gönderildi",
    "tamirde": "Tamirde",
    "tamir_edildi": "Tamir Edildi",
    "iade_edildi": "İade Edildi",
}

_FABRIKA_FONT_KAYITLI = False


def _fabrika_durum_etiketi(durum):
    return FABRIKA_DURUM_ETIKETLERI.get(durum, durum or "")


def _fabrika_font_hazirla():
    """Sayaç Durum Raporu PDF'inde Türkçe karakterlerin doğru görünmesi için
    depoya gömülen DejaVuSans fontunu reportlab'a kayıt eder."""
    global _FABRIKA_FONT_KAYITLI
    if _FABRIKA_FONT_KAYITLI or pdfmetrics is None:
        return
    taban = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "fonts")
    pdfmetrics.registerFont(TTFont("DejaVuSans", os.path.join(taban, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", os.path.join(taban, "DejaVuSans-Bold.ttf")))
    _FABRIKA_FONT_KAYITLI = True


def _gg_aa_yyyy_veya(deger):
    """Bir tarih değerini GG.AA.YYYY biçimine çevirir; boşsa boş string döner."""
    if not deger:
        return ""
    if isinstance(deger, str):
        try:
            deger = datetime.strptime(deger[:10], "%Y-%m-%d").date()
        except ValueError:
            return deger
    try:
        return deger.strftime("%d.%m.%Y")
    except Exception:
        return str(deger)


def _fabrika_filigran_ciz(canvas, belge):
    """Her sayfanın arkasına, sayfanın ortasına soluk su damlası görselini filigran olarak çizer."""
    if RLImage is None:
        return
    taban = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
    yol = os.path.join(taban, "fabrika_su_damla_filigran.png")
    if not os.path.exists(yol):
        return
    genislik = 100 * mm
    yukseklik = genislik * (168 / 286)
    sayfa_genislik, sayfa_yukseklik = A4
    x = (sayfa_genislik - genislik) / 2
    y = (sayfa_yukseklik - yukseklik) / 2
    canvas.saveState()
    try:
        canvas.drawImage(yol, x, y, width=genislik, height=yukseklik, mask="auto")
    except Exception:
        pass
    canvas.restoreState()


def _fabrika_rapor_pdf_olustur(gonderim, koliler, satici):
    """Sayaç Durum Raporu'nu reportlab ile PDF olarak üretir. Her koli kendi sayfasıdır."""
    _fabrika_font_hazirla()
    arabellek = io.BytesIO()
    belge = SimpleDocTemplate(
        arabellek, pagesize=A4,
        topMargin=14 * mm, bottomMargin=14 * mm, leftMargin=14 * mm, rightMargin=14 * mm,
    )
    static_taban = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
    logo_yolu = os.path.join(static_taban, "fabrika_logo.jpg")
    baslik_stili = ParagraphStyle("fabrika_baslik", fontName="DejaVuSans-Bold", fontSize=13, alignment=1, spaceAfter=10)
    normal_stil = ParagraphStyle("fabrika_normal", fontName="DejaVuSans", fontSize=9, leading=11)
    normal_kalin = ParagraphStyle("fabrika_normal_kalin", fontName="DejaVuSans-Bold", fontSize=9, leading=11)
    hucre_stil = ParagraphStyle("fabrika_hucre", fontName="DejaVuSans", fontSize=8, leading=10)
    baslik_hucre_stil = ParagraphStyle("fabrika_baslik_hucre", fontName="DejaVuSans-Bold", fontSize=8, leading=10, alignment=1)

    urun_tanimi = gonderim.get("urun_tanimi") or "Elektronik Kartlı Ön Ödemeli Su Sayacı"
    adres = gonderim.get("adres") or ""
    kargo_firmasi = gonderim.get("kargo_firmasi") or ""
    kargo_takip_no = gonderim.get("kargo_takip_no") or ""
    toplam_koli = len(koliler)

    yetkili_bayii_adi = gonderim.get("yetkili_bayii") or ""

    ogeler = []
    for sira, koli in enumerate(koliler, start=1):
        if sira > 1:
            ogeler.append(PageBreak())

        tarih_str = _gg_aa_yyyy_veya(koli.get("koli_tarihi") or gonderim.get("gonderim_tarihi"))

        if os.path.exists(logo_yolu):
            logo_genislik = 56 * mm
            logo_yukseklik = logo_genislik * (229 / 1600)
            logo_resmi = RLImage(logo_yolu, width=logo_genislik, height=logo_yukseklik)
            logo_resmi.hAlign = "LEFT"
            baslik_hucresi = [logo_resmi, Paragraph("ARIZALI SAYAÇ BİLGİ FORMU", baslik_stili)]
        else:
            baslik_hucresi = [Paragraph("ARIZALI SAYAÇ BİLGİ FORMU", baslik_stili)]

        ust_bilgi = Table(
            [
                [baslik_hucresi, "", "", ""],
                [Paragraph("ADRES", normal_kalin), Paragraph(adres, normal_stil),
                 Paragraph("TARİH", normal_kalin), Paragraph(tarih_str, normal_stil)],
                [Paragraph("ÜRÜN TANIMI", normal_kalin), Paragraph(urun_tanimi, normal_stil),
                 Paragraph("KOLİ", normal_kalin), Paragraph(f"{sira} / {toplam_koli}", normal_stil)],
                [Paragraph("KARGO FİRMASI", normal_kalin), Paragraph(kargo_firmasi, normal_stil),
                 Paragraph("TAKİP NO", normal_kalin), Paragraph(kargo_takip_no, normal_stil)],
            ],
            colWidths=[28 * mm, 70 * mm, 22 * mm, 52 * mm],
        )
        ust_bilgi.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("VALIGN", (0, 0), (-1, 0), "TOP"),
            ("BACKGROUND", (0, 1), (0, -1), colors.whitesmoke),
            ("BACKGROUND", (2, 1), (2, -1), colors.whitesmoke),
            ("SPAN", (0, 0), (-1, 0)),
        ]))
        ogeler.append(ust_bilgi)

        basliklar = ["S.NO", "SERİ NO", "ÜRETİM\nYILI", "ARIZA DURUMU", "SAYAÇ SAHİBİ"]
        veri = [[Paragraph(b, baslik_hucre_stil) for b in basliklar]]
        kayitlar = koli.get("kayitlar") or []
        for i in range(FABRIKA_KOLI_KAPASITESI):
            if i < len(kayitlar):
                k = kayitlar[i]
                sahip = f"{k.get('koy_adi') or ''}\n{k.get('abone_adi') or ''}".strip()
                veri.append([
                    Paragraph(str(i + 1), hucre_stil),
                    Paragraph(k.get("seri_no") or "", hucre_stil),
                    Paragraph(k.get("uretim_yili") or "", hucre_stil),
                    Paragraph(k.get("tespit_edilen_ariza") or "", hucre_stil),
                    Paragraph(sahip, hucre_stil),
                ])
            else:
                veri.append([Paragraph(str(i + 1), hucre_stil), "", "", "", ""])
        veri.append([
            Paragraph("TOPLAM", normal_kalin), "", "",
            Paragraph(f"{len(kayitlar)} Adet Su Sayacı", normal_kalin), "",
        ])
        bosluk_satir_sayisi = 3
        for _ in range(bosluk_satir_sayisi):
            veri.append(["", "", "", "", ""])
        gövde_satir_sayisi = 1 + FABRIKA_KOLI_KAPASITESI + 1
        satir_yukseklikleri = ([None] * gövde_satir_sayisi) + [14.5 * mm, 9.5 * mm, 6 * mm]
        tablo = Table(
            veri,
            colWidths=[12 * mm, 21 * mm, 17 * mm, 99 * mm, 23 * mm],
            rowHeights=satir_yukseklikleri,
            repeatRows=1,
        )
        gri_satir = gövde_satir_sayisi + 1
        tablo.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, gövde_satir_sayisi - 1), 0.5, colors.grey),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
            ("LINEABOVE", (0, gövde_satir_sayisi), (-1, gövde_satir_sayisi), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, gövde_satir_sayisi - 1), (-1, gövde_satir_sayisi - 1), colors.whitesmoke),
            ("SPAN", (0, gövde_satir_sayisi - 1), (2, gövde_satir_sayisi - 1)),
            ("SPAN", (3, gövde_satir_sayisi - 1), (4, gövde_satir_sayisi - 1)),
            ("BACKGROUND", (0, gri_satir), (-1, gri_satir), colors.HexColor("#f2f2f2")),
        ]))
        ogeler.append(tablo)

        sol_hucre = [Paragraph("YETKİLİ BAYİİ", normal_kalin)]
        if yetkili_bayii_adi:
            sol_hucre.append(Paragraph(yetkili_bayii_adi, normal_kalin))
        sol_hucre.append(Paragraph("İMZA", normal_stil))
        sag_metin = f"{satici.get('unvan', '')} – {satici.get('yetkili', '')}".strip(" –")
        imza = Table(
            [
                [sol_hucre, Paragraph(sag_metin, normal_kalin)],
                [Spacer(1, 24), Spacer(1, 24)],
            ],
            colWidths=[86 * mm, 86 * mm],
        )
        imza.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        ogeler.append(imza)

    belge.build(ogeler, onFirstPage=_fabrika_filigran_ciz, onLaterPages=_fabrika_filigran_ciz)
    return arabellek.getvalue()


def _fabrika_rapor_verisi(db, gonderim_id):
    """Bir gönderime ait kargo/koli/kayıt verilerini ve satıcı bilgisini toplar."""
    cur = db.cursor()
    cur.execute("SELECT * FROM fabrika_gonderim WHERE id = %s", (gonderim_id,))
    gonderim = cur.fetchone()
    if gonderim is None:
        cur.close()
        return None, None, None
    cur.execute("SELECT * FROM fabrika_koli WHERE gonderim_id = %s ORDER BY koli_no", (gonderim_id,))
    koliler = cur.fetchall()
    for koli in koliler:
        cur.execute("SELECT * FROM fabrika_tamir WHERE koli_id = %s AND silindi_mi IS NOT TRUE ORDER BY id", (koli["id"],))
        koli["kayitlar"] = cur.fetchall()
    cur.close()
    satici_unvan = _ayar_getir(db, "fatura_satici_unvan") or ""
    satici_ad = _ayar_getir(db, "fatura_satici_ad") or ""
    satici_soyad = _ayar_getir(db, "fatura_satici_soyad") or ""
    satici = {"unvan": satici_unvan, "yetkili": f"{satici_ad} {satici_soyad}".strip()}
    return gonderim, koliler, satici


def _fabrika_yetkili_bayii_listesi(db):
    """Sayaç Durum Raporu'nda 'YETKİLİ BAYİİ' tarafında imza atacak kişi için,
    daha önce kullanılmış isimlerden oluşan listeyi döndürür."""
    cur = db.cursor()
    cur.execute(
        "SELECT DISTINCT yetkili_bayii FROM fabrika_gonderim "
        "WHERE yetkili_bayii IS NOT NULL AND yetkili_bayii <> '' ORDER BY yetkili_bayii"
    )
    isimler = [r["yetkili_bayii"] for r in cur.fetchall()]
    cur.close()
    return isimler


def _fabrika_gonderim_sira_numaralarini_yenile(db):
    """Gönderimler listesindeki '#' sırası artık id'ye (kayıt eklenme sırasına)
    değil, Gönderim Tarihi'ne göre (eskiden yeniye) hesaplanır ve 1'den
    başlayarak boşluksuz yeniden numaralandırılır — bkz. abone/ariza
    kayıtlarındaki aynı mantık (_abone_sira_numaralarini_yenile /
    _ariza_sira_numaralarini_yenile). Sonradan girilen ama tarihi daha eski
    olan bir gönderim sıraya doğru yerine otursun, silinen bir gönderimin
    sırası da boşluk bırakmasın diye bu fonksiyon her ekleme/silme
    sonrasında çağrılıyor."""
    cur = db.cursor()
    cur.execute(
        """
        UPDATE fabrika_gonderim fg
        SET sira_no = t.yeni_sira
        FROM (
            SELECT id, ROW_NUMBER() OVER (ORDER BY gonderim_tarihi ASC NULLS LAST, id ASC) AS yeni_sira
            FROM fabrika_gonderim
        ) t
        WHERE fg.id = t.id AND fg.sira_no IS DISTINCT FROM t.yeni_sira
        """
    )
    db.commit()
    cur.close()


@app.route("/fabrika")
@login_required
def fabrika_listesi():
    """Fabrikaya/üreticiye tamire gönderilen arızalı sayaç kayıtlarını,
    Abone Listesi'ndekiyle aynı arama/sütun filtreleme sistemiyle listeler."""
    yonlendirme = _filtre_durumu_uygula("fabrika_listesi")
    if yonlendirme:
        return yonlendirme

    db = get_db()
    durum_filtre = request.args.get("durum", "").strip()
    koy = request.args.get("koy", "").strip()
    q = request.args.get("q", "").strip()
    alanlar_secili = request.args.getlist("alan")
    alan_listesi = [(k, etiket) for k, etiket, _, _ in _FABRIKA_ALAN_TANIMLARI]

    sql = (
        "SELECT ft.*, fk.koli_no, fk.gonderim_id AS koli_gonderim_id "
        "FROM fabrika_tamir ft LEFT JOIN fabrika_koli fk ON fk.id = ft.koli_id "
        "WHERE ft.silindi_mi IS NOT TRUE"
    )
    params = []
    if durum_filtre and durum_filtre in FABRIKA_DURUM_ETIKETLERI:
        sql += " AND ft.durum = %s"
        params.append(durum_filtre)
    if koy:
        sql += " AND ft.koy_adi = %s"
        params.append(koy)

    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in _FABRIKA_ALAN_TANIMLARI]
        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None
        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in _FABRIKA_ALAN_HARITASI:
                kolon, sayisal = _FABRIKA_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            sql += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params

    kolon_listesi, kolon_bilgi, sayisal_kolonlar, _ozel = _fabrika_kolon_takimi()
    deger_secili = {}
    haric_secili = {}
    for anahtar, _e in kolon_listesi:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")
        haric_secili[anahtar] = request.args.getlist(f"haric_{anahtar}")
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi

    sql += f" ORDER BY ft.sira_no {'DESC' if _sira_yonu_al() == 'desc' else 'ASC'}"

    cur = db.cursor()
    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()
    cur.execute("SELECT DISTINCT koy_adi FROM fabrika_tamir WHERE koy_adi IS NOT NULL AND koy_adi <> '' ORDER BY koy_adi")
    koyler = cur.fetchall()
    cur.execute("SELECT COUNT(*) AS c FROM fabrika_tamir WHERE silindi_mi IS NOT TRUE")
    toplam_kayit = cur.fetchone()["c"]
    cur.close()

    satirlar = [_fabrika_satir_sozlugu(k) for k in kayitlar_ham]
    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "fabrika_listesi.html", satirlar=satirlar, durum_filtre=durum_filtre,
        durum_etiketleri=FABRIKA_DURUM_ETIKETLERI, koyler=koyler, secili_koy=koy,
        q=q, secili_alanlar=alanlar_secili, alan_listesi=alan_listesi,
        kolon_listesi=kolon_listesi, deger_secili=deger_secili, haric_secili=haric_secili,
        sayisal_kolonlar=sayisal_kolonlar,
        arama_satir=_izgara_satir(len(alan_listesi)),
        arama_satir_2=_izgara_satir(len(alan_listesi), 2),
        filtreli_kayit=filtreli_kayit, toplam_kayit=toplam_kayit,
        sira=_sira_yonu_al(), sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
    )


@app.route("/fabrika/yeni", methods=["GET", "POST"])
@login_required
def fabrika_yeni():
    """Yeni bir arızalı sayaç / tamir kaydı oluşturur (henüz 'beklemede')."""
    if request.method == "POST":
        seri_no = request.form.get("seri_no", "").strip()
        if not seri_no:
            flash("Seri No zorunludur.")
            return redirect(url_for("fabrika_yeni"))
        yerine_takildi = request.form.get("yerine_sayac_takildi") == "takildi"
        db = get_db()
        cur = db.cursor()
        abone_karti = request.form.get("abone_karti", "").strip()
        if abone_karti not in ("alindi", "alinmadi"):
            abone_karti = "alinmadi"
        cur.execute(
            "INSERT INTO fabrika_tamir "
            "(seri_no, abone_adi, koy_adi, telefon, ilk_montaj_tarihi, uretim_yili, "
            "tespit_edilen_ariza, yerine_sayac_takildi, takilan_sayac_serisi, "
            "tamir_ucreti, parca_maliyeti, odeyen, abone_karti, olusturan_kullanici) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id",
            (
                seri_no,
                request.form.get("abone_adi", "").strip(),
                request.form.get("koy_adi", "").strip(),
                request.form.get("telefon", "").strip(),
                request.form.get("ilk_montaj_tarihi", "").strip() or None,
                request.form.get("uretim_yili", "").strip(),
                request.form.get("tespit_edilen_ariza", "").strip(),
                yerine_takildi,
                request.form.get("takilan_sayac_serisi", "").strip() if yerine_takildi else "",
                _sayi_veya(request.form.get("tamir_ucreti"), 0),
                _sayi_veya(request.form.get("parca_maliyeti"), 0),
                request.form.get("odeyen", "").strip(),
                abone_karti,
                session.get("kullanici_adi", ""),
            ),
        )
        yeni_id = cur.fetchone()["id"]
        db.commit()
        cur.close()
        _fabrika_tamir_sira_numaralarini_yenile(db)
        _fabrika_fotograflarini_kaydet(db, yeni_id, request.files.getlist("fotograflar"))
        flash("Tamir kaydı eklendi (Beklemede). Gönderime dahil etmek için Fabrika/Tamir listesinden seçip "
              "'Gönderim Oluştur' deyin.")
        return redirect(url_for("fabrika_listesi"))
    return render_template("fabrika_form.html", kayit=None, fotograflar=[])


@app.route("/fabrika/<int:kayit_id>/duzenle", methods=["GET", "POST"])
@login_required
def fabrika_duzenle(kayit_id):
    """Bir tamir kaydını düzenler."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM fabrika_tamir WHERE id = %s", (kayit_id,))
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Tamir kaydı bulunamadı.")
        return redirect(url_for("fabrika_listesi"))
    if kayit["silindi_mi"]:
        flash("Bu kayıt silinmiş durumda. Düzenleyebilmek için önce Silinenler sayfasından geri yükleyin.")
        return redirect(url_for("fabrika_silinenler"))

    if request.method == "POST":
        seri_no = request.form.get("seri_no", "").strip()
        if not seri_no:
            flash("Seri No zorunludur.")
            return redirect(url_for("fabrika_duzenle", kayit_id=kayit_id))
        yerine_takildi = request.form.get("yerine_sayac_takildi") == "takildi"
        abone_karti = request.form.get("abone_karti", "").strip()
        if abone_karti not in ("alindi", "alinmadi"):
            abone_karti = "alinmadi"

        yeni_durum = kayit["durum"]
        if kayit["koli_id"] is not None:
            secilen_durum = request.form.get("durum", "").strip()
            if secilen_durum in ("gonderildi", "tamirde", "tamir_edildi", "iade_edildi"):
                yeni_durum = secilen_durum

        cur = db.cursor()
        cur.execute(
            "UPDATE fabrika_tamir SET seri_no=%s, abone_adi=%s, koy_adi=%s, telefon=%s, "
            "ilk_montaj_tarihi=%s, uretim_yili=%s, tespit_edilen_ariza=%s, "
            "yerine_sayac_takildi=%s, takilan_sayac_serisi=%s, durum=%s, donus_tarihi=%s, "
            "tamir_ucreti=%s, parca_maliyeti=%s, odeyen=%s, abone_karti=%s, updated_at=NOW() WHERE id=%s",
            (
                seri_no,
                request.form.get("abone_adi", "").strip(),
                request.form.get("koy_adi", "").strip(),
                request.form.get("telefon", "").strip(),
                request.form.get("ilk_montaj_tarihi", "").strip() or None,
                request.form.get("uretim_yili", "").strip(),
                request.form.get("tespit_edilen_ariza", "").strip(),
                yerine_takildi,
                request.form.get("takilan_sayac_serisi", "").strip() if yerine_takildi else "",
                yeni_durum,
                request.form.get("donus_tarihi", "").strip() or None,
                _sayi_veya(request.form.get("tamir_ucreti"), 0),
                _sayi_veya(request.form.get("parca_maliyeti"), 0),
                request.form.get("odeyen", "").strip(),
                abone_karti,
                kayit_id,
            ),
        )
        db.commit()
        cur.close()
        _fabrika_fotograflarini_kaydet(db, kayit_id, request.files.getlist("fotograflar"))
        flash("Tamir kaydı güncellendi.")
        return redirect(url_for("fabrika_listesi"))

    koli_bilgi = None
    if kayit["koli_id"] is not None:
        cur = db.cursor()
        cur.execute(
            "SELECT fk.koli_no, fk.gonderim_id, fg.kargo_firmasi, fg.kargo_takip_no "
            "FROM fabrika_koli fk JOIN fabrika_gonderim fg ON fg.id = fk.gonderim_id "
            "WHERE fk.id = %s", (kayit["koli_id"],),
        )
        koli_bilgi = cur.fetchone()
        cur.close()
    cur = db.cursor()
    cur.execute(
        "SELECT id, dosya_adi, content_type FROM fabrika_fotograf WHERE kayit_id = %s ORDER BY id",
        (kayit_id,),
    )
    fotograflar = cur.fetchall()
    cur.close()
    return render_template(
        "fabrika_form.html", kayit=kayit, koli_bilgi=koli_bilgi, fotograflar=fotograflar,
        durum_etiketleri=FABRIKA_DURUM_ETIKETLERI,
    )


@app.route("/fabrika/<int:kayit_id>/sil", methods=["POST"])
@login_required
def fabrika_sil(kayit_id):
    """Bir tamir kaydını siler. Kayıt veritabanından kalıcı olarak silinmez,
    'silindi_mi' işaretlenip Silinenler sayfasına taşınır — yanlışlıkla
    silinen bir kayıt oradan geri yüklenebilir (bkz. fabrika_silinenler /
    fabrika_geri_yukle)."""
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "UPDATE fabrika_tamir SET silindi_mi = TRUE, silinme_tarihi = NOW() WHERE id = %s",
        (kayit_id,),
    )
    db.commit()
    cur.close()
    _fabrika_tamir_sira_numaralarini_yenile(db)
    flash("Tamir kaydı silindi. Yanlışlıkla sildiyseniz 'Silinenler' sayfasından geri yükleyebilirsiniz.")
    return redirect(url_for("fabrika_listesi"))


@app.route("/fabrika/silinenler")
@login_required
def fabrika_silinenler():
    """Silinmiş (silindi_mi=TRUE) tamir kayıtlarını listeler; buradan bir
    kayıt geri yüklenebilir ya da kalıcı olarak silinebilir."""
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT ft.*, fk.koli_no, fk.gonderim_id AS koli_gonderim_id "
        "FROM fabrika_tamir ft LEFT JOIN fabrika_koli fk ON fk.id = ft.koli_id "
        "WHERE ft.silindi_mi IS TRUE ORDER BY ft.silinme_tarihi DESC"
    )
    kayitlar = cur.fetchall()
    cur.close()
    return render_template(
        "fabrika_silinenler.html", kayitlar=kayitlar, durum_etiketleri=FABRIKA_DURUM_ETIKETLERI,
    )


@app.route("/fabrika/<int:kayit_id>/geri-yukle", methods=["POST"])
@login_required
def fabrika_geri_yukle(kayit_id):
    """Silinenler sayfasından bir kaydı eski durumuna geri getirir."""
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "UPDATE fabrika_tamir SET silindi_mi = FALSE, silinme_tarihi = NULL WHERE id = %s",
        (kayit_id,),
    )
    db.commit()
    cur.close()
    _fabrika_tamir_sira_numaralarini_yenile(db)
    flash("Tamir kaydı geri yüklendi.")
    return redirect(url_for("fabrika_silinenler"))


@app.route("/fabrika/<int:kayit_id>/kalici-sil", methods=["POST"])
@login_required
def fabrika_kalici_sil(kayit_id):
    """Silinenler sayfasındaki bir kaydı veritabanından kalıcı olarak siler
    (fotoğrafları da fabrika_fotograf tablosundaki ON DELETE CASCADE ile
    birlikte gider). Bu işlem geri alınamaz."""
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM fabrika_tamir WHERE id = %s AND silindi_mi IS TRUE", (kayit_id,))
    db.commit()
    cur.close()
    flash("Tamir kaydı kalıcı olarak silindi.")
    return redirect(url_for("fabrika_silinenler"))


@app.route("/fabrika-fotograf/<int:foto_id>")
@login_required
def fabrika_fotograf_goster(foto_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT content_type, icerik FROM fabrika_fotograf WHERE id = %s", (foto_id,))
    foto = cur.fetchone()
    cur.close()
    if foto is None:
        return "Fotoğraf bulunamadı.", 404
    yanit = Response(bytes(foto["icerik"]), mimetype=foto["content_type"] or "application/octet-stream")
    yanit.headers["X-Content-Type-Options"] = "nosniff"
    return yanit


@app.route("/fabrika-fotograf/<int:foto_id>/sil", methods=["POST"])
@login_required
def fabrika_fotograf_sil(foto_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT kayit_id FROM fabrika_fotograf WHERE id = %s", (foto_id,))
    foto = cur.fetchone()
    if foto:
        cur.execute("DELETE FROM fabrika_fotograf WHERE id = %s", (foto_id,))
        db.commit()
    cur.close()
    if foto:
        return redirect(url_for("fabrika_duzenle", kayit_id=foto["kayit_id"]))
    return redirect(url_for("fabrika_listesi"))


@app.route("/fabrika/gonderim-olustur", methods=["POST"])
@login_required
def fabrika_gonderim_olustur():
    """1. adım: işaretlenen kayıtları alır, kaç koli oluşacağını hesaplar ve
    kargo bilgisi girilecek onay ekranını gösterir."""
    secili_idler = request.form.getlist("secili_id")
    if not secili_idler:
        flash("Gönderim oluşturmak için en az bir kayıt seçmelisiniz.")
        return redirect(url_for("fabrika_listesi"))
    try:
        id_listesi = sorted({int(x) for x in secili_idler})
    except ValueError:
        flash("Geçersiz kayıt seçimi.")
        return redirect(url_for("fabrika_listesi"))

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT * FROM fabrika_tamir WHERE id = ANY(%s) AND durum = 'beklemede' AND koli_id IS NULL AND silindi_mi IS NOT TRUE "
        "ORDER BY id", (id_listesi,),
    )
    kayitlar = cur.fetchall()
    cur.close()
    if not kayitlar:
        flash("Seçilen kayıtlar artık 'Beklemede' durumunda değil (belki başka bir gönderime dahil edildi). "
              "Lütfen listeyi yenileyip tekrar deneyin.")
        return redirect(url_for("fabrika_listesi"))

    koli_sayisi = math.ceil(len(kayitlar) / FABRIKA_KOLI_KAPASITESI)
    varsayilan_adres = _ayar_getir(db, "fatura_satici_adres") or ""
    return render_template(
        "fabrika_gonderim_onay.html", kayitlar=kayitlar, koli_sayisi=koli_sayisi,
        koli_kapasitesi=FABRIKA_KOLI_KAPASITESI, varsayilan_adres=varsayilan_adres,
        bugun=datetime.now().strftime("%Y-%m-%d"),
        yetkili_bayii_listesi=_fabrika_yetkili_bayii_listesi(db),
    )


@app.route("/fabrika/gonderim-kaydet", methods=["POST"])
@login_required
def fabrika_gonderim_kaydet():
    """2. adım: onay ekranından gelen kargo bilgisiyle birlikte gönderimi ve
    içindeki koli'leri veritabanına kaydeder."""
    secili_idler = request.form.getlist("secili_id")
    if not secili_idler:
        flash("Gönderim oluşturmak için en az bir kayıt seçmelisiniz.")
        return redirect(url_for("fabrika_listesi"))
    try:
        id_listesi = sorted({int(x) for x in secili_idler})
    except ValueError:
        flash("Geçersiz kayıt seçimi.")
        return redirect(url_for("fabrika_listesi"))

    gonderim_tarihi = request.form.get("gonderim_tarihi", "").strip() or datetime.now().strftime("%Y-%m-%d")
    kargo_firmasi = request.form.get("kargo_firmasi", "").strip()
    kargo_takip_no = request.form.get("kargo_takip_no", "").strip()
    urun_tanimi = request.form.get("urun_tanimi", "").strip() or "Elektronik Kartlı Ön Ödemeli Su Sayacı"
    adres = request.form.get("adres", "").strip()
    yetkili_bayii = (request.form.get("yetkili_bayii_yeni", "").strip()
                     or request.form.get("yetkili_bayii", "").strip())

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT id FROM fabrika_tamir WHERE id = ANY(%s) AND durum = 'beklemede' AND koli_id IS NULL AND silindi_mi IS NOT TRUE "
        "ORDER BY id", (id_listesi,),
    )
    gecerli_idler = [r["id"] for r in cur.fetchall()]
    if not gecerli_idler:
        cur.close()
        flash("Seçilen kayıtlar artık 'Beklemede' durumunda değil. Lütfen listeyi yenileyip tekrar deneyin.")
        return redirect(url_for("fabrika_listesi"))

    cur.execute(
        "INSERT INTO fabrika_gonderim (kargo_firmasi, kargo_takip_no, urun_tanimi, adres, gonderim_tarihi, yetkili_bayii) "
        "VALUES (%s, %s, %s, %s, %s, %s) RETURNING id",
        (kargo_firmasi, kargo_takip_no, urun_tanimi, adres, gonderim_tarihi, yetkili_bayii or None),
    )
    gonderim_id = cur.fetchone()["id"]

    koli_no = 0
    for i in range(0, len(gecerli_idler), FABRIKA_KOLI_KAPASITESI):
        koli_no += 1
        parca = gecerli_idler[i:i + FABRIKA_KOLI_KAPASITESI]
        cur.execute(
            "INSERT INTO fabrika_koli (gonderim_id, koli_no, koli_tarihi) VALUES (%s, %s, %s) RETURNING id",
            (gonderim_id, koli_no, gonderim_tarihi),
        )
        koli_id = cur.fetchone()["id"]
        cur.execute(
            "UPDATE fabrika_tamir SET koli_id=%s, durum='gonderildi', gonderim_tarihi=%s, updated_at=NOW() "
            "WHERE id = ANY(%s)",
            (koli_id, gonderim_tarihi, parca),
        )
    db.commit()
    cur.close()
    _fabrika_gonderim_sira_numaralarini_yenile(db)
    flash(f"Gönderim oluşturuldu: {len(gecerli_idler)} sayaç, {koli_no} koli halinde paketlendi.")
    return redirect(url_for("fabrika_gonderim_detay", gonderim_id=gonderim_id))


@app.route("/fabrika/gonderimler")
@login_required
def fabrika_gonderim_listesi():
    """Tüm gönderimleri (kargo bilgisi, koli/sayaç sayısı) listeler."""
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT fg.*, "
        "(SELECT COUNT(*) FROM fabrika_koli fk WHERE fk.gonderim_id = fg.id) AS koli_sayisi, "
        "(SELECT COUNT(*) FROM fabrika_tamir ft JOIN fabrika_koli fk ON fk.id = ft.koli_id "
        " WHERE fk.gonderim_id = fg.id) AS sayac_sayisi "
        "FROM fabrika_gonderim fg ORDER BY fg.sira_no DESC"
    )
    gonderimler = cur.fetchall()
    cur.close()
    return render_template("fabrika_gonderim_listesi.html", gonderimler=gonderimler)


@app.route("/fabrika/gonderim/<int:gonderim_id>", methods=["GET", "POST"])
@login_required
def fabrika_gonderim_detay(gonderim_id):
    """Bir gönderimin kargo bilgisini gösterir/düzenler ve içindeki koli'leri listeler."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM fabrika_gonderim WHERE id = %s", (gonderim_id,))
    gonderim = cur.fetchone()
    if gonderim is None:
        cur.close()
        flash("Gönderim bulunamadı.")
        return redirect(url_for("fabrika_gonderim_listesi"))

    if request.method == "POST":
        yetkili_bayii = (request.form.get("yetkili_bayii_yeni", "").strip()
                         or request.form.get("yetkili_bayii", "").strip())
        cur.execute(
            "UPDATE fabrika_gonderim SET kargo_firmasi=%s, kargo_takip_no=%s, urun_tanimi=%s, adres=%s, "
            "yetkili_bayii=%s WHERE id=%s",
            (
                request.form.get("kargo_firmasi", "").strip(),
                request.form.get("kargo_takip_no", "").strip(),
                request.form.get("urun_tanimi", "").strip() or "Elektronik Kartlı Ön Ödemeli Su Sayacı",
                request.form.get("adres", "").strip(),
                yetkili_bayii or None,
                gonderim_id,
            ),
        )
        db.commit()
        cur.close()
        flash("Gönderim/kargo bilgisi güncellendi.")
        return redirect(url_for("fabrika_gonderim_detay", gonderim_id=gonderim_id))

    cur.execute("SELECT * FROM fabrika_koli WHERE gonderim_id = %s ORDER BY koli_no", (gonderim_id,))
    koliler = cur.fetchall()
    for koli in koliler:
        cur.execute("SELECT * FROM fabrika_tamir WHERE koli_id = %s AND silindi_mi IS NOT TRUE ORDER BY id", (koli["id"],))
        koli["kayitlar"] = cur.fetchall()
    yetkili_bayii_listesi = _fabrika_yetkili_bayii_listesi(db)
    cur.close()
    return render_template(
        "fabrika_gonderim_detay.html", gonderim=gonderim, koliler=koliler,
        yetkili_bayii_listesi=yetkili_bayii_listesi,
    )


@app.route("/fabrika/koli/<int:koli_id>/duzenle", methods=["POST"])
@login_required
def fabrika_koli_duzenle(koli_id):
    """Bir kolinin kendi tarihini ve notunu günceller."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT gonderim_id FROM fabrika_koli WHERE id = %s", (koli_id,))
    koli = cur.fetchone()
    if koli is None:
        cur.close()
        flash("Koli bulunamadı.")
        return redirect(url_for("fabrika_gonderim_listesi"))
    cur.execute(
        "UPDATE fabrika_koli SET koli_tarihi=%s, aciklama=%s WHERE id=%s",
        (
            request.form.get("koli_tarihi", "").strip() or None,
            request.form.get("aciklama", "").strip(),
            koli_id,
        ),
    )
    db.commit()
    gonderim_id = koli["gonderim_id"]
    cur.close()
    flash("Koli bilgisi güncellendi.")
    return redirect(url_for("fabrika_gonderim_detay", gonderim_id=gonderim_id))


@app.route("/fabrika/koli/<int:koli_id>/kayit-ekle", methods=["GET", "POST"])
@login_required
def fabrika_koli_kayit_ekle(koli_id):
    """Var olan bir koliye, normal 'Beklemede -> Gönderim Oluştur' akışından
    geçmeden doğrudan yeni bir tamir kaydı ekler. Özellikle yanlışlıkla
    kalıcı olarak silinmiş (Silinenler'den bile geri gelemeyen, çünkü bu
    özellik eklenmeden önce silinmiş) bir kaydı, ait olduğu koliye elle
    tekrar girmek için kullanılır."""
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT fk.*, fg.gonderim_tarihi AS gonderim_ana_tarih "
        "FROM fabrika_koli fk JOIN fabrika_gonderim fg ON fg.id = fk.gonderim_id "
        "WHERE fk.id = %s", (koli_id,),
    )
    koli = cur.fetchone()
    if koli is None:
        cur.close()
        flash("Koli bulunamadı.")
        return redirect(url_for("fabrika_gonderim_listesi"))

    if request.method == "POST":
        seri_no = request.form.get("seri_no", "").strip()
        if not seri_no:
            flash("Seri No zorunludur.")
            return redirect(url_for("fabrika_koli_kayit_ekle", koli_id=koli_id))
        yerine_takildi = request.form.get("yerine_sayac_takildi") == "takildi"
        abone_karti = request.form.get("abone_karti", "").strip()
        if abone_karti not in ("alindi", "alinmadi"):
            abone_karti = "alinmadi"
        durum = request.form.get("durum", "").strip()
        if durum not in ("gonderildi", "tamirde", "tamir_edildi", "iade_edildi"):
            durum = "gonderildi"
        gonderim_tarihi = koli["koli_tarihi"] or koli["gonderim_ana_tarih"]

        cur.execute(
            "INSERT INTO fabrika_tamir "
            "(seri_no, abone_adi, koy_adi, telefon, ilk_montaj_tarihi, uretim_yili, "
            "tespit_edilen_ariza, yerine_sayac_takildi, takilan_sayac_serisi, durum, "
            "donus_tarihi, gonderim_tarihi, tamir_ucreti, parca_maliyeti, odeyen, "
            "abone_karti, koli_id, olusturan_kullanici) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (
                seri_no,
                request.form.get("abone_adi", "").strip(),
                request.form.get("koy_adi", "").strip(),
                request.form.get("telefon", "").strip(),
                request.form.get("ilk_montaj_tarihi", "").strip() or None,
                request.form.get("uretim_yili", "").strip(),
                request.form.get("tespit_edilen_ariza", "").strip(),
                yerine_takildi,
                request.form.get("takilan_sayac_serisi", "").strip() if yerine_takildi else "",
                durum,
                request.form.get("donus_tarihi", "").strip() or None,
                gonderim_tarihi,
                _sayi_veya(request.form.get("tamir_ucreti"), 0),
                _sayi_veya(request.form.get("parca_maliyeti"), 0),
                request.form.get("odeyen", "").strip(),
                abone_karti,
                koli_id,
                session.get("kullanici_adi", ""),
            ),
        )
        db.commit()
        cur.close()
        _fabrika_tamir_sira_numaralarini_yenile(db)
        flash(f"Kayıt Koli {koli['koli_no']} içine eklendi.")
        return redirect(url_for("fabrika_gonderim_detay", gonderim_id=koli["gonderim_id"]))

    cur.close()
    return render_template(
        "fabrika_koli_kayit_ekle.html", koli=koli, durum_etiketleri=FABRIKA_DURUM_ETIKETLERI,
    )


@app.route("/fabrika/gonderim/<int:gonderim_id>/rapor")
@login_required
def fabrika_rapor(gonderim_id):
    """Sayaç Durum Raporu — yazdırılabilir HTML çıktısı."""
    db = get_db()
    gonderim, koliler, satici = _fabrika_rapor_verisi(db, gonderim_id)
    if gonderim is None:
        flash("Gönderim bulunamadı.")
        return redirect(url_for("fabrika_gonderim_listesi"))
    return render_template(
        "fabrika_rapor.html", gonderim=gonderim, koliler=koliler, satici=satici,
        koli_kapasitesi=FABRIKA_KOLI_KAPASITESI,
    )


@app.route("/fabrika/gonderim/<int:gonderim_id>/rapor.pdf")
@login_required
def fabrika_rapor_pdf(gonderim_id):
    """Sayaç Durum Raporu'nu indirilebilir PDF olarak üretir (reportlab)."""
    if SimpleDocTemplate is None:
        flash("PDF oluşturma bileşeni (reportlab) sunucuda kurulu değil. Yazdırılabilir ekran görünümünü kullanabilirsiniz.")
        return redirect(url_for("fabrika_rapor", gonderim_id=gonderim_id))
    db = get_db()
    gonderim, koliler, satici = _fabrika_rapor_verisi(db, gonderim_id)
    if gonderim is None:
        flash("Gönderim bulunamadı.")
        return redirect(url_for("fabrika_gonderim_listesi"))
    pdf_bayt = _fabrika_rapor_pdf_olustur(gonderim, koliler, satici)
    return Response(
        pdf_bayt, mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sayac-durum-raporu-{gonderim_id}.pdf"},
    )


@app.route("/fabrika/gonderim/<int:gonderim_id>/sil", methods=["POST"])
@login_required
def fabrika_gonderim_sil(gonderim_id):
    """Bir gönderim kaydını (kargo/koli bilgisiyle birlikte) siler."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id FROM fabrika_gonderim WHERE id = %s", (gonderim_id,))
    if cur.fetchone() is None:
        cur.close()
        flash("Gönderim bulunamadı.")
        return redirect(url_for("fabrika_gonderim_listesi"))
    cur.execute(
        "UPDATE fabrika_tamir SET koli_id = NULL, durum = 'beklemede' "
        "WHERE koli_id IN (SELECT id FROM fabrika_koli WHERE gonderim_id = %s)",
        (gonderim_id,),
    )
    cur.execute("DELETE FROM fabrika_gonderim WHERE id = %s", (gonderim_id,))
    db.commit()
    cur.close()
    _fabrika_gonderim_sira_numaralarini_yenile(db)
    flash("Gönderim kaydı silindi.")
    return redirect(url_for("fabrika_gonderim_listesi"))


@app.route("/yedek-al")
@login_required
def yedek_al():
    db = get_db()
    cur = db.cursor()
    parcalar = []
    for t in YEDEKLENECEK_TABLOLAR:
        cur.execute(f"SELECT * FROM {t}")
        for row in cur.fetchall():
            parcalar.append(cur.mogrify(f"INSERT INTO {t} VALUES %s;", (tuple(row.values()),)).decode())
    cur.close()
    icerik = "\n".join(parcalar)
    tarih = datetime.now().strftime("%d_%m_%Y")
    dosya_adi = f"algi_bilisim_yedek_{tarih}.sql"

    alici_eposta = _ayar_getir(db, "yedek_alici_eposta")
    if alici_eposta:
        basarili, hata = _yedek_eposta_gonder(icerik.encode("utf-8"), dosya_adi, alici_eposta)
        if basarili:
            flash(f"Yedek dosyası ayrıca {alici_eposta} adresine e-posta ile gönderildi.")
        else:
            flash(f"Yedek dosyası indirildi, ancak e-posta gönderilemedi: {hata}")

    return Response(
        icerik,
        mimetype="text/plain",
        headers={"Content-Disposition": f"attachment; filename={dosya_adi}"},
    )


def _csv_olustur(kolon_listesi, goster_kolonlari, satirlar, dosya_adi):
    cikti = io.StringIO()
    yazici = csv.writer(cikti, delimiter=';')
    basliklar = [etiket for anahtar, etiket in kolon_listesi if anahtar in goster_kolonlari]
    yazici.writerow(basliklar)
    for s in satirlar:
        satir = [s[anahtar] for anahtar, etiket in kolon_listesi if anahtar in goster_kolonlari]
        yazici.writerow(satir)
    icerik = "\ufeff" + cikti.getvalue()
    return Response(
        icerik,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={dosya_adi}"},
    )


def _abone_kaynak_paketle(satirlar):
    ilk = satirlar[0]
    digerleri = [
        {"id": s["id"], "adi": s["adi"] or "", "soyadi": s["soyadi"] or "", "koy_adi": s["koy_adi"] or ""}
        for s in satirlar
    ]
    return {
        "bulundu": True,
        "adi": ilk["adi"] or "", "soyadi": ilk["soyadi"] or "",
        "telefon": ilk["telefon"] or "", "telefon2": ilk["telefon2"] or "",
        "koy_adi": ilk["koy_adi"] or "",
        "montaj_tarihi": _tarih_iso_hale_getir(ilk["montaj_tarihi"]),
        "coklu": len(satirlar) > 1,
        "digerleri": digerleri,
        "kaynak": "abone",
    }


def _koy_kaynak_paketle(satirlar):
    ilk = satirlar[0]
    digerleri = [
        {"id": s["id"], "adi": s["adi"] or "", "soyadi": s["soyadi"] or "", "koy_adi": s["koy_adi"] or ""}
        for s in satirlar
    ]
    return {
        "bulundu": True,
        "adi": ilk["adi"] or "", "soyadi": ilk["soyadi"] or "",
        "telefon": "", "telefon2": "",
        "koy_adi": ilk["koy_adi"] or "",
        "montaj_tarihi": _tarih_iso_hale_getir(ilk["abonelik_tarihi"]),
        "coklu": len(satirlar) > 1,
        "digerleri": digerleri,
        "kaynak": "koy_listesi",
    }


def _isim_anahtari(adi, soyadi):
    return ((adi or "").strip().upper(), (soyadi or "").strip().upper())


def _tum_adaylari_olustur(abone_satirlari, koy_satirlari):
    """Hem 'abone' tablosundaki, hem de 'koy_abone' tablosundaki tüm satırları
    tarayıp isme göre tekilleştirilmiş TÜM farklı kişi adaylarının listesini döndürür."""
    adaylar = []
    gorulenler = set()
    for s in abone_satirlari:
        anahtar = _isim_anahtari(s["adi"], s["soyadi"])
        if anahtar in gorulenler:
            continue
        gorulenler.add(anahtar)
        adaylar.append({
            "kaynak": "abone",
            "adi": s["adi"] or "", "soyadi": s["soyadi"] or "",
            "telefon": s["telefon"] or "", "telefon2": s["telefon2"] or "",
            "koy_adi": s["koy_adi"] or "",
            "montaj_tarihi": _tarih_iso_hale_getir(s["montaj_tarihi"]),
        })
    for s in koy_satirlari:
        anahtar = _isim_anahtari(s["adi"], s["soyadi"])
        if anahtar in gorulenler:
            continue
        gorulenler.add(anahtar)
        adaylar.append({
            "kaynak": "koy_listesi",
            "adi": s["adi"] or "", "soyadi": s["soyadi"] or "",
            "telefon": "", "telefon2": "",
            "koy_adi": s["koy_adi"] or "",
            "montaj_tarihi": _tarih_iso_hale_getir(s["abonelik_tarihi"]),
        })
    return adaylar


@app.route("/api/abone-ara")
@login_required
def abone_ara():
    sayac_no = request.args.get("sayac_no", "").strip()
    if not sayac_no:
        return jsonify({"bulundu": False})
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT id, adi, soyadi, telefon, telefon2, koy_adi, montaj_tarihi FROM abone WHERE sayac_no = %s ORDER BY id",
        (sayac_no,),
    )
    abone_satirlari = cur.fetchall()

    cur.execute(
        "SELECT id, adi, soyadi, koy_adi, abonelik_tarihi FROM koy_abone WHERE cihaz_no = %s ORDER BY id",
        (sayac_no,),
    )
    koy_satirlari = cur.fetchall()
    cur.close()

    if not abone_satirlari and not koy_satirlari:
        return jsonify({"bulundu": False, "digerleri": []})

    abone_paket = _abone_kaynak_paketle(abone_satirlari) if abone_satirlari else None
    koy_paket = _koy_kaynak_paketle(koy_satirlari) if koy_satirlari else None
    birincil = abone_paket or koy_paket

    adaylar = _tum_adaylari_olustur(abone_satirlari, koy_satirlari)

    sonuc = dict(birincil)
    sonuc["secenekler"] = adaylar if len(adaylar) > 1 else []
    return jsonify(sonuc)


_SESLI_SORGU_NITELIKLER = [
    # (anahtar kelime öbekleri — en uzun/özelinden en genele), [(tablo, kolon ifadesi, etiket, biçim), ...]
    (['telefon numarası', 'telefon numarasi', 'telefonu', 'telefon'],
        [('abone', 'telefon', 'Telefon', None), ('ariza', 'telefon', 'Telefon', None)]),
    (['ikinci telefon', 'telefon iki', 'telefon 2'],
        [('abone', 'telefon2', 'Telefon 2', None), ('ariza', 'telefon2', 'Telefon 2', None)]),
    (['adresi', 'adres'],
        [('abone', 'adres', 'Adres', None), ('ariza', 'adres', 'Adres', None)]),
    (['baba adı', 'baba adi'],
        [('abone', 'baba_adi', 'Baba Adı', None)]),
    (['sayaç numarası', 'sayac numarasi', 'sayaç no', 'sayac no'],
        [('abone', 'sayac_no', 'Sayaç No', None)]),
    (['seri numarası', 'seri numarasi', 'seri no'],
        [('ariza', 'seri_no', 'Seri No', None)]),
    (['senet numarası', 'senet numarasi', 'senet no'],
        [('abone', 'senet_no', 'Senet No', None)]),
    (['fatura numarası', 'fatura numarasi', 'fatura no'],
        [('abone', 'fatura_no', 'Fatura No', None)]),
    (['kalan bakiyesi', 'bakiyesi kalan', 'kalan bakiye', 'kalan borcu', 'borcu kalan',
      'bakiyesi', 'borcu', 'bakiye', 'kalan borç', 'kalan'],
        [('abone', '(sayac_tutari - alinan_tutar)', 'Kalan Bakiye', 'tl'),
         ('ariza', '(ariza_ucret - alinan_ucret)', 'Kalan Bakiye', 'tl')]),
    (['sayaç tutarı', 'sayac tutari', 'sayaç ücreti', 'sayac ucreti'],
        [('abone', 'sayac_tutari', 'Sayaç Tutarı', 'tl')]),
    (['alınan tutar', 'alinan tutar', 'alınan ücret', 'alinan ucret', 'alınan', 'alinan'],
        [('abone', 'alinan_tutar', 'Alınan Tutar', 'tl'), ('ariza', 'alinan_ucret', 'Alınan Ücret', 'tl')]),
    (['malzeme kalanı', 'malzeme kalani'],
        [('abone', '(malzeme_tutari - malzeme_alinan)', 'Malzeme Kalan', 'tl')]),
    (['ödeme tarihi', 'odeme tarihi'],
        [('abone', 'odeme_tarihi', 'Ödeme Tarihi', None)]),
    (['ödeme şekli', 'odeme sekli'],
        [('abone', 'odeme_sekli', 'Ödeme Şekli', None)]),
    (['montaj tarihi'],
        [('abone', 'montaj_tarihi', 'Montaj Tarihi', None)]),
    (['geliş tarihi', 'gelis tarihi'],
        [('ariza', 'gelis_tarihi', 'Geliş Tarihi', None)]),
    (['tespit edilen arıza', 'tespit edilen ariza', 'arızası', 'arizasi'],
        [('ariza', 'tespit_edilen_ariza', 'Tespit Edilen Arıza', None)]),
    (['yapılan işlemler', 'yapilan islemler', 'yapılan işlem', 'yapilan islem'],
        [('ariza', 'yapilan_islemler', 'Yapılan İşlemler', None)]),
    (['açıklaması', 'aciklamasi', 'açıklama', 'aciklama'],
        [('abone', 'aciklama', 'Açıklama', None)]),
]


@app.route("/api/sesli-sorgu", methods=["POST"])
@login_required
def sesli_sorgu_api():
    """Genel sesli mikrofonun, hiçbir form alanı/düğmesiyle eşleşmeyen bir
    söylemi (örn. \"Ahmet Yılmaz'ın telefon numarası\" ya da \"Sarıhasanlı
    köyü kalan bakiyesi\") DOĞAL BİR SORU olarak yorumlamaya çalıştığı uç
    nokta. Basit bir örüntü tanıma kullanır (apostrof/olmuş kelime öbeği +
    isim ya da köy adı ayrıştırma) — tam bir yapay zekâ değildir, bu yüzden
    her ifadeyi anlayamayabilir."""
    veri = request.get_json(silent=True) or {}
    sorgu_ham = (veri.get("sorgu") or "").strip()
    if not sorgu_ham:
        return jsonify({"tur": "bulunamadi", "mesaj": "Boş sorgu"})

    db = get_db()
    cur = db.cursor()
    sorgu_norm = _turkce_normallestir(sorgu_ham)

    # --- Köy modu: "... köy(ü) ..." kalıbı -----------------------------------
    koy_eslesme = re.search(r'\bköyü?\b', sorgu_norm)
    if not koy_eslesme:
        # normalize edilmiş metinde 'ö' de katlanmış olabileceğinden (bkz.
        # _turkce_normallestir SADECE İ/I/ı/i, Ç/Ş/Ğ/Ö/Ü içindir; 'köy'
        # zaten Latin harflerle yazılabildiği için katlanmıyor) — yine de
        # güvenlik amaçlı iki biçimi de deniyoruz.
        koy_eslesme = re.search(r'\bkoyu?\b', sorgu_norm)
    if koy_eslesme:
        koy_adi_ham = sorgu_ham[:koy_eslesme.start()].strip(" '").strip()
        kalan_metin_norm = sorgu_norm[koy_eslesme.end():].strip()
        if not koy_adi_ham:
            cur.close()
            return jsonify({"tur": "bulunamadi", "mesaj": "Köy adı anlaşılamadı"})
        sadece_kalanlar = any(k in kalan_metin_norm for k in ("kalan", "bakiye", "borc"))
        sql = (
            "SELECT id, adi, soyadi, koy_adi, (sayac_tutari - alinan_tutar) AS kalan "
            "FROM abone WHERE " + _turkce_esle_kosul("koy_adi") + " LIKE %s"
        )
        params = [_turkce_normallestir(f"%{koy_adi_ham}%")]
        if sadece_kalanlar:
            sql += " AND (sayac_tutari - alinan_tutar) > 0"
        sql += " ORDER BY (sayac_tutari - alinan_tutar) DESC LIMIT 50"
        cur.execute(sql, params)
        satirlar = cur.fetchall()
        cur.close()
        if not satirlar:
            return jsonify({"tur": "bulunamadi", "mesaj": f'"{koy_adi_ham}" için kayıt bulunamadı'})
        sonuclar = [{
            "baslik": f"{s['adi']} {s['soyadi']}",
            "alt": f"{s['koy_adi']} — Kalan: {tl_format(s['kalan'])}",
            "url": url_for("abone_duzenle", abone_id=s["id"]),
        } for s in satirlar]
        return jsonify({
            "tur": "liste",
            "baslik": (koy_adi_ham + " — Kalan Bakiyesi Olanlar") if sadece_kalanlar else (koy_adi_ham + " Köyü"),
            "sonuclar": sonuclar,
        })

    # --- Kişi modu -----------------------------------------------------------
    # Konuşma tanıma, Türkçe imla kuralı gereği özel isim + iyelik ekini
    # genelde bir kesme işaretiyle ayırır (örn. "Ahmet Yılmaz'ın telefonu").
    # Bu işaret varsa isim/nitelik sınırı olarak öncelikli kullanılır. Ayrıca
    # konuşma tanıma bazen düz (') yerine eğik (' ' gibi) bir kesme işareti
    # üretebiliyor, o yüzden birkaç türü birden deniyoruz.
    isim_kismi_ham = None
    nitelik_metin_norm = sorgu_norm
    kesme_isaretleri = ("'", "\u2019", "\u2018")
    kesme_konumu = -1
    for isaret in kesme_isaretleri:
        pos = sorgu_ham.find(isaret)
        if pos != -1 and (kesme_konumu == -1 or pos < kesme_konumu):
            kesme_konumu = pos
    if kesme_konumu != -1:
        isim_kismi_ham = sorgu_ham[:kesme_konumu].strip()
        kalan_norm = _turkce_normallestir(sorgu_ham[kesme_konumu + 1:])
        kalan_kelimeler = kalan_norm.split(None, 1)  # ilk kelime = iyelik eki (nın/nin/ın/in vb.), atılır
        nitelik_metin_norm = kalan_kelimeler[1] if len(kalan_kelimeler) > 1 else ""

    def _nitelik_bul(metin_norm):
        """Bir nitelik öbeğinin (örn. "kalan borcu") metinde geçip
        geçmediğini üç farklı şekilde dener: düz alt-dize, kelime kümesi
        (sıra/araya başka kelime girmesi önemsiz) ve bitişik yazım
        (konuşma tanımanın kelimeleri birleştirmesi durumunda). En uzun
        (en özel) eşleşen öbeği döndürür."""
        kelime_kumesi = set(metin_norm.split())
        metin_bitisik = metin_norm.replace(" ", "")
        en_iyi = None
        en_iyi_uzunluk = -1
        for anahtarlar, hedefler in _SESLI_SORGU_NITELIKLER:
            for anahtar in anahtarlar:
                anahtar_norm = _turkce_normallestir(anahtar)
                anahtar_kelimeleri = anahtar_norm.split()
                eslesti = (
                    anahtar_norm in metin_norm
                    or anahtar_norm.replace(" ", "") in metin_bitisik
                    or all(k in kelime_kumesi for k in anahtar_kelimeleri)
                )
                if eslesti and len(anahtar_norm) > en_iyi_uzunluk:
                    en_iyi = (anahtar, hedefler)
                    en_iyi_uzunluk = len(anahtar_norm)
        return en_iyi

    bulunan_nitelik = _nitelik_bul(nitelik_metin_norm)

    if isim_kismi_ham is None:
        # Kesme işareti yoksa: bulunan niteliğin kelimelerini metinden tek
        # tek çıkarıp geri kalanını isim/konu olarak kullan (nitelik
        # kelimeleri bitişik ya da araya başka kelime girmiş olsa bile).
        if bulunan_nitelik:
            kalan_kelimeler = sorgu_norm.split()
            for k in _turkce_normallestir(bulunan_nitelik[0]).split():
                if k in kalan_kelimeler:
                    kalan_kelimeler.remove(k)
            isim_kismi_ham = " ".join(kalan_kelimeler)
        else:
            isim_kismi_ham = sorgu_ham

    if not bulunan_nitelik:
        cur.close()
        return jsonify({"tur": "bulunamadi", "mesaj": f'Anlaşılamadı: "{sorgu_ham}"'})

    isim_kelimeler = [k for k in _turkce_normallestir(isim_kismi_ham).split() if k]
    if not isim_kelimeler:
        cur.close()
        return jsonify({"tur": "bulunamadi", "mesaj": f'İsim anlaşılamadı: "{sorgu_ham}"'})

    _anahtar, hedefler = bulunan_nitelik
    sonuc_listesi = []
    for tablo, kolon_ifadesi, etiket, bicim in hedefler:
        ad_ifadesi = "(COALESCE(adi, '') || ' ' || COALESCE(soyadi, ''))"
        kosul_parcalari = []
        params = []
        for k in isim_kelimeler:
            kosul_parcalari.append(_turkce_esle_kosul(ad_ifadesi) + " LIKE %s")
            params.append(_turkce_normallestir(f"%{k}%"))
        sql = (
            f"SELECT id, adi, soyadi, koy_adi, {kolon_ifadesi} AS deger FROM {tablo} "
            "WHERE " + " AND ".join(kosul_parcalari) + " LIMIT 20"
        )
        cur.execute(sql, params)
        for satir in cur.fetchall():
            deger = satir["deger"]
            if bicim == "tl":
                deger_metin = tl_format(deger) if deger is not None else "-"
            elif deger in (None, ""):
                deger_metin = "(boş)"
            else:
                deger_metin = str(deger)
            url_fonk = "abone_duzenle" if tablo == "abone" else "ariza_duzenle"
            id_param = "abone_id" if tablo == "abone" else "ariza_id"
            sonuc_listesi.append({
                "baslik": f"{satir['adi']} {satir['soyadi']}",
                "alt": f"{etiket}: {deger_metin}" + (f" ({satir['koy_adi']})" if satir["koy_adi"] else ""),
                "url": url_for(url_fonk, **{id_param: satir["id"]}),
            })
    cur.close()

    if not sonuc_listesi:
        return jsonify({"tur": "bulunamadi", "mesaj": f'"{isim_kismi_ham}" için kayıt bulunamadı (Anlaşılan: "{sorgu_ham}")'})
    return jsonify({"tur": "liste", "baslik": bulunan_nitelik[1][0][2], "sonuclar": sonuc_listesi})


@app.route("/api/kolon-secenekleri")
@login_required
def kolon_secenekleri_api():
    """Liste sayfalarındaki sütun filtre kutuları için, kullanıcı bir sütunun
    filtre kutusunu AÇTIĞINDA seçenekler bu uç nokta üzerinden istenir."""
    tablo = request.args.get("tablo", "").strip()
    anahtar = request.args.get("anahtar", "").strip()
    if tablo not in ("abone", "ariza", "fabrika_tamir", "fatura"):
        return jsonify({"hata": "geçersiz tablo"}), 400
    db = get_db()
    if tablo == "abone":
        _kolon_listesi, bilgi_sozlugu, _sayisal, _ozel = _abone_kolon_takimi(db)
    elif tablo == "ariza":
        _kolon_listesi, bilgi_sozlugu, _sayisal, _ozel = _ariza_kolon_takimi(db)
    elif tablo == "fabrika_tamir":
        _kolon_listesi, bilgi_sozlugu, _sayisal, _ozel = _fabrika_kolon_takimi(db)
    else:
        _kolon_listesi, bilgi_sozlugu, _sayisal, _ozel = _fatura_kolon_takimi(db)
    if anahtar not in bilgi_sozlugu:
        return jsonify({"hata": "geçersiz sütun"}), 400
    if tablo == "abone":
        ekstra_kosul, ekstra_params = _abone_filtre_kosulu_olustur(anahtar, bilgi_sozlugu)
    elif tablo == "ariza":
        ekstra_kosul, ekstra_params = _ariza_filtre_kosulu_olustur(anahtar, bilgi_sozlugu)
    elif tablo == "fabrika_tamir":
        ekstra_kosul, ekstra_params = _fabrika_filtre_kosulu_olustur(anahtar, bilgi_sozlugu)
    else:
        ekstra_kosul, ekstra_params = _fatura_filtre_kosulu_olustur(anahtar, bilgi_sozlugu)
    gercek_tablo = _FATURA_ALT_SORGU if tablo == "fatura" else tablo
    secenekler = _kolon_secenekleri(db, anahtar, gercek_tablo, bilgi_sozlugu, ekstra_kosul, ekstra_params)
    return jsonify({"secenekler": secenekler})


@app.route("/api/ariza-gecmisi")
@login_required
def ariza_gecmisi():
    """Yeni arıza kaydı açılırken, girilen seri no'ya ait daha önce oluşturulmuş
    arıza kayıtlarını (varsa) döndürür."""
    seri_no = request.args.get("seri_no", "").strip()
    if not seri_no:
        return jsonify({"bulundu": False, "kayitlar": []})
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT id, gelis_tarihi, takilan_tarih, tespit_edilen_ariza, yapilan_islemler, "
        "ariza_ucret, alinan_ucret FROM ariza WHERE seri_no = %s ORDER BY id DESC",
        (seri_no,),
    )
    satirlar = cur.fetchall()
    cur.close()

    kayitlar = []
    for s in satirlar:
        kalan = (s["ariza_ucret"] or 0) - (s["alinan_ucret"] or 0)
        kayitlar.append({
            "id": s["id"],
            "duzenle_url": url_for("ariza_duzenle", ariza_id=s["id"]),
            "gelis_tarihi": _gg_aa_yyyy(s["gelis_tarihi"]),
            "takilan_tarih": _gg_aa_yyyy(s["takilan_tarih"]),
            "tespit_edilen_ariza": s["tespit_edilen_ariza"] or "",
            "yapilan_islemler": s["yapilan_islemler"] or "",
            "kalan_ucret": tl_format(kalan),
        })
    return jsonify({"bulundu": len(kayitlar) > 0, "kayitlar": kayitlar})


@app.route("/abone")
@login_required
def abone_listesi():
    yonlendirme = _filtre_durumu_uygula("abone_listesi")
    if yonlendirme:
        return yonlendirme

    q = request.args.get("q", "").strip()
    koy = request.args.get("koy", "").strip()
    alanlar_secili = request.args.getlist("alan")
    db = get_db()

    ALAN_TANIMLARI = _ABONE_ALAN_TANIMLARI
    ALAN_HARITASI = _ABONE_ALAN_HARITASI
    alan_listesi = [(k, etiket) for k, etiket, _, _ in ALAN_TANIMLARI]

    sql = "SELECT * FROM abone WHERE 1=1"
    params = []
    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in ALAN_TANIMLARI]

        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None

        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in ALAN_HARITASI:
                kolon, sayisal = ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            sql += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params
    if koy:
        sql += " AND koy_adi = %s"
        params.append(koy)

    kolon_listesi, kolon_bilgi, sayisal_kolonlar, ozel_alanlar = _abone_kolon_takimi(db)
    deger_secili = {}
    haric_secili = {}
    for anahtar, _ in kolon_listesi:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")
        haric_secili[anahtar] = request.args.getlist(f"haric_{anahtar}")
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi

    sql += f" ORDER BY s_no {'DESC' if _sira_yonu_al() == 'desc' else 'ASC'}"

    cur = db.cursor()
    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()
    cur.execute("SELECT DISTINCT koy_adi FROM abone ORDER BY koy_adi")
    koyler = cur.fetchall()
    cur.execute("SELECT COUNT(*) AS c FROM abone")
    toplam_kayit = cur.fetchone()["c"]

    toplam_ucret = sum(float(k["sayac_tutari"] or 0) + float(k["malzeme_tutari"] or 0) for k in kayitlar_ham)
    tahsil_edilen_ucret = sum(float(k["alinan_tutar"] or 0) + float(k["malzeme_alinan"] or 0) for k in kayitlar_ham)
    kalan_bakiye = toplam_ucret - tahsil_edilen_ucret

    bugun_iso = datetime.now().strftime("%Y-%m-%d")
    cur.execute(
        """
        SELECT id, koy_adi, adi, soyadi, odeme_gun_sozu,
               (sayac_tutari + malzeme_tutari - alinan_tutar - malzeme_alinan) AS toplam_kalan
        FROM abone
        WHERE odeme_gun_sozu IS NOT NULL AND odeme_gun_sozu <> ''
          AND odeme_gun_sozu <= %s
          AND (sayac_tutari + malzeme_tutari - alinan_tutar - malzeme_alinan) > 0
        ORDER BY odeme_gun_sozu ASC
        """,
        (bugun_iso,),
    )
    odeme_hatirlatmalari = [
        {
            "id": r["id"],
            "koy_adi": r["koy_adi"],
            "adi": r["adi"],
            "soyadi": r["soyadi"],
            "odeme_gun_sozu": _gg_aa_yyyy(r["odeme_gun_sozu"]),
            "toplam_kalan": tl_format(r["toplam_kalan"]),
        }
        for r in cur.fetchall()
    ]

    satirlar = [_abone_satir_sozlugu(k, ozel_alanlar) for k in kayitlar_ham]

    cur.close()

    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "abone_list.html", satirlar=satirlar, koyler=koyler, q=q, secili_koy=koy,
        secili_alanlar=alanlar_secili, alan_listesi=alan_listesi,
        kolon_listesi=kolon_listesi, deger_secili=deger_secili, haric_secili=haric_secili,
        sayisal_kolonlar=sayisal_kolonlar,
        arama_satir=_izgara_satir(len(alan_listesi)),
        arama_satir_2=_izgara_satir(len(alan_listesi), 2),
        filtreli_kayit=filtreli_kayit, toplam_kayit=toplam_kayit,
        odeme_hatirlatmalari=odeme_hatirlatmalari,
        toplam_ucret=toplam_ucret, tahsil_edilen_ucret=tahsil_edilen_ucret, kalan_bakiye=kalan_bakiye,
        sira=_sira_yonu_al(), sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
    )


def _montaj_formu_secili_id(db, istenen_id):
    """İstenen sablon_id gerçekten var mı diye kontrol eder; yoksa kayıtlı ilk tasarımın id'sine döner."""
    sablonlar = _montaj_formu_sablonlar_listele(db)
    if not sablonlar:
        return None
    if istenen_id:
        for s in sablonlar:
            if s["id"] == istenen_id:
                return istenen_id
    return sablonlar[0]["id"]


@app.route("/montaj-formu/tasarim", methods=["GET", "POST"])
@login_required
def montaj_formu_tasarim():
    db = get_db()

    if request.method == "POST":
        sablon_id = request.form.get("sablon_id", type=int)
        yeni_icerik = request.form.get("icerik", "")
        _montaj_formu_sablon_kaydet(db, sablon_id, yeni_icerik)
        flash("Montaj Formu tasarımı kaydedildi.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    secili_id = _montaj_formu_secili_id(db, request.args.get("sablon_id", type=int))
    sablon = _montaj_formu_sablon_getir(db, secili_id)
    sablonlar = _montaj_formu_sablonlar_listele(db)

    ornek_satir = {
        "adi": "AHMET", "soyadi": "YILMAZ", "koy_adi": "ÖRNEK KÖYÜ",
        "sayac_no": "12345678", "telefon": "0555 555 55 55", "telefon2": "",
        "montaj_tarihi": datetime.now().strftime("%d.%m.%Y"),
        "sayac_tutari": "1.500,00", "alinan_tutar": "1.500,00",
        "malzeme_tutari": "750,00", "malzeme_alinan": "750,00",
        "montaj_personeli": "ÖRNEK PERSONEL",
        "_montaj_imza_veri_url": _ORNEK_IMZA_YER_TUTUCU,
        "_abone_imza_veri_url": _ORNEK_IMZA_YER_TUTUCU,
    }
    onizleme_html, onizleme_hata = _montaj_formu_render_tek(sablon["icerik"], ornek_satir)

    return render_template(
        "montaj_formu_tasarim.html",
        sablon=sablon, sablonlar=sablonlar,
        icerik=sablon["icerik"], onizleme_html=onizleme_html, onizleme_hata=onizleme_hata,
    )


@app.route("/montaj-formu/tasarim/yeni", methods=["POST"])
@login_required
def montaj_formu_tasarim_yeni():
    """Boş/varsayılan içerikle yeni, isimli bir Montaj Formu tasarımı oluşturur."""
    db = get_db()
    ad = (request.form.get("ad") or "").strip() or "Yeni Tasarım"
    yeni_id = _montaj_formu_sablon_olustur(db, ad, _MONTAJ_FORMU_VARSAYILAN_SABLON)
    flash(f'"{ad}" adında yeni bir Montaj Formu tasarımı oluşturuldu.')
    return redirect(url_for("montaj_formu_tasarim", sablon_id=yeni_id))


@app.route("/montaj-formu/tasarim/yeniden-adlandir", methods=["POST"])
@login_required
def montaj_formu_tasarim_yeniden_adlandir():
    db = get_db()
    sablon_id = request.form.get("sablon_id", type=int)
    yeni_ad = (request.form.get("ad") or "").strip()
    if sablon_id and yeni_ad:
        _montaj_formu_sablon_yeniden_adlandir(db, sablon_id, yeni_ad)
        flash("Tasarımın adı güncellendi.")
    return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))


@app.route("/montaj-formu/tasarim/sil", methods=["POST"])
@login_required
def montaj_formu_tasarim_sil():
    """Bir tasarımı siler."""
    db = get_db()
    sablon_id = request.form.get("sablon_id", type=int)
    sablonlar = _montaj_formu_sablonlar_listele(db)
    if len(sablonlar) <= 1:
        flash("Son kalan Montaj Formu tasarımı silinemez — en az bir tasarım kayıtlı olmalı.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))
    if sablon_id:
        _montaj_formu_sablon_sil(db, sablon_id)
        flash("Tasarım silindi.")
    return redirect(url_for("montaj_formu_tasarim"))


@app.route("/montaj-formu/sablonlar.json")
@login_required
def montaj_formu_sablonlar_json():
    """M.Form penceresinde 'hangi tasarımla açmak istiyorsun' listesini doldurur."""
    db = get_db()
    sablonlar = _montaj_formu_sablonlar_listele(db)
    return jsonify([{"id": s["id"], "ad": s["ad"]} for s in sablonlar])


@app.route("/montaj-formu/tasarim/sifirla", methods=["POST"])
@login_required
def montaj_formu_tasarim_sifirla():
    """Seçili Montaj Formu tasarımını, koddaki güncel varsayılan tasarıma sıfırlar."""
    db = get_db()
    sablon_id = request.form.get("sablon_id", type=int)
    _montaj_formu_sablon_kaydet(db, sablon_id, _MONTAJ_FORMU_VARSAYILAN_SABLON)
    flash("Montaj Formu tasarımı, programın güncel varsayılan tasarımına sıfırlandı.")
    return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))


@app.route("/montaj-formu/tasarim/dosya-yukle", methods=["POST"])
@login_required
def montaj_formu_tasarim_dosya_yukle():
    """Montaj Formu tasarımını, .html dosyası yükleyerek değiştirir."""
    sablon_id = request.form.get("sablon_id", type=int)
    dosya = request.files.get("sablon_dosyasi")
    if dosya is None or not dosya.filename:
        flash("Lütfen yüklemek için bir tasarım dosyası (.html) seçin.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    try:
        yeni_icerik = dosya.read().decode("utf-8")
    except UnicodeDecodeError:
        flash("Dosya okunamadı — lütfen UTF-8 kodlamalı bir .html dosyası yükleyin.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    _, hata = _montaj_formu_render_tek(yeni_icerik, _MONTAJ_FORMU_TEST_VERISI)
    if hata:
        flash(f"Yüklenen dosyada bir hata var, tasarım kaydedilmedi: {hata}")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))
    db = get_db()
    _montaj_formu_sablon_kaydet(db, sablon_id, yeni_icerik)
    flash("Montaj Formu tasarımı, yüklediğiniz dosyadan güncellendi.")
    return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))


@app.route("/montaj-formu/tasarim/word-yukle", methods=["POST"])
@login_required
def montaj_formu_tasarim_word_yukle():
    """Montaj Formu tasarımını Word (.docx) belgesinden yükleyerek değiştirir."""
    sablon_id = request.form.get("sablon_id", type=int)
    if mammoth is None:
        flash("Word'den tasarım yükleme özelliği şu anda kullanılamıyor (sunucu tarafında "
              "'mammoth' kütüphanesi kurulu değil). Lütfen bizimle iletişime geçin.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    dosya = request.files.get("sablon_word")
    if dosya is None or not dosya.filename:
        flash("Lütfen yüklemek için bir Word belgesi (.docx) seçin.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    try:
        sonuc = mammoth.convert_to_html(dosya)
        yeni_icerik = _word_tasarimini_onar(sonuc.value)
    except Exception as e:
        flash(f"Word belgesi okunamadı: {e}")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    if not yeni_icerik.strip():
        flash("Word belgesinden hiçbir içerik okunamadı — lütfen dosyayı kontrol edin.")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    _, hata = _montaj_formu_render_tek(yeni_icerik, _MONTAJ_FORMU_TEST_VERISI)
    if hata:
        flash(f"Word belgesinden dönüştürülen tasarımda bir hata var, tasarım "
              f"kaydedilmedi: {hata}")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))

    db = get_db()
    _montaj_formu_sablon_kaydet(db, sablon_id, yeni_icerik)
    uyari_sayisi = len(getattr(sonuc, "messages", []) or [])
    if uyari_sayisi:
        flash(f"Montaj Formu tasarımı, Word belgesinden güncellendi ({uyari_sayisi} "
              f"küçük dönüşüm notu var — önizlemeyi kontrol edin).")
    else:
        flash("Montaj Formu tasarımı, Word belgesinden güncellendi.")
    return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon_id))


@app.route("/abone/<int:abone_id>/montaj-formu")
@login_required
def abone_montaj_formu(abone_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Abone bulunamadı.")
        return redirect(url_for("abone_listesi"))

    sablon_id = _montaj_formu_secili_id(db, request.args.get("sablon_id", type=int))
    sablon = _montaj_formu_sablon_getir(db, sablon_id)
    satir = _abone_satir_sozlugu(kayit)
    satir["_montaj_imza_veri_url"] = _abone_imza_veri_url(db, abone_id, "montaj")
    satir["_abone_imza_veri_url"] = _abone_imza_veri_url(db, abone_id, "abone")
    render_edilmis, hata = _montaj_formu_render_tek(sablon["icerik"], satir)
    if hata:
        flash(f"Montaj Formu tasarımında hata var, lütfen tasarımı kontrol edin: {hata}")
        return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon["id"]))

    return render_template(
        "montaj_formu.html",
        sayfalar=[render_edilmis],
        baslik=f"{satir['adi'] or ''} {satir['soyadi'] or ''}".strip(),
        geri_url=url_for("abone_listesi"),
    )


@app.route("/abone/<int:abone_id>/montaj-formu/onizle/<int:sablon_id>")
@login_required
def abone_montaj_formu_onizle(abone_id, sablon_id):
    """M.Form penceresinde formu tam açmadan ÖNCE hangi tasarım olduğunu gösteren önizleme."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        return "<p>Abone bulunamadı.</p>", 404

    sablon = _montaj_formu_sablon_getir(db, sablon_id)
    satir = _abone_satir_sozlugu(kayit)
    satir["_montaj_imza_veri_url"] = _abone_imza_veri_url(db, abone_id, "montaj")
    satir["_abone_imza_veri_url"] = _abone_imza_veri_url(db, abone_id, "abone")
    render_edilmis, hata = _montaj_formu_render_tek(sablon["icerik"], satir)
    if hata:
        return f"<p>Bu tasarımda bir hata var: {hata}</p>"

    return render_template("montaj_formu_onizle_parca.html", icerik=render_edilmis)


@app.route("/abone/montaj-formu/toplu")
@login_required
def abone_montaj_formu_toplu():
    db = get_db()
    kayitlar_ham = _abone_filtreli_kayitlari_getir(db)
    if not kayitlar_ham:
        flash("Filtreye uyan abone bulunamadı.")
        return redirect(url_for("abone_listesi"))

    sablon_id = _montaj_formu_secili_id(db, request.args.get("sablon_id", type=int))
    sablon = _montaj_formu_sablon_getir(db, sablon_id)
    imzalar = _abone_imzalari_toplu_getir(db, [k["id"] for k in kayitlar_ham])

    sayfalar = []
    for kayit in kayitlar_ham:
        satir = _abone_satir_sozlugu(kayit)
        satir["_montaj_imza_veri_url"] = imzalar.get((kayit["id"], "montaj"))
        satir["_abone_imza_veri_url"] = imzalar.get((kayit["id"], "abone"))
        render_edilmis, hata = _montaj_formu_render_tek(sablon["icerik"], satir)
        if hata:
            flash(f"Montaj Formu tasarımında hata var, lütfen tasarımı kontrol edin: {hata}")
            return redirect(url_for("montaj_formu_tasarim", sablon_id=sablon["id"]))
        sayfalar.append(render_edilmis)

    return render_template(
        "montaj_formu.html",
        sayfalar=sayfalar,
        baslik=f"Toplu ({len(sayfalar)} kayıt)",
        geri_url=url_for("abone_listesi") + "?" + request.query_string.decode(),
    )


@app.route("/abone/coklu-sayac")
@login_required
def abone_coklu_sayac():
    """Aynı sayaç no'ya birden fazla abonede rastlanan kayıtları gruplanmış şekilde gösterir."""
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT sayac_no FROM abone WHERE sayac_no IS NOT NULL AND sayac_no != '' "
        "GROUP BY sayac_no HAVING COUNT(*) > 1 ORDER BY sayac_no"
    )
    gruplar = [r["sayac_no"] for r in cur.fetchall()]
    renk_haritasi = {sn: GRUP_RENK_PALETI[i % len(GRUP_RENK_PALETI)] for i, sn in enumerate(gruplar)}

    kayitlar_ham = []
    if gruplar:
        cur.execute(
            "SELECT * FROM abone WHERE sayac_no = ANY(%s) ORDER BY sayac_no, s_no",
            (gruplar,),
        )
        kayitlar_ham = cur.fetchall()
    cur.close()

    satirlar = [_abone_satir_sozlugu(k) for k in kayitlar_ham]
    for s in satirlar:
        s["_grup_renk"] = renk_haritasi.get(s["sayac_no"], "#333333")

    return render_template(
        "abone_coklu_sayac.html", satirlar=satirlar, grup_sayisi=len(gruplar),
    )


def _koy_abone_satir_sozlugu(k):
    return {
        "id": k["id"],
        "koy_adi": k["koy_adi"] or "",
        "sira_no": k["sira_no"] or "",
        "abonelik_tarihi": _gg_aa_yyyy(k["abonelik_tarihi"]),
        "abone_no": k["abone_no"] or "",
        "cihaz_no": k["cihaz_no"] or "",
        "adi": k["adi"] or "",
        "soyadi": k["soyadi"] or "",
        "adres": k["adres"] or "",
    }


def _koy_excel_hucre_metin(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


_KOY_EXCEL_BASLIK_ETIKETLERI = {
    "sira_no": ["SIRA NO", "S.NO", "S NO", "SNO"],
    "abonelik_tarihi": ["ABONELİK TARİHİ", "ABONELIK TARIHI"],
    "abone_no": ["ABONE NO"],
    "cihaz_no": ["CİHAZ NO", "CIHAZ NO"],
    "adi_soyadi": ["ADI SOYADI"],
    "adres": ["ADRES"],
}

_KOY_EXCEL_VARSAYILAN_SUTUNLAR = {
    "sira_no": 0, "abonelik_tarihi": 3, "abone_no": 5,
    "cihaz_no": 9, "adi_soyadi": 13, "adres": 15,
}


def _koy_excel_baslik_normallestir(v):
    metin = _koy_excel_hucre_metin(v)
    return re.sub(r"\s+", " ", metin).strip().upper()


def _koy_excel_sutunlarini_bul(ham_satirlar):
    """SIRA NO / ADI SOYADI gibi başlıkların hangi sütunda olduğunu bulan satırı arar."""
    for satir in ham_satirlar:
        if not satir:
            continue
        bulunanlar = {}
        for idx, deger in enumerate(satir):
            metin = _koy_excel_baslik_normallestir(deger)
            if not metin:
                continue
            for alan, olasi_etiketler in _KOY_EXCEL_BASLIK_ETIKETLERI.items():
                if alan in bulunanlar:
                    continue
                if metin in olasi_etiketler:
                    bulunanlar[alan] = idx
        if "sira_no" in bulunanlar and "adi_soyadi" in bulunanlar:
            return bulunanlar
    return None


def _koy_excel_satirlarini_ayikla(ham_satirlar):
    """TEKSAN tarzı köy abone listesi export'undaki ham satırlardan gerçek veri satırlarını ayıklar."""
    sutunlar = _koy_excel_sutunlarini_bul(ham_satirlar) or _KOY_EXCEL_VARSAYILAN_SUTUNLAR

    sonuc = []
    for satir in ham_satirlar:
        if not satir:
            continue
        ilk_metin = _koy_excel_hucre_metin(satir[sutunlar["sira_no"]] if sutunlar["sira_no"] < len(satir) else None)
        if not ilk_metin:
            continue
        try:
            int(float(ilk_metin.replace(",", ".")))
        except ValueError:
            continue

        def al(alan):
            idx = sutunlar.get(alan)
            if idx is None or idx >= len(satir):
                return None
            return satir[idx]

        abonelik_tarihi_ham = _koy_excel_hucre_metin(al("abonelik_tarihi"))
        adi_soyadi_ham = _koy_excel_hucre_metin(al("adi_soyadi"))
        parcalar = [p for p in re.split(r"\s{2,}", adi_soyadi_ham) if p]
        if len(parcalar) >= 2:
            adi, soyadi = parcalar[0], " ".join(parcalar[1:])
        elif parcalar:
            adi, soyadi = parcalar[0], ""
        else:
            adi, soyadi = "", ""

        sonuc.append({
            "sira_no": ilk_metin,
            "abonelik_tarihi": _tarih_iso_hale_getir(abonelik_tarihi_ham) or abonelik_tarihi_ham,
            "abone_no": _koy_excel_hucre_metin(al("abone_no")),
            "cihaz_no": _koy_excel_hucre_metin(al("cihaz_no")),
            "adi": adi,
            "soyadi": soyadi,
            "adres": _koy_excel_hucre_metin(al("adres")),
        })
    return sonuc


def _koy_excel_ayikla(dosya):
    """Yüklenen .xls/.xlsx dosyasından ham satırları okuyup köy abone kayıtlarına çevirir."""
    dosya_adi = (dosya.filename or "").lower()
    icerik = dosya.read()
    ham_satirlar = []

    if dosya_adi.endswith(".xls") and not dosya_adi.endswith(".xlsx"):
        import xlrd
        kitap = xlrd.open_workbook(file_contents=icerik)
        sayfa = kitap.sheet_by_index(0)
        for r in range(sayfa.nrows):
            satir = []
            for c in range(sayfa.ncols):
                hucre = sayfa.cell(r, c)
                deger = hucre.value
                if hucre.ctype == xlrd.XL_CELL_DATE:
                    try:
                        deger = xlrd.xldate_as_datetime(deger, kitap.datemode).strftime("%d.%m.%Y")
                    except Exception:
                        pass
                satir.append(deger)
            ham_satirlar.append(satir)
    else:
        import openpyxl
        kitap = openpyxl.load_workbook(io.BytesIO(icerik), data_only=True)
        sayfa = kitap.worksheets[0]
        for row in sayfa.iter_rows(values_only=True):
            ham_satirlar.append(list(row))

    return _koy_excel_satirlarini_ayikla(ham_satirlar)


@app.route("/koy-abone-listesi")
@login_required
def koy_abone_listesi():
    q = request.args.get("q", "").strip()
    koy = request.args.get("koy", "").strip()
    db = get_db()
    cur = db.cursor()

    sql = "SELECT * FROM koy_abone WHERE 1=1"
    params = []
    if koy:
        sql += " AND koy_adi = %s"
        params.append(koy)
    if q:
        _kk = _turkce_esle_kosul
        sql += (f" AND ({_kk('adi')} LIKE %s OR {_kk('soyadi')} LIKE %s OR {_kk('cihaz_no')} LIKE %s "
                f"OR {_kk('abone_no')} LIKE %s OR {_kk('adres')} LIKE %s)")
        params += [_turkce_normallestir(f"%{q}%")] * 5

    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()

    cur.execute("SELECT DISTINCT koy_adi FROM koy_abone ORDER BY koy_adi")
    koyler = cur.fetchall()

    cur.execute("SELECT COUNT(*) AS c FROM koy_abone")
    toplam_kayit = cur.fetchone()["c"]

    secili_koy_toplam = None
    if koy:
        cur.execute("SELECT COUNT(*) AS c FROM koy_abone WHERE koy_adi = %s", (koy,))
        secili_koy_toplam = cur.fetchone()["c"]
    cur.close()

    satirlar = [_koy_abone_satir_sozlugu(k) for k in kayitlar_ham]

    def _siralama_anahtari(s):
        try:
            return (s["koy_adi"], 0, int(s["sira_no"]))
        except (TypeError, ValueError):
            return (s["koy_adi"], 1, s["sira_no"])

    sira = _sira_yonu_al()
    satirlar.sort(key=_siralama_anahtari, reverse=(sira == "desc"))

    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "koy_abone_listesi.html", satirlar=satirlar, koyler=koyler,
        q=q, secili_koy=koy,
        filtreli_kayit=filtreli_kayit, toplam_kayit=toplam_kayit,
        secili_koy_toplam=secili_koy_toplam,
        sira=sira, sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
    )


@app.route("/koy-abone-listesi/koy-sil", methods=["POST"])
@login_required
def koy_abone_koy_sil():
    koy_adi = request.form.get("koy_adi", "").strip()
    if koy_adi:
        db = get_db()
        cur = db.cursor()
        cur.execute("DELETE FROM koy_abone WHERE koy_adi = %s", (koy_adi,))
        silinen = cur.rowcount
        db.commit()
        cur.close()
        flash(f"\"{koy_adi}\" köyüne ait {silinen} kayıt silindi.")
    return redirect(url_for("koy_abone_listesi"))


def _koy_abone_kaydet(kayit_id):
    f = request.form
    alanlar = dict(
        koy_adi=f.get("koy_adi", "").strip(),
        sira_no=f.get("sira_no", "").strip(),
        abonelik_tarihi=f.get("abonelik_tarihi", "").strip(),
        abone_no=f.get("abone_no", "").strip(),
        cihaz_no=f.get("cihaz_no", "").strip(),
        adi=f.get("adi", "").strip(),
        soyadi=f.get("soyadi", "").strip(),
        adres=f.get("adres", "").strip(),
    )
    db = get_db()
    cur = db.cursor()
    if kayit_id is None:
        kolonlar = ", ".join(alanlar.keys())
        yer_tutucular = ", ".join(["%s"] * len(alanlar))
        cur.execute(f"INSERT INTO koy_abone ({kolonlar}) VALUES ({yer_tutucular})", list(alanlar.values()))
    else:
        set_ifadesi = ", ".join([f"{k} = %s" for k in alanlar.keys()])
        cur.execute(f"UPDATE koy_abone SET {set_ifadesi}, updated_at = NOW() WHERE id = %s", list(alanlar.values()) + [kayit_id])
    db.commit()
    cur.close()


@app.route("/koy-abone-listesi/yeni", methods=["GET", "POST"])
@login_required
def koy_abone_yeni():
    if request.method == "POST":
        _koy_abone_kaydet(None)
        return redirect(url_for("koy_abone_listesi"))
    return render_template("koy_abone_form.html", kayit=None)


@app.route("/koy-abone-listesi/<int:kayit_id>/duzenle", methods=["GET", "POST"])
@login_required
def koy_abone_duzenle(kayit_id):
    db = get_db()
    if request.method == "POST":
        _koy_abone_kaydet(kayit_id)
        return redirect(url_for("koy_abone_listesi"))
    cur = db.cursor()
    cur.execute("SELECT * FROM koy_abone WHERE id = %s", (kayit_id,))
    kayit = cur.fetchone()
    cur.close()
    if not kayit:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("koy_abone_listesi"))
    return render_template("koy_abone_form.html", kayit=kayit)


@app.route("/koy-abone-listesi/<int:kayit_id>/sil", methods=["POST"])
@login_required
def koy_abone_sil(kayit_id):
    geri = request.args.get("geri", "")
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM koy_abone WHERE id = %s", (kayit_id,))
    db.commit()
    cur.close()
    hedef = url_for("koy_abone_listesi")
    if geri:
        hedef += "?" + geri
    return redirect(hedef)


@app.route("/koy-abone-listesi/yukle", methods=["GET", "POST"])
@login_required
def koy_abone_yukle():
    db = get_db()
    if request.method == "GET":
        cur = db.cursor()
        cur.execute("SELECT DISTINCT koy_adi FROM koy_abone ORDER BY koy_adi")
        koyler = cur.fetchall()
        cur.close()
        return render_template("koy_abone_yukle.html", koyler=koyler)

    koy_adi = request.form.get("koy_adi", "").strip()
    dosya = request.files.get("dosya")

    if not koy_adi:
        flash("Köy adı boş bırakılamaz.")
        return redirect(url_for("koy_abone_yukle"))
    if not dosya or not dosya.filename:
        flash("Bir Excel dosyası (.xls veya .xlsx) seçmediniz.")
        return redirect(url_for("koy_abone_yukle"))

    try:
        yeni_kayitlar = _koy_excel_ayikla(dosya)
    except Exception as e:
        flash(f"Dosya okunamadı, format desteklenmiyor olabilir: {e}")
        return redirect(url_for("koy_abone_yukle"))

    if not yeni_kayitlar:
        flash("Dosyada okunabilir abone kaydı bulunamadı. Dosya formatını kontrol edin.")
        return redirect(url_for("koy_abone_yukle"))

    cur = db.cursor()
    cur.execute("DELETE FROM koy_abone WHERE koy_adi = %s", (koy_adi,))
    kolonlar = ["koy_adi", "sira_no", "abonelik_tarihi", "abone_no", "cihaz_no", "adi", "soyadi", "adres"]
    toplu_degerler = [
        (koy_adi, k["sira_no"], k["abonelik_tarihi"], k["abone_no"], k["cihaz_no"], k["adi"], k["soyadi"], k["adres"])
        for k in yeni_kayitlar
    ]
    kolonlar_sql = ", ".join(kolonlar)
    yer_tutucular = ", ".join(["%s"] * len(kolonlar))
    cur.executemany(f"INSERT INTO koy_abone ({kolonlar_sql}) VALUES ({yer_tutucular})", toplu_degerler)
    db.commit()
    cur.close()

    flash(f"\"{koy_adi}\" köyü için {len(yeni_kayitlar)} kayıt yüklendi (önceki liste silinip yenisiyle değiştirildi).")
    return redirect(url_for("koy_abone_listesi", koy=koy_adi))


def _sayilastir(deger):
    try:
        return float(str(deger).replace(",", ".").strip() or 0)
    except ValueError:
        return 0.0


def _konum_sayilastir(deger):
    """Konum alanları için: boşsa/hatalıysa None döner."""
    try:
        deger = str(deger).strip()
        return float(deger) if deger else None
    except (ValueError, TypeError):
        return None


def _ayar_getir(db, anahtar):
    """Basit anahtar/değer ayar deposundan bir değeri okur."""
    cur = db.cursor()
    cur.execute("SELECT deger FROM ayar WHERE anahtar = %s", (anahtar,))
    satir = cur.fetchone()
    cur.close()
    return satir["deger"] if satir else None


def _ayar_kaydet(db, anahtar, deger):
    """Basit anahtar/değer ayar deposuna bir değer yazar."""
    cur = db.cursor()
    cur.execute(
        "INSERT INTO ayar (anahtar, deger) VALUES (%s, %s) "
        "ON CONFLICT (anahtar) DO UPDATE SET deger = EXCLUDED.deger",
        (anahtar, deger),
    )
    db.commit()
    cur.close()


def _sonraki_s_no(db):
    """Yeni Abone formunda gösterilen, KABACA bir öngörü."""
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM abone")
    satir = cur.fetchone()
    cur.close()
    return (satir["c"] or 0) + 1


def _abone_sira_numaralarini_yenile(db):
    """S.No, abone kayıtlarında montaj tarihine göre (eskiden yeniye) sıralı
    tutuluyor ve bu sıraya göre 1'den başlayarak boşluksuz yeniden numaralandırılıyor."""
    cur = db.cursor()
    cur.execute(
        """
        UPDATE abone a
        SET s_no = t.yeni_sira
        FROM (
            SELECT id, ROW_NUMBER() OVER (ORDER BY montaj_tarihi ASC NULLS LAST, id ASC) AS yeni_sira
            FROM abone
        ) t
        WHERE a.id = t.id AND a.s_no IS DISTINCT FROM t.yeni_sira
        """
    )
    db.commit()
    cur.close()


def _fabrika_tamir_sira_numaralarini_yenile(db):
    """Sıra No, Fabrika/Tamir kayıtlarında oluşturuluş sırasına göre (eskiden
    yeniye) sıralı tutuluyor ve bu sıraya göre 1'den başlayarak boşluksuz
    yeniden numaralandırılıyor; silinmiş (silindi_mi) kayıtlar sayılmıyor."""
    cur = db.cursor()
    cur.execute(
        """
        UPDATE fabrika_tamir ft
        SET sira_no = t.yeni_sira
        FROM (
            SELECT id, ROW_NUMBER() OVER (ORDER BY created_at ASC NULLS LAST, id ASC) AS yeni_sira
            FROM fabrika_tamir
            WHERE silindi_mi IS NOT TRUE
        ) t
        WHERE ft.id = t.id AND ft.sira_no IS DISTINCT FROM t.yeni_sira
        """
    )
    db.commit()
    cur.close()


def _sonraki_senet_no(db):
    cur = db.cursor()
    cur.execute("SELECT senet_no FROM abone WHERE senet_no IS NOT NULL AND senet_no != ''")
    satirlar = cur.fetchall()
    cur.close()
    en_buyuk = 0
    for s in satirlar:
        try:
            deger = int(s["senet_no"])
            if deger > en_buyuk:
                en_buyuk = deger
        except (ValueError, TypeError):
            pass
    return str(en_buyuk + 1)


@app.route("/abone/yeni", methods=["GET", "POST"])
@login_required
def abone_yeni():
    if request.method == "POST":
        yeni_id = _abone_kaydet(None)
        db = get_db()
        _abone_fotograflarini_kaydet(db, yeni_id, request.files.getlist("fotograflar"))
        _abone_imzalarini_kaydet(db, yeni_id, request.form)
        return redirect(url_for("abone_listesi"))
    db = get_db()
    return render_template(
        "abone_form.html", kayit=None,
        sonraki_s_no=_sonraki_s_no(db),
        sonraki_senet_no=_sonraki_senet_no(db),
        fotograflar=[], imzalar={"montaj": False, "abone": False},
        ozel_alan_harita=_ozel_alan_harita(_ozel_alanlari_getir(db, "abone")),
        bugun=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/abone/<int:abone_id>/duzenle", methods=["GET", "POST"])
@login_required
def abone_duzenle(abone_id):
    db = get_db()
    geri = request.args.get("geri", "") or request.form.get("geri", "")
    sonraki_hedef = request.args.get("hedef", "") or request.form.get("hedef", "")
    if request.method == "POST":
        _abone_kaydet(abone_id)
        _abone_fotograflarini_kaydet(db, abone_id, request.files.getlist("fotograflar"))
        _abone_imzalarini_kaydet(db, abone_id, request.form)
        if sonraki_hedef:
            return redirect(sonraki_hedef)
        hedef = url_for("abone_listesi")
        if geri:
            hedef += "?" + geri
        return redirect(hedef)
    cur = db.cursor()
    cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("abone_listesi"))
    cur = db.cursor()
    cur.execute(
        "SELECT id, dosya_adi, content_type FROM abone_fotograf WHERE abone_id = %s ORDER BY id",
        (abone_id,),
    )
    fotograflar = cur.fetchall()
    cur.close()
    return render_template(
        "abone_form.html", kayit=kayit, geri=geri, hedef=sonraki_hedef, fotograflar=fotograflar,
        imzalar=_abone_imzalari_getir(db, abone_id),
        ozel_alan_harita=_ozel_alan_harita(_ozel_alanlari_getir(db, "abone")),
    )


@app.route("/abone/<int:abone_id>/sil", methods=["POST"])
@login_required
def abone_sil(abone_id):
    geri = request.args.get("geri", "")
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM abone WHERE id = %s", (abone_id,))
    db.commit()
    cur.close()
    _abone_sira_numaralarini_yenile(db)
    hedef = url_for("abone_listesi")
    if geri:
        hedef += "?" + geri
    return redirect(hedef)


@app.route("/abone-fotograf/<int:foto_id>")
@login_required
def abone_fotograf_goster(foto_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT content_type, icerik FROM abone_fotograf WHERE id = %s", (foto_id,))
    foto = cur.fetchone()
    cur.close()
    if foto is None:
        return "Fotoğraf bulunamadı.", 404
    yanit = Response(bytes(foto["icerik"]), mimetype=foto["content_type"] or "application/octet-stream")
    yanit.headers["X-Content-Type-Options"] = "nosniff"
    return yanit


@app.route("/abone-fotograf/<int:foto_id>/sil", methods=["POST"])
@login_required
def abone_fotograf_sil(foto_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT abone_id FROM abone_fotograf WHERE id = %s", (foto_id,))
    foto = cur.fetchone()
    if foto:
        cur.execute("DELETE FROM abone_fotograf WHERE id = %s", (foto_id,))
        db.commit()
    cur.close()
    if foto:
        return redirect(url_for("abone_duzenle", abone_id=foto["abone_id"]))
    return redirect(url_for("abone_listesi"))


def _abone_kaydet(abone_id):
    f = request.form
    sayac_tutari = _sayilastir(f.get("sayac_tutari"))
    malzeme_tutari = _sayilastir(f.get("malzeme_tutari"))

    db = get_db()
    cur = db.cursor()

    yeni_kayit_mi = abone_id is None

    if yeni_kayit_mi:
        girilen_alinan_toplam = _sayilastir(f.get("alinan_tutar"))
        malzeme_alinan = min(girilen_alinan_toplam, malzeme_tutari) if malzeme_tutari > 0 else 0.0
        alinan_tutar = girilen_alinan_toplam - malzeme_alinan
    else:
        cur.execute("SELECT alinan_tutar, malzeme_alinan FROM abone WHERE id = %s", (abone_id,))
        mevcut = cur.fetchone()
        alinan_tutar = (mevcut["alinan_tutar"] or 0) if mevcut else 0.0
        malzeme_alinan = (mevcut["malzeme_alinan"] or 0) if mevcut else 0.0

    senet_tutari_hesap = sayac_tutari + malzeme_tutari - alinan_tutar - malzeme_alinan

    if senet_tutari_hesap == 0:
        senet_no_final = ""
    else:
        mevcut_senet_no = ""
        if abone_id is not None:
            cur.execute("SELECT senet_no FROM abone WHERE id = %s", (abone_id,))
            satir = cur.fetchone()
            if satir and satir["senet_no"]:
                mevcut_senet_no = satir["senet_no"]
        senet_no_final = mevcut_senet_no if mevcut_senet_no else _sonraki_senet_no(db)

    alanlar = dict(
        koy_adi=f.get("koy_adi", "").strip(),
        adi=f.get("adi", "").strip(),
        soyadi=f.get("soyadi", "").strip(),
        sayac_no=f.get("sayac_no", "").strip(),
        senet_tutari=senet_tutari_hesap,
        sayac_tutari=sayac_tutari,
        alinan_tutar=alinan_tutar,
        malzeme_tutari=malzeme_tutari,
        malzeme_alinan=malzeme_alinan,
        senet_no=senet_no_final,
        senet_sahibi_adi=f.get("senet_sahibi_adi", "").strip(),
        senet_sahibi_soyadi=f.get("senet_sahibi_soyadi", "").strip(),
        telefon=_telefon_formatla(f.get("telefon", "")),
        telefon2=_telefon_formatla(f.get("telefon2", "")),
        baba_adi=f.get("baba_adi", "").strip(),
        montaj_tarihi=f.get("montaj_tarihi", "").strip(),
        montaj_personeli=f.get("montaj_personeli", "").strip(),
        odeme_tarihi=f.get("odeme_tarihi", "").strip(),
        odeme_sekli=f.get("odeme_sekli", "").strip(),
        odeme_gun_sozu=f.get("odeme_gun_sozu", "").strip(),
        odemeyi_gonderen=f.get("odemeyi_gonderen", "").strip(),
        aciklama=f.get("aciklama", "").strip(),
        muhtara_odenecek=_sayilastir(f.get("muhtara_odenecek")),
        muhtara_odenen=_sayilastir(f.get("muhtara_odenen")),
        fatura_no=f.get("fatura_no", "").strip(),
        konum_enlem=_konum_sayilastir(f.get("konum_enlem")),
        konum_boylam=_konum_sayilastir(f.get("konum_boylam")),
        kimlik_no=_kimlik_no_temizle(f.get("kimlik_no", "")),
        vergi_dairesi=f.get("vergi_dairesi", "").strip(),
        adres=f.get("adres", "").strip(),
        eposta=f.get("eposta", "").strip(),
    )

    for oa in _ozel_alanlari_getir(db, "abone"):
        if oa["tur"] == "sayi":
            alanlar[oa["kolon_adi"]] = _sayilastir(f.get(oa["kolon_adi"]))
        else:
            alanlar[oa["kolon_adi"]] = f.get(oa["kolon_adi"], "").strip()

    if yeni_kayit_mi:
        kolonlar = ", ".join(alanlar.keys())
        yer_tutucular = ", ".join(["%s"] * len(alanlar))
        cur.execute(
            f"INSERT INTO abone ({kolonlar}) VALUES ({yer_tutucular}) RETURNING id",
            list(alanlar.values()),
        )
        abone_id = cur.fetchone()["id"]

        tahsilat_tarihi = (
            f.get("odeme_tarihi", "").strip()
            or f.get("montaj_tarihi", "").strip()
            or datetime.now().strftime("%Y-%m-%d")
        )
        odeme_sekli = f.get("odeme_sekli", "").strip()
        odemeyi_yapan = f.get("odemeyi_gonderen", "").strip()
        for tur, tutar, aciklama in (
            ("sayac", alinan_tutar, "Abone kaydı sırasında alınan (sayaç)"),
            ("malzeme", malzeme_alinan, "Abone kaydı sırasında alınan (malzeme)"),
        ):
            if tutar:
                cur.execute(
                    "INSERT INTO tahsilat (abone_id, tarih, tur, tutar, odeme_sekli, odemeyi_yapan, aciklama) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (abone_id, tahsilat_tarihi, tur, tutar, odeme_sekli, odemeyi_yapan, aciklama),
                )
    else:
        set_ifadesi = ", ".join([f"{k} = %s" for k in alanlar.keys()])
        cur.execute(
            f"UPDATE abone SET {set_ifadesi}, updated_at = NOW() WHERE id = %s",
            list(alanlar.values()) + [abone_id],
        )
    db.commit()
    cur.close()
    _abone_sira_numaralarini_yenile(db)
    return abone_id


def _base64_imza_decode(veri_url):
    """Bir <canvas>.toDataURL('image/png') çıktısını ham PNG baytlarına çevirir."""
    if not veri_url or not veri_url.startswith("data:image/"):
        return None
    try:
        _baslik, b64_kisim = veri_url.split(",", 1)
        return base64.b64decode(b64_kisim)
    except Exception:
        return None


def _abone_imzalarini_kaydet(db, abone_id, form):
    """Montaj Personeli ve Abone imza alanlarını işler; ayrı 'abone_imza' tablosuna yazar."""
    cur = db.cursor()
    for form_alani, tur in (
        ("imza_montaj_personeli_veri", "montaj"),
        ("imza_abone_veri", "abone"),
    ):
        deger = form.get(form_alani, "")
        if not deger:
            continue
        if deger == "TEMIZLE":
            cur.execute("DELETE FROM abone_imza WHERE abone_id = %s AND tur = %s", (abone_id, tur))
        else:
            png_baytlari = _base64_imza_decode(deger)
            if png_baytlari:
                cur.execute(
                    "INSERT INTO abone_imza (abone_id, tur, icerik) VALUES (%s, %s, %s) "
                    "ON CONFLICT (abone_id, tur) DO UPDATE SET icerik = EXCLUDED.icerik, created_at = NOW()",
                    (abone_id, tur, psycopg2.Binary(png_baytlari)),
                )
    db.commit()
    cur.close()


def _abone_imzalari_getir(db, abone_id):
    """Bir abonenin kayıtlı imzalarını {'montaj': True/False, 'abone': True/False} şeklinde döner."""
    cur = db.cursor()
    cur.execute("SELECT tur FROM abone_imza WHERE abone_id = %s", (abone_id,))
    turler = {r["tur"] for r in cur.fetchall()}
    cur.close()
    return {"montaj": "montaj" in turler, "abone": "abone" in turler}


def _abone_imza_veri_url(db, abone_id, tur):
    """Bir abonenin TEK bir imzasını data: URL olarak döner; imza yoksa None döner."""
    cur = db.cursor()
    cur.execute("SELECT icerik FROM abone_imza WHERE abone_id = %s AND tur = %s", (abone_id, tur))
    kayit = cur.fetchone()
    cur.close()
    if not kayit:
        return None
    b64 = base64.b64encode(bytes(kayit["icerik"])).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _abone_imzalari_toplu_getir(db, abone_idler):
    """Birden fazla abonenin imzalarını TEK sorguda alıp sözlük olarak döner."""
    if not abone_idler:
        return {}
    cur = db.cursor()
    cur.execute(
        "SELECT abone_id, tur, icerik FROM abone_imza WHERE abone_id = ANY(%s)",
        (list(abone_idler),),
    )
    sonuc = {}
    for r in cur.fetchall():
        b64 = base64.b64encode(bytes(r["icerik"])).decode("ascii")
        sonuc[(r["abone_id"], r["tur"])] = f"data:image/png;base64,{b64}"
    cur.close()
    return sonuc


@app.route("/abone-imza/<int:abone_id>/<hangi>")
@login_required
def abone_imza_goster(abone_id, hangi):
    """Montaj Personeli ('montaj') veya Abone ('abone') imzasının PNG'ini gösterir."""
    if hangi not in ("montaj", "abone"):
        return "Geçersiz istek.", 404
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT icerik FROM abone_imza WHERE abone_id = %s AND tur = %s", (abone_id, hangi))
    kayit = cur.fetchone()
    cur.close()
    if not kayit:
        return "İmza bulunamadı.", 404
    yanit = Response(bytes(kayit["icerik"]), mimetype="image/png")
    yanit.headers["X-Content-Type-Options"] = "nosniff"
    return yanit


@app.route("/abone/<int:abone_id>/tahsilat", methods=["GET", "POST"])
@login_required
def abone_tahsilat(abone_id):
    db = get_db()
    cur = db.cursor()
    geri = request.args.get("geri", "") or request.form.get("geri", "")

    if request.method == "POST":
        tutar = _sayilastir(request.form.get("tutar"))
        tarih = request.form.get("tarih", "").strip()
        odeme_sekli = request.form.get("odeme_sekli", "").strip()
        odemeyi_yapan = request.form.get("odemeyi_yapan", "").strip()
        aciklama = request.form.get("aciklama", "").strip()

        if tutar:
            cur.execute(
                "SELECT sayac_tutari, alinan_tutar, malzeme_tutari, malzeme_alinan FROM abone WHERE id = %s",
                (abone_id,),
            )
            mevcut = cur.fetchone()
            malzeme_kalan = (mevcut["malzeme_tutari"] or 0) - (mevcut["malzeme_alinan"] or 0) if mevcut else 0

            malzeme_payi = min(tutar, malzeme_kalan) if malzeme_kalan > 0 else 0
            sayac_payi = tutar - malzeme_payi

            for tur, pay in (("malzeme", malzeme_payi), ("sayac", sayac_payi)):
                if pay:
                    cur.execute(
                        "INSERT INTO tahsilat (abone_id, tarih, tur, tutar, odeme_sekli, odemeyi_yapan, aciklama) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                        (abone_id, tarih, tur, pay, odeme_sekli, odemeyi_yapan, aciklama),
                    )
                    kolon = "alinan_tutar" if tur == "sayac" else "malzeme_alinan"
                    cur.execute(f"UPDATE abone SET {kolon} = {kolon} + %s WHERE id = %s", (pay, abone_id))
            db.commit()

        cur.close()
        hedef = url_for("abone_tahsilat", abone_id=abone_id)
        if geri:
            hedef += "?geri=" + _url_quote(geri, safe="")
        return redirect(hedef)

    cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
    abone = cur.fetchone()
    if abone is None:
        cur.close()
        flash("Kayıt bulunamadı.")
        return redirect(url_for("abone_listesi"))

    cur.execute("SELECT * FROM tahsilat WHERE abone_id = %s ORDER BY tarih DESC, id DESC", (abone_id,))
    tahsilatlar = cur.fetchall()
    cur.close()

    return render_template(
        "abone_tahsilat.html", abone=abone, tahsilatlar=tahsilatlar, geri=geri,
        bugun=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/abone/<int:abone_id>/fatura-kes", methods=["GET", "POST"])
@login_required
def abone_fatura_kes(abone_id):
    """Abone kaydındaki Sayaç Tutarı ve Malzeme Tutarı üzerinden e-Fatura/e-Arşiv Fatura keser."""
    db = get_db()
    geri = request.args.get("geri", "") or request.form.get("geri", "")
    cur = db.cursor()
    cur.execute("SELECT * FROM abone WHERE id = %s", (abone_id,))
    abone = cur.fetchone()
    cur.close()
    if abone is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("abone_listesi"))

    onerilen_tur = _fatura_turu_belirle(abone["kimlik_no"])

    if request.method == "POST":
        fatura_turu = request.form.get("fatura_turu")
        if fatura_turu not in ("earsiv", "efatura"):
            flash("Geçerli bir fatura türü seçin.")
            return redirect(url_for("abone_fatura_kes", abone_id=abone_id, geri=geri))
        if not abone["kimlik_no"] or not abone["adres"]:
            flash("Fatura kesmeden önce abonenin TC Kimlik No/Vergi No ve Adres bilgilerini doldurun.")
            return redirect(url_for("abone_duzenle", abone_id=abone_id))

        kalemler = _fatura_kalemlerini_formdan_oku(request.form)
        if not kalemler:
            flash("En az bir fatura kalemi girilmeli (açıklama, miktar ve birim fiyat).")
            return redirect(url_for("abone_fatura_kes", abone_id=abone_id, geri=geri))

        fatura_tarihi_form = request.form.get("fatura_tarihi", "").strip()
        try:
            fatura_tarihi_dt = datetime.strptime(fatura_tarihi_form, "%Y-%m-%d")
        except ValueError:
            flash("Geçerli bir fatura tarihi seçin.")
            return redirect(url_for("abone_fatura_kes", abone_id=abone_id, geri=geri))
        if fatura_tarihi_dt.date() > datetime.now().date():
            flash("Fatura tarihi bugünden ileri bir tarih olamaz.")
            return redirect(url_for("abone_fatura_kes", abone_id=abone_id, geri=geri))

        fatura_id = _hizli_fatura_gonder(
            db, "abone", abone_id, dict(abone), kalemler, fatura_turu,
            session.get("kullanici_adi", ""), fatura_tarihi=fatura_tarihi_form,
        )
        hedef = url_for("fatura_goruntule", fatura_id=fatura_id)
        return redirect(hedef)

    onizleme = []
    if float(abone["sayac_tutari"] or 0) > 0:
        taban, _kdv = _hizli_kdv_ayir(float(abone["sayac_tutari"]))
        onizleme.append({"aciklama": "ÖN ÖDEMELİ SU SAYACI", "miktar": 1, "birim_fiyat": taban})
    if float(abone["malzeme_tutari"] or 0) > 0:
        taban, _kdv = _hizli_kdv_ayir(float(abone["malzeme_tutari"]))
        onizleme.append({"aciklama": "FİTTİNGS MALZEMESİ", "miktar": 1, "birim_fiyat": taban})

    fatura_kes_url = url_for("abone_fatura_kes", abone_id=abone_id) + (f"?geri={_url_quote(geri, safe='')}" if geri else "")
    return render_template(
        "fatura_kes.html", kaynak=abone, kaynak_tur="abone", kaynak_ad=f"{abone['adi']} {abone['soyadi']}",
        duzenle_url=url_for("abone_duzenle", abone_id=abone_id) + f"?hedef={_url_quote(fatura_kes_url, safe='')}",
        geri_url=fatura_kes_url,
        onizleme=onizleme,
        onerilen_tur=onerilen_tur, hizli_ayarli_mi=_hizli_ayarli_mi(), geri=geri,
        bugun=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/tahsilat/<int:tahsilat_id>/sil", methods=["POST"])
@login_required
def tahsilat_sil(tahsilat_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT abone_id, tur, tutar FROM tahsilat WHERE id = %s", (tahsilat_id,))
    kayit = cur.fetchone()
    abone_id = None
    if kayit:
        abone_id = kayit["abone_id"]
        kolon = "alinan_tutar" if kayit["tur"] == "sayac" else "malzeme_alinan"
        cur.execute(f"UPDATE abone SET {kolon} = {kolon} - %s WHERE id = %s", (kayit["tutar"], abone_id))
        cur.execute("DELETE FROM tahsilat WHERE id = %s", (tahsilat_id,))
        db.commit()
    cur.close()
    if abone_id:
        return redirect(url_for("abone_tahsilat", abone_id=abone_id))
    return redirect(url_for("abone_listesi"))


@app.route("/tahsilat/<int:tahsilat_id>/makbuz")
@login_required
def tahsilat_makbuz(tahsilat_id):
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT t.*, a.adi AS abone_adi, a.soyadi AS abone_soyadi, a.sayac_no AS abone_sayac_no "
        "FROM tahsilat t JOIN abone a ON a.id = t.abone_id WHERE t.id = %s",
        (tahsilat_id,),
    )
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("abone_listesi"))

    ad_soyad = f"{kayit['abone_adi']} {kayit['abone_soyadi']}"
    aciklama_basligi = "Sayaç Ücreti Ödemesi" if kayit["tur"] == "sayac" else "Malzeme Ücreti Ödemesi"
    odemeyi_yapan = kayit["odemeyi_yapan"] or ad_soyad
    not_goster = bool(kayit["odemeyi_yapan"]) and kayit["odemeyi_yapan"].strip().upper() != ad_soyad.strip().upper()

    return render_template(
        "makbuz.html",
        sira_no=str(kayit["id"]).zfill(6),
        tarih=_gg_aa_yyyy(kayit["tarih"]),
        sayin=ad_soyad,
        tutar=kayit["tutar"],
        tutar_yazi=_tutar_yaziya_cevir(kayit["tutar"]),
        odeme_esles=_odeme_sekli_esle(kayit["odeme_sekli"]),
        odeme_sekli_metin=kayit["odeme_sekli"],
        aciklama_basligi=aciklama_basligi,
        seri_no=kayit["abone_sayac_no"],
        odemeyi_yapan=odemeyi_yapan,
        ad_soyad=ad_soyad,
        not_goster=not_goster,
        geri_url=url_for("abone_tahsilat", abone_id=kayit["abone_id"]),
    )


KOY_KOLONLARI = [
    ("koy_adi", "Köy Adı", "metin"),
    ("sayac_tutari_toplami", "Sayaç Tutarı", "sayi"),
    ("malzeme_tutari_toplami", "Malzeme Tutarı", "sayi"),
    ("genel_satis_tutari", "Genel Satış", "sayi"),
    ("tahsil_edilen_tutar", "Tahsil Edilen", "sayi"),
    ("kalan_tutar", "Kalan Tutar", "sayi"),
    ("muhtara_odenecek", "Muhtara Ödenecek", "sayi"),
    ("muhtara_odenen", "Muhtara Ödenen", "sayi"),
    ("muhtara_kalan", "Muhtar Kalan", "sayi"),
]


@app.route("/tahsilat")
@login_required
def tahsilat():
    yonlendirme = _filtre_durumu_uygula("tahsilat")
    if yonlendirme:
        return yonlendirme

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT koy_adi, SUM(sayac_tutari) AS sayac_tutari_toplami, SUM(malzeme_tutari) AS malzeme_tutari_toplami, SUM(sayac_tutari + malzeme_tutari) AS genel_satis_tutari, SUM(alinan_tutar + malzeme_alinan) AS tahsil_edilen_tutar, SUM(sayac_tutari + malzeme_tutari - alinan_tutar - malzeme_alinan) AS kalan_tutar, SUM(muhtara_odenecek) AS muhtara_odenecek, SUM(muhtara_odenen) AS muhtara_odenen, SUM(muhtara_odenecek - muhtara_odenen) AS muhtara_kalan FROM abone GROUP BY koy_adi ORDER BY koy_adi"
    )
    satirlar_tum = cur.fetchall()
    cur.close()

    deger_secenekleri = {}
    for anahtar, etiket, tur in KOY_KOLONLARI:
        secenekler = []
        gorulen = set()
        for s in satirlar_tum:
            ham = s[anahtar]
            if ham is None:
                continue
            if tur == "sayi":
                ham_yuvarlanmis = round(float(ham), 2)
                anahtar_kelime = str(ham_yuvarlanmis)
                metin = tl_format(ham_yuvarlanmis)
            else:
                anahtar_kelime = str(ham)
                metin = str(ham)
            if anahtar_kelime not in gorulen:
                gorulen.add(anahtar_kelime)
                secenekler.append((anahtar_kelime, metin))
        secenekler.sort(key=lambda x: x[1])
        deger_secenekleri[anahtar] = secenekler

    deger_secili = {}
    for anahtar, etiket, tur in KOY_KOLONLARI:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")

    def _satir_uyumlu(s):
        for anahtar, etiket, tur in KOY_KOLONLARI:
            secilenler = deger_secili[anahtar]
            if not secilenler:
                continue
            ham = s[anahtar]
            if tur == "sayi":
                deger_kelime = str(round(float(ham or 0), 2))
            else:
                deger_kelime = str(ham)
            if deger_kelime not in secilenler:
                return False
        return True

    satirlar = [s for s in satirlar_tum if _satir_uyumlu(s)]

    genel = {
        "sayac_tutari_toplami": sum(s["sayac_tutari_toplami"] or 0 for s in satirlar),
        "malzeme_tutari_toplami": sum(s["malzeme_tutari_toplami"] or 0 for s in satirlar),
        "genel_satis_tutari": sum(s["genel_satis_tutari"] or 0 for s in satirlar),
        "tahsil_edilen_tutar": sum(s["tahsil_edilen_tutar"] or 0 for s in satirlar),
        "kalan_tutar": sum(s["kalan_tutar"] or 0 for s in satirlar),
        "muhtara_odenecek": sum(s["muhtara_odenecek"] or 0 for s in satirlar),
        "muhtara_odenen": sum(s["muhtara_odenen"] or 0 for s in satirlar),
        "muhtara_kalan": sum(s["muhtara_kalan"] or 0 for s in satirlar),
    }
    genel["firma_asil_alacagi"] = genel["kalan_tutar"] - genel["muhtara_kalan"]

    return render_template(
        "tahsilat.html", satirlar=satirlar, genel=genel,
        kolon_listesi=KOY_KOLONLARI, deger_secili=deger_secili,
        deger_secenekleri=deger_secenekleri,
        filtreli_kayit=len(satirlar), toplam_kayit=len(satirlar_tum),
    )


def _tahsilat_ciktisi_satirlar():
    db = get_db()
    kolon_listesi, kolon_bilgi, _sayisal, ozel_alanlar = _abone_kolon_takimi(db)
    kolonlar_secili = request.args.getlist("kolon")
    goster_kolonlari = kolonlar_secili if kolonlar_secili else [k for k, _ in kolon_listesi]
    koy = request.args.get("koy", "").strip()
    sql = "SELECT * FROM abone WHERE 1=1"
    params = []
    if koy:
        sql += " AND koy_adi = %s"
        params.append(koy)
    for anahtar in goster_kolonlari:
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi
    sql += f" ORDER BY s_no {'DESC' if _sira_yonu_al() == 'desc' else 'ASC'}"
    cur = db.cursor()
    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()
    cur.close()
    satirlar = [_abone_satir_sozlugu(k, ozel_alanlar) for k in kayitlar_ham]

    toplam_ucret = sum(float(k["sayac_tutari"] or 0) + float(k["malzeme_tutari"] or 0) for k in kayitlar_ham)
    tahsil_edilen_ucret = sum(float(k["alinan_tutar"] or 0) + float(k["malzeme_alinan"] or 0) for k in kayitlar_ham)
    ucret_toplamlari = {
        "toplam_ucret": toplam_ucret,
        "tahsil_edilen_ucret": tahsil_edilen_ucret,
        "kalan_bakiye": toplam_ucret - tahsil_edilen_ucret,
    }
    return satirlar, goster_kolonlari, kolon_listesi, ucret_toplamlari


@app.route("/tahsilat-ciktisi")
@login_required
def tahsilat_ciktisi():
    yonlendirme = _filtre_durumu_uygula("tahsilat_ciktisi")
    if yonlendirme:
        return yonlendirme

    satirlar, goster_kolonlari, kolon_listesi, ucret_toplamlari = _tahsilat_ciktisi_satirlar()
    kolonlar_secili = request.args.getlist("kolon")
    db = get_db()
    _kl, _kb, sayisal_kolonlar, _ozel = _abone_kolon_takimi(db)
    kolon_secim_listesi = DISPLAY_KOLONLARI_ALFABETIK + [
        (k, e) for k, e in kolon_listesi if k not in _DISPLAY_KOLON_HARITASI
    ]

    deger_secili = {}
    haric_secili = {}
    for anahtar in goster_kolonlari:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")
        haric_secili[anahtar] = request.args.getlist(f"haric_{anahtar}")

    cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM abone")
    toplam_kayit = cur.fetchone()["c"]
    cur.execute("SELECT DISTINCT koy_adi FROM abone ORDER BY koy_adi")
    koyler = cur.fetchall()
    cur.close()

    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "tahsilat_ciktisi.html",
        satirlar=satirlar,
        kolon_listesi=kolon_listesi, goster_kolonlari=goster_kolonlari,
        kolon_secim_listesi=kolon_secim_listesi,
        secili_kolonlar=kolonlar_secili,
        koyler=koyler, secili_koy=request.args.get("koy", "").strip(),
        deger_secili=deger_secili, haric_secili=haric_secili,
        sayisal_kolonlar=sayisal_kolonlar,
        kolon_satir=_izgara_satir(len(kolon_secim_listesi)),
        kolon_satir_2=_izgara_satir(len(kolon_secim_listesi), 2),
        filtreli_kayit=filtreli_kayit, toplam_kayit=toplam_kayit,
        toplam_ucret=ucret_toplamlari["toplam_ucret"],
        tahsil_edilen_ucret=ucret_toplamlari["tahsil_edilen_ucret"],
        kalan_bakiye=ucret_toplamlari["kalan_bakiye"],
        sira=_sira_yonu_al(), sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
        tumunu_goster_qs=_tumunu_goster_qs(),
    )


@app.route("/tahsilat-ciktisi-excel")
@login_required
def tahsilat_ciktisi_excel():
    satirlar, goster_kolonlari, kolon_listesi, _ucret_toplamlari = _tahsilat_ciktisi_satirlar()
    tarih = datetime.now().strftime("%d_%m_%Y")
    return _csv_olustur(kolon_listesi, goster_kolonlari, satirlar, f"tahsilat_ciktisi_{tarih}.csv")


def _ariza_secenek_baglami(db):
    """Arıza formundaki iki onay kutusu listesini ve ızgara satır sayılarını hazırlar."""
    tespit = _form_secenekleri_getir(db, "tespit_edilen_ariza")
    islem = _form_secenekleri_getir(db, "yapilan_islemler")
    return dict(
        tespit_secenekleri=tespit,
        islem_secenekleri=islem,
        tespit_satir=_izgara_satir(len(tespit), 6),
        tespit_satir_2=_izgara_satir(len(tespit), 2),
        islem_satir=_izgara_satir(len(islem), 6),
        islem_satir_2=_izgara_satir(len(islem), 2),
    )


def _ariza_sonraki_s_no(db):
    """Yeni Arıza formunda gösterilen, KABACA bir öngörü."""
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM ariza")
    satir = cur.fetchone()
    cur.close()
    return (satir["c"] or 0) + 1


def _ariza_sira_numaralarini_yenile(db):
    """S.No, arıza kayıtlarında geliş tarihine göre (eskiden yeniye) sıralı
    tutuluyor ve bu sıraya göre 1'den başlayarak boşluksuz yeniden numaralandırılıyor."""
    cur = db.cursor()
    cur.execute(
        """
        UPDATE ariza a
        SET s_no = t.yeni_sira
        FROM (
            SELECT id, ROW_NUMBER() OVER (ORDER BY gelis_tarihi ASC NULLS LAST, id ASC) AS yeni_sira
            FROM ariza
        ) t
        WHERE a.id = t.id AND a.s_no IS DISTINCT FROM t.yeni_sira
        """
    )
    db.commit()
    cur.close()


def _ariza_kaydet(ariza_id):
    f = request.form
    ariza_ucret = _sayilastir(f.get("ariza_ucret"))
    alinan_ucret = _sayilastir(f.get("alinan_ucret"))
    tespit_metni = ", ".join(f.getlist("tespit_edilen_ariza"))
    islem_metni = ", ".join(f.getlist("yapilan_islemler"))

    alanlar = dict(
        ozel_s_no=f.get("ozel_s_no", "").strip(),
        koy_adi=f.get("koy_adi", "").strip(),
        yeni_seri_no=f.get("yeni_seri_no", "").strip(),
        seri_no=f.get("seri_no", "").strip(),
        adi=f.get("adi", "").strip(),
        soyadi=f.get("soyadi", "").strip(),
        telefon=_telefon_formatla(f.get("telefon", "")),
        telefon2=_telefon_formatla(f.get("telefon2", "")),
        ariza_ucret=ariza_ucret,
        alinan_ucret=alinan_ucret,
        gelis_tarihi=f.get("gelis_tarihi", "").strip(),
        takilan_tarih=f.get("takilan_tarih", "").strip(),
        teslim_tarihi=f.get("teslim_tarihi", "").strip(),
        sayac_kredisi=f.get("sayac_kredisi", "").strip(),
        tespit_edilen_ariza=tespit_metni,
        tespit_aciklama=f.get("tespit_aciklama", "").strip(),
        yapilan_islemler=islem_metni,
        islem_aciklama=f.get("islem_aciklama", "").strip(),
        konum_enlem=_konum_sayilastir(f.get("konum_enlem")),
        konum_boylam=_konum_sayilastir(f.get("konum_boylam")),
        kimlik_no=_kimlik_no_temizle(f.get("kimlik_no", "")),
        vergi_dairesi=f.get("vergi_dairesi", "").strip(),
        adres=f.get("adres", "").strip(),
        eposta=f.get("eposta", "").strip(),
    )

    db = get_db()
    for oa in _ozel_alanlari_getir(db, "ariza"):
        if oa["tur"] == "sayi":
            alanlar[oa["kolon_adi"]] = _sayilastir(f.get(oa["kolon_adi"]))
        else:
            alanlar[oa["kolon_adi"]] = f.get(oa["kolon_adi"], "").strip()

    cur = db.cursor()
    if ariza_id is None:
        kolonlar = ", ".join(alanlar.keys())
        yer_tutucular = ", ".join(["%s"] * len(alanlar))
        cur.execute(
            f"INSERT INTO ariza ({kolonlar}) VALUES ({yer_tutucular}) RETURNING id",
            list(alanlar.values()),
        )
        ariza_id = cur.fetchone()["id"]

        if alinan_ucret:
            tahsilat_tarihi = (
                f.get("teslim_tarihi", "").strip()
                or f.get("gelis_tarihi", "").strip()
                or datetime.now().strftime("%Y-%m-%d")
            )
            cur.execute(
                "INSERT INTO ariza_tahsilat (ariza_id, tarih, tutar, odeme_sekli, odemeyi_yapan, aciklama) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                (ariza_id, tahsilat_tarihi, alinan_ucret, "", "", "Arıza kaydı sırasında alınan"),
            )
    else:
        set_ifadesi = ", ".join([f"{k} = %s" for k in alanlar.keys()])
        cur.execute(f"UPDATE ariza SET {set_ifadesi}, updated_at = NOW() WHERE id = %s", list(alanlar.values()) + [ariza_id])
    db.commit()
    cur.close()
    _ariza_sira_numaralarini_yenile(db)
    return ariza_id


_FOTOGRAF_MAKS_BOYUT = 60 * 1024 * 1024

_VIDEO_IMZALARI = (
    (4, b"ftyp"),
    (0, b"\x1A\x45\xDF\xA3"),
    (0, b"RIFF"),
    (0, b"\x00\x00\x00\x0C\x6A\x50"),
)

_RESIM_FORMAT_MIME_HARITASI = {
    "JPEG": "image/jpeg", "PNG": "image/png", "GIF": "image/gif",
    "WEBP": "image/webp", "BMP": "image/bmp", "HEIF": "image/heif", "TIFF": "image/tiff",
}

_FOTOGRAF_MAKS_KENAR = 1920
_FOTOGRAF_JPEG_KALITE = 82


def _fotografi_kucult(icerik):
    """Yüklenen bir fotoğrafı ekranda göstermek için fazlasıyla yeterli ama
    çok daha küçük bir JPEG'e dönüştürmeyi dener."""
    try:
        from PIL import Image, ImageOps
        with Image.open(io.BytesIO(icerik)) as img:
            if img.format == "GIF":
                return icerik, None
            img = ImageOps.exif_transpose(img)
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            genislik, yukseklik = img.size
            en_uzun_kenar = max(genislik, yukseklik)
            if en_uzun_kenar > _FOTOGRAF_MAKS_KENAR:
                oran = _FOTOGRAF_MAKS_KENAR / en_uzun_kenar
                yeni_boyut = (max(1, round(genislik * oran)), max(1, round(yukseklik * oran)))
                img = img.resize(yeni_boyut, Image.LANCZOS)
            tampon = io.BytesIO()
            img.save(tampon, format="JPEG", quality=_FOTOGRAF_JPEG_KALITE, optimize=True)
            yeni_icerik = tampon.getvalue()
            if len(yeni_icerik) < len(icerik):
                return yeni_icerik, "image/jpeg"
            return icerik, None
    except Exception:
        return icerik, None


def _dosya_gercek_turu_dogrula(icerik, beyan_edilen_tur):
    """Yüklenen dosyanın GERÇEKTEN bir resim/video olup olmadığını dosyanın kendi
    baytlarına bakarak doğrular."""
    if beyan_edilen_tur.startswith("image/"):
        try:
            from PIL import Image
            with Image.open(io.BytesIO(icerik)) as img:
                img.verify()
            with Image.open(io.BytesIO(icerik)) as img2:
                return _RESIM_FORMAT_MIME_HARITASI.get(img2.format)
        except Exception:
            return None
    if beyan_edilen_tur.startswith("video/"):
        for konum, imza in _VIDEO_IMZALARI:
            if icerik[konum:konum + len(imza)] == imza:
                if imza == b"RIFF" and icerik[8:12] != b"AVI ":
                    continue
                return beyan_edilen_tur if beyan_edilen_tur.startswith("video/") else "video/mp4"
        return None
    return None


def _medya_kaydet(db, tablo, sahip_kolonu, sahip_id, dosyalar):
    """Arıza ve abone kayıtlarındaki fotoğraf/video yükleme kutuları aynı mantığı paylaşıyor."""
    cur = db.cursor()
    for dosya in dosyalar:
        if not dosya or not dosya.filename:
            continue
        icerik = dosya.read()
        if not icerik:
            continue
        if len(icerik) > _FOTOGRAF_MAKS_BOYUT:
            flash(f'"{dosya.filename}" dosyası çok büyük (60 MB üstü), yüklenmedi.')
            continue
        tur = dosya.mimetype or ""
        if not (tur.startswith("image/") or tur.startswith("video/")):
            flash(f'"{dosya.filename}" bir resim/video dosyası gibi görünmüyor, yüklenmedi.')
            continue
        dogrulanmis_tur = _dosya_gercek_turu_dogrula(icerik, tur)
        if not dogrulanmis_tur:
            flash(f'"{dosya.filename}" dosyasının içeriği bir resim/video ile eşleşmiyor, yüklenmedi.')
            continue
        if dogrulanmis_tur.startswith("image/"):
            icerik, kucultulmus_tur = _fotografi_kucult(icerik)
            if kucultulmus_tur:
                dogrulanmis_tur = kucultulmus_tur
        cur.execute(
            f"INSERT INTO {tablo} ({sahip_kolonu}, dosya_adi, content_type, icerik) VALUES (%s, %s, %s, %s)",
            (sahip_id, dosya.filename, dogrulanmis_tur, psycopg2.Binary(icerik)),
        )
    db.commit()
    cur.close()


def _ariza_fotograflarini_kaydet(db, ariza_id, dosyalar):
    _medya_kaydet(db, "ariza_fotograf", "ariza_id", ariza_id, dosyalar)


def _abone_fotograflarini_kaydet(db, abone_id, dosyalar):
    _medya_kaydet(db, "abone_fotograf", "abone_id", abone_id, dosyalar)


def _fabrika_fotograflarini_kaydet(db, kayit_id, dosyalar):
    _medya_kaydet(db, "fabrika_fotograf", "kayit_id", kayit_id, dosyalar)


@app.route("/admin/secenek-yonetimi")
@login_required
def secenek_yonetimi():
    db = get_db()
    cur = db.cursor()
    gruplar = {}
    for anahtar, baslik in FORM_SECENEK_GRUPLARI.items():
        cur.execute(
            "SELECT id, deger, sira FROM form_secenegi WHERE grup = %s ORDER BY sira, id",
            (anahtar,),
        )
        gruplar[anahtar] = {"baslik": baslik, "secenekler": cur.fetchall()}
    cur.close()
    return render_template(
        "secenek_yonetimi.html",
        gruplar=gruplar, grup_sirasi=list(FORM_SECENEK_GRUPLARI.keys()),
    )


@app.route("/admin/secenek-yonetimi/ekle", methods=["POST"])
@login_required
def secenek_yonetimi_ekle():
    grup = request.form.get("grup", "")
    deger = request.form.get("deger", "").strip()
    if grup not in FORM_SECENEK_GRUPLARI:
        flash("Geçersiz seçenek grubu.")
        return redirect(url_for("secenek_yonetimi"))
    if not deger:
        flash("Boş seçenek eklenemez.")
        return redirect(url_for("secenek_yonetimi"))

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT id FROM form_secenegi WHERE grup = %s AND LOWER(deger) = LOWER(%s)",
        (grup, deger),
    )
    if cur.fetchone():
        flash(f'"{deger}" zaten bu listede var.')
        cur.close()
        return redirect(url_for("secenek_yonetimi"))

    cur.execute("SELECT COALESCE(MAX(sira), -1) AS m FROM form_secenegi WHERE grup = %s", (grup,))
    sonraki_sira = cur.fetchone()["m"] + 1
    cur.execute(
        "INSERT INTO form_secenegi (grup, deger, sira) VALUES (%s, %s, %s)",
        (grup, deger, sonraki_sira),
    )
    db.commit()
    cur.close()
    flash(f'"{deger}" eklendi.')
    return redirect(url_for("secenek_yonetimi"))


@app.route("/admin/secenek-yonetimi/duzenle/<int:secenek_id>", methods=["POST"])
@login_required
def secenek_yonetimi_duzenle(secenek_id):
    yeni_deger = request.form.get("deger", "").strip()
    if not yeni_deger:
        flash("Boş değer kaydedilemez.")
        return redirect(url_for("secenek_yonetimi"))
    db = get_db()
    cur = db.cursor()
    cur.execute("UPDATE form_secenegi SET deger = %s WHERE id = %s", (yeni_deger, secenek_id))
    db.commit()
    cur.close()
    flash("Seçenek güncellendi. (Daha önce kaydedilmiş arıza kayıtlarındaki eski metin değişmeden kalır.)")
    return redirect(url_for("secenek_yonetimi"))


@app.route("/admin/secenek-yonetimi/sil/<int:secenek_id>", methods=["POST"])
@login_required
def secenek_yonetimi_sil(secenek_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM form_secenegi WHERE id = %s", (secenek_id,))
    db.commit()
    cur.close()
    flash("Seçenek silindi.")
    return redirect(url_for("secenek_yonetimi"))


@app.route("/admin/secenek-yonetimi/sirala", methods=["POST"])
@login_required
def secenek_yonetimi_sirala():
    """Onay Kutusu Ayarları'ndaki sürükle-bırak listesi için sıralama günceller."""
    veri = request.get_json(silent=True) or {}
    grup = veri.get("grup", "")
    sira_listesi = veri.get("sira", [])
    if grup not in FORM_SECENEK_GRUPLARI or not isinstance(sira_listesi, list):
        return jsonify({"hata": "geçersiz istek"}), 400
    db = get_db()
    cur = db.cursor()
    for i, secenek_id in enumerate(sira_listesi):
        try:
            secenek_id = int(secenek_id)
        except (TypeError, ValueError):
            continue
        cur.execute(
            "UPDATE form_secenegi SET sira = %s WHERE id = %s AND grup = %s",
            (i, secenek_id, grup),
        )
    db.commit()
    cur.close()
    return jsonify({"tamam": True})


def _isim_normalize(metin):
    """Karşılaştırma için sadece baştaki/sondaki ve kelimeler arasındaki fazla boşlukları temizler."""
    return re.sub(r"\s+", " ", (metin or "")).strip()


def _isim_grup_anahtari(metin):
    """Gruplama anahtarı için boşluk + büyük/küçük harf farkını yok sayarak normalize eder."""
    return re.sub(r"\s+", " ", (metin or "")).strip().upper()


@app.route("/admin/isim-tutarlilik")
@login_required
def isim_tutarlilik():
    """Abone Listesi, Köy Abone Listesi ve Arıza Takip'te aynı sayaç/cihaz
    numarasına ve aynı köy adına sahip olduğu halde adı/soyadı yazılışı
    farklı olan kayıtları bulup listeler."""
    db = get_db()
    cur = db.cursor()

    kayitlar = []

    cur.execute("SELECT id, koy_adi, sayac_no, adi, soyadi FROM abone")
    for r in cur.fetchall():
        if (r["sayac_no"] or "").strip() and (r["koy_adi"] or "").strip():
            kayitlar.append({
                "kaynak": "abone", "kaynak_etiket": "Abone Listesi",
                "id": r["id"], "koy_adi": r["koy_adi"], "no": r["sayac_no"],
                "no_etiket": "Sayaç No", "adi": r["adi"], "soyadi": r["soyadi"],
            })

    cur.execute("SELECT id, koy_adi, cihaz_no, adi, soyadi FROM koy_abone")
    for r in cur.fetchall():
        if (r["cihaz_no"] or "").strip() and (r["koy_adi"] or "").strip():
            kayitlar.append({
                "kaynak": "koy_abone", "kaynak_etiket": "Köy Abone Listesi",
                "id": r["id"], "koy_adi": r["koy_adi"], "no": r["cihaz_no"],
                "no_etiket": "Cihaz No", "adi": r["adi"], "soyadi": r["soyadi"],
            })

    cur.execute("SELECT id, koy_adi, seri_no, yeni_seri_no, adi, soyadi FROM ariza")
    for r in cur.fetchall():
        koy = r["koy_adi"]
        if not (koy or "").strip():
            continue
        adaylar = []
        if (r["seri_no"] or "").strip():
            adaylar.append(("Seri No", r["seri_no"]))
        if (r["yeni_seri_no"] or "").strip() and r["yeni_seri_no"] != r["seri_no"]:
            adaylar.append(("Yeni Seri No", r["yeni_seri_no"]))
        for no_etiket, no_degeri in adaylar:
            kayitlar.append({
                "kaynak": "ariza", "kaynak_etiket": "Arıza Takip",
                "id": r["id"], "koy_adi": koy, "no": no_degeri,
                "no_etiket": no_etiket, "adi": r["adi"], "soyadi": r["soyadi"],
            })

    cur.close()

    gruplar = {}
    for k in kayitlar:
        anahtar = (_isim_grup_anahtari(k["koy_adi"]), _isim_grup_anahtari(k["no"]))
        gruplar.setdefault(anahtar, []).append(k)

    tutarsiz_gruplar = []
    for (koy_anahtar, no_anahtar), grup_kayitlari in gruplar.items():
        if len(grup_kayitlari) < 2:
            continue
        isimler = set()
        for k in grup_kayitlari:
            isimler.add(_isim_normalize(k["adi"]) + " " + _isim_normalize(k["soyadi"]))
        if len(isimler) > 1:
            tutarsiz_gruplar.append({
                "koy_adi": grup_kayitlari[0]["koy_adi"],
                "no": grup_kayitlari[0]["no"],
                "kayitlar": sorted(grup_kayitlari, key=lambda k: k["kaynak_etiket"]),
            })

    tutarsiz_gruplar.sort(key=lambda g: (_isim_grup_anahtari(g["koy_adi"]), _isim_grup_anahtari(g["no"])))

    return render_template(
        "isim_tutarlilik.html",
        gruplar=tutarsiz_gruplar,
        taranan_kayit_sayisi=len(kayitlar),
    )


@app.route("/admin/ozel-alan-ayarlari")
@login_required
def ozel_alan_ayarlari():
    """Abone/Arıza formuna kod yazmadan yeni bir bilgi kutusu eklenebildiği ayarlar ekranı."""
    db = get_db()
    abone_alanlari = _ozel_alanlari_getir(db, "abone")
    ariza_alanlari = _ozel_alanlari_getir(db, "ariza")
    return render_template(
        "ozel_alan_ayarlari.html",
        abone_sirasi=_form_onizleme_sirasi(ABONE_FORM_ALAN_SIRASI, abone_alanlari),
        ariza_sirasi=_form_onizleme_sirasi(ARIZA_FORM_ALAN_SIRASI, ariza_alanlari),
        tur_etiketleri=_OZEL_ALAN_TUR_ETIKETLERI,
    )


@app.route("/admin/ozel-alan-ayarlari/ekle", methods=["POST"])
@login_required
def ozel_alan_ayarlari_ekle():
    tablo = request.form.get("tablo", "")
    etiket = request.form.get("etiket", "").strip()
    tur = request.form.get("tur", "")
    if tablo not in ("abone", "ariza"):
        flash("Geçersiz tablo.")
        return redirect(url_for("ozel_alan_ayarlari"))
    if not etiket:
        flash("Boş alan adı eklenemez.")
        return redirect(url_for("ozel_alan_ayarlari"))
    if tur not in _OZEL_ALAN_TUR_PG:
        flash("Geçersiz alan türü.")
        return redirect(url_for("ozel_alan_ayarlari"))

    db = get_db()
    _ozel_alan_ekle(db, tablo, etiket, tur)
    flash(f'"{etiket}" alanı eklendi.')
    return redirect(url_for("ozel_alan_ayarlari"))


@app.route("/admin/ozel-alan-ayarlari/sil/<int:ozel_alan_id>", methods=["POST"])
@login_required
def ozel_alan_ayarlari_sil(ozel_alan_id):
    db = get_db()
    _ozel_alan_sil(db, ozel_alan_id)
    flash("Alan kaldırıldı. (Daha önce bu alana girilmiş veriler kaybolmadı, sadece form/listeden gizlendi.)")
    return redirect(url_for("ozel_alan_ayarlari"))


@app.route("/admin/ozel-alan-ayarlari/sirala", methods=["POST"])
@login_required
def ozel_alan_ayarlari_sirala():
    """Özel Alan Ayarları'ndaki sürükle-bırak önizlemesi için sıralamayı kaydeder."""
    veri = request.get_json(silent=True) or {}
    tablo = veri.get("tablo", "")
    sira_listesi = veri.get("sira", [])
    if tablo not in ("abone", "ariza") or not isinstance(sira_listesi, list):
        return jsonify({"hata": "geçersiz istek"}), 400
    db = get_db()
    _ozel_alan_sirala(db, tablo, [str(a) for a in sira_listesi])
    return jsonify({"tamam": True})


@app.route("/ariza/yeni", methods=["GET", "POST"])
@login_required
def ariza_yeni():
    if request.method == "POST":
        yeni_id = _ariza_kaydet(None)
        db = get_db()
        _ariza_fotograflarini_kaydet(db, yeni_id, request.files.getlist("fotograflar"))
        return redirect(url_for("ariza_listesi"))
    db = get_db()
    return render_template(
        "ariza_form.html", kayit=None,
        sonraki_s_no=_ariza_sonraki_s_no(db),
        **_ariza_secenek_baglami(db),
        secili_tespit=set(), secili_islem=set(),
        ilk_montaj_tarihi="",
        bugun=datetime.now().strftime("%Y-%m-%d"),
        fotograflar=[],
        ozel_alan_harita=_ozel_alan_harita(_ozel_alanlari_getir(db, "ariza")),
    )


@app.route("/ariza/<int:ariza_id>/duzenle", methods=["GET", "POST"])
@login_required
def ariza_duzenle(ariza_id):
    db = get_db()
    sonraki_hedef = request.args.get("hedef", "") or request.form.get("hedef", "")
    if request.method == "POST":
        _ariza_kaydet(ariza_id)
        _ariza_fotograflarini_kaydet(db, ariza_id, request.files.getlist("fotograflar"))
        if sonraki_hedef:
            return redirect(sonraki_hedef)
        flash("Kayıt kaydedildi.")
        return redirect(url_for("ariza_duzenle", ariza_id=ariza_id))
    cur = db.cursor()
    cur.execute("SELECT * FROM ariza WHERE id = %s", (ariza_id,))
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("ariza_listesi"))
    secili_tespit = set((kayit["tespit_edilen_ariza"] or "").split(", ")) if kayit["tespit_edilen_ariza"] else set()
    secili_islem = set((kayit["yapilan_islemler"] or "").split(", ")) if kayit["yapilan_islemler"] else set()

    ilk_montaj_tarihi = ""
    if kayit["seri_no"]:
        cur = db.cursor()
        cur.execute(
            "SELECT montaj_tarihi FROM abone WHERE sayac_no = %s ORDER BY id LIMIT 1",
            (kayit["seri_no"],),
        )
        abone_satir = cur.fetchone()
        cur.close()
        if abone_satir:
            ilk_montaj_tarihi = _tarih_iso_hale_getir(abone_satir["montaj_tarihi"])

    cur = db.cursor()
    cur.execute(
        "SELECT id, dosya_adi, content_type FROM ariza_fotograf WHERE ariza_id = %s ORDER BY id",
        (ariza_id,),
    )
    fotograflar = cur.fetchall()
    cur.close()

    return render_template(
        "ariza_form.html", kayit=kayit, hedef=sonraki_hedef,
        **_ariza_secenek_baglami(db),
        secili_tespit=secili_tespit, secili_islem=secili_islem,
        ilk_montaj_tarihi=ilk_montaj_tarihi,
        bugun=datetime.now().strftime("%Y-%m-%d"),
        fotograflar=fotograflar,
        ozel_alan_harita=_ozel_alan_harita(_ozel_alanlari_getir(db, "ariza")),
    )


@app.route("/ariza/<int:ariza_id>/sil", methods=["POST"])
@login_required
def ariza_sil(ariza_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM ariza WHERE id = %s", (ariza_id,))
    db.commit()
    cur.close()
    _ariza_sira_numaralarini_yenile(db)
    return redirect(url_for("ariza_listesi"))


@app.route("/ariza-fotograf/<int:foto_id>")
@login_required
def ariza_fotograf_goster(foto_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT content_type, icerik FROM ariza_fotograf WHERE id = %s", (foto_id,))
    foto = cur.fetchone()
    cur.close()
    if foto is None:
        return "Fotoğraf bulunamadı.", 404
    yanit = Response(bytes(foto["icerik"]), mimetype=foto["content_type"] or "application/octet-stream")
    yanit.headers["X-Content-Type-Options"] = "nosniff"
    return yanit


@app.route("/ariza-fotograf/<int:foto_id>/sil", methods=["POST"])
@login_required
def ariza_fotograf_sil(foto_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT ariza_id FROM ariza_fotograf WHERE id = %s", (foto_id,))
    foto = cur.fetchone()
    if foto:
        cur.execute("DELETE FROM ariza_fotograf WHERE id = %s", (foto_id,))
        db.commit()
    cur.close()
    if foto:
        return redirect(url_for("ariza_duzenle", ariza_id=foto["ariza_id"]))
    return redirect(url_for("ariza_listesi"))


@app.route("/ariza")
@login_required
def ariza_listesi():
    yonlendirme = _filtre_durumu_uygula("ariza_listesi")
    if yonlendirme:
        return yonlendirme

    db = get_db()
    q = request.args.get("q", "").strip()
    koy = request.args.get("koy", "").strip()
    alanlar_secili = request.args.getlist("alan")
    ARIZA_ALAN_HARITASI = {k: (kolon, sayisal) for k, _, kolon, sayisal in ARIZA_ALAN_TANIMLARI}
    alan_listesi = [(k, etiket) for k, etiket, _, _ in ARIZA_ALAN_TANIMLARI]

    sql = "SELECT * FROM ariza WHERE 1=1"
    params = []

    if koy:
        sql += " AND koy_adi = %s"
        params.append(koy)

    if q:
        secili = alanlar_secili if alanlar_secili else [k for k, *_ in ARIZA_ALAN_TANIMLARI]

        q_sayi = None
        q_temiz = q.replace(",", ".").strip()
        try:
            q_sayi = float(q_temiz)
        except ValueError:
            q_sayi = None

        kosul_listesi = []
        kosul_params = []
        for s in secili:
            if s in ARIZA_ALAN_HARITASI:
                kolon, sayisal = ARIZA_ALAN_HARITASI[s]
                if sayisal and q_sayi is not None:
                    kosul_listesi.append(f"ROUND(CAST({kolon} AS NUMERIC), 2) = %s")
                    kosul_params.append(round(q_sayi, 2))
                elif sayisal:
                    kosul_listesi.append(f"{_turkce_esle_kosul(f'CAST({kolon} AS TEXT)')} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
                else:
                    kosul_listesi.append(f"{_turkce_esle_kosul(kolon)} LIKE %s")
                    kosul_params.append(_turkce_normallestir(f"%{q}%"))
        if kosul_listesi:
            sql += " AND (" + " OR ".join(kosul_listesi) + ")"
            params += kosul_params

    kolon_listesi, kolon_bilgi, sayisal_kolonlar, ozel_alanlar = _ariza_kolon_takimi(db)
    deger_secili = {}
    haric_secili = {}
    for anahtar, _ in kolon_listesi:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")
        haric_secili[anahtar] = request.args.getlist(f"haric_{anahtar}")
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi
    sql += f" ORDER BY s_no {'DESC' if _sira_yonu_al() == 'desc' else 'ASC'}"

    cur = db.cursor()
    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()
    cur.execute("SELECT DISTINCT koy_adi FROM ariza ORDER BY koy_adi")
    koyler = cur.fetchall()
    cur.execute("SELECT COUNT(*) AS c FROM ariza")
    toplam_kayit = cur.fetchone()["c"]
    cur.close()

    satirlar = [_ariza_satir_sozlugu(k, ozel_alanlar) for k in kayitlar_ham]

    toplam_ariza_ucreti = sum(float(k["ariza_ucret"] or 0) for k in kayitlar_ham)
    tahsil_edilen_ucret = sum(float(k["alinan_ucret"] or 0) for k in kayitlar_ham)
    kalan_bakiye = toplam_ariza_ucreti - tahsil_edilen_ucret

    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "ariza_listesi.html", satirlar=satirlar,
        kolon_listesi=kolon_listesi,
        q=q, koyler=koyler, secili_koy=koy, secili_alanlar=alanlar_secili, alan_listesi=alan_listesi,
        deger_secili=deger_secili, haric_secili=haric_secili,
        sayisal_kolonlar=sayisal_kolonlar,
        arama_satir=_izgara_satir(len(alan_listesi)),
        arama_satir_2=_izgara_satir(len(alan_listesi), 2),
        filtreli_kayit=filtreli_kayit, toplam_kayit=toplam_kayit,
        toplam_ariza_ucreti=toplam_ariza_ucreti,
        tahsil_edilen_ucret=tahsil_edilen_ucret,
        kalan_bakiye=kalan_bakiye,
        sira=_sira_yonu_al(), sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
    )


def _ariza_ciktisi_satirlar():
    db = get_db()
    kolon_listesi, kolon_bilgi, _sayisal, ozel_alanlar = _ariza_kolon_takimi(db)
    kolonlar_secili = request.args.getlist("kolon")
    goster_kolonlari = kolonlar_secili if kolonlar_secili else [k for k, _ in kolon_listesi]
    koy = request.args.get("koy", "").strip()
    sql = "SELECT * FROM ariza WHERE 1=1"
    params = []
    if koy:
        sql += " AND koy_adi = %s"
        params.append(koy)
    for anahtar in goster_kolonlari:
        kosul, param_listesi = _kolon_secim_kosulu(anahtar, kolon_bilgi)
        if kosul:
            sql += f" AND {kosul}"
            params += param_listesi
    sql += f" ORDER BY s_no {'DESC' if _sira_yonu_al() == 'desc' else 'ASC'}"
    cur = db.cursor()
    cur.execute(sql, params)
    kayitlar_ham = cur.fetchall()
    cur.close()
    satirlar = [_ariza_satir_sozlugu(k, ozel_alanlar) for k in kayitlar_ham]

    toplam_ucret = sum(float(k["ariza_ucret"] or 0) for k in kayitlar_ham)
    tahsil_edilen_ucret = sum(float(k["alinan_ucret"] or 0) for k in kayitlar_ham)
    ucret_toplamlari = {
        "toplam_ucret": toplam_ucret,
        "tahsil_edilen_ucret": tahsil_edilen_ucret,
        "kalan_bakiye": toplam_ucret - tahsil_edilen_ucret,
    }
    return satirlar, goster_kolonlari, kolon_listesi, ucret_toplamlari


@app.route("/ariza-ciktisi")
@login_required
def ariza_ciktisi():
    yonlendirme = _filtre_durumu_uygula("ariza_ciktisi")
    if yonlendirme:
        return yonlendirme

    satirlar, goster_kolonlari, kolon_listesi, ucret_toplamlari = _ariza_ciktisi_satirlar()
    kolonlar_secili = request.args.getlist("kolon")
    db = get_db()

    kolon_secim_listesi = ARIZA_DISPLAY_KOLONLARI_ALFABETIK + [
        (k, e) for k, e in kolon_listesi if k not in _ARIZA_DISPLAY_KOLON_HARITASI
    ]
    _kl, _kb, sayisal_kolonlar, _ozel = _ariza_kolon_takimi(db)

    deger_secili = {}
    haric_secili = {}
    for anahtar in goster_kolonlari:
        deger_secili[anahtar] = request.args.getlist(f"deger_{anahtar}")
        haric_secili[anahtar] = request.args.getlist(f"haric_{anahtar}")

    cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM ariza")
    toplam_kayit = cur.fetchone()["c"]
    cur.execute("SELECT DISTINCT koy_adi FROM ariza ORDER BY koy_adi")
    koyler = cur.fetchall()
    cur.close()

    satirlar, filtreli_kayit, sayfa, toplam_sayfa = _sayfala(satirlar)

    return render_template(
        "ariza_ciktisi.html",
        satirlar=satirlar,
        kolon_listesi=kolon_listesi, goster_kolonlari=goster_kolonlari,
        kolon_secim_listesi=kolon_secim_listesi,
        secili_kolonlar=kolonlar_secili,
        koyler=koyler, secili_koy=request.args.get("koy", "").strip(),
        deger_secili=deger_secili, haric_secili=haric_secili,
        sayisal_kolonlar=sayisal_kolonlar,
        kolon_satir=_izgara_satir(len(kolon_secim_listesi)),
        kolon_satir_2=_izgara_satir(len(kolon_secim_listesi), 2),
        filtreli_kayit=filtreli_kayit, toplam_kayit=toplam_kayit,
        toplam_ucret=ucret_toplamlari["toplam_ucret"],
        tahsil_edilen_ucret=ucret_toplamlari["tahsil_edilen_ucret"],
        kalan_bakiye=ucret_toplamlari["kalan_bakiye"],
        sira=_sira_yonu_al(), sira_toggle_qs=_sira_toggle_qs(),
        sayfa=sayfa, toplam_sayfa=toplam_sayfa, sayfalama_qs=_sayfalama_qs,
        tumunu_goster_qs=_tumunu_goster_qs(),
    )


@app.route("/ariza-ciktisi-excel")
@login_required
def ariza_ciktisi_excel():
    satirlar, goster_kolonlari, kolon_listesi, _ucret_toplamlari = _ariza_ciktisi_satirlar()
    tarih = datetime.now().strftime("%d_%m_%Y")
    return _csv_olustur(kolon_listesi, goster_kolonlari, satirlar, f"ariza_ciktisi_{tarih}.csv")


@app.route("/ariza/<int:ariza_id>/tahsilat", methods=["GET", "POST"])
@login_required
def ariza_tahsilat(ariza_id):
    db = get_db()
    cur = db.cursor()

    if request.method == "POST":
        tutar = _sayilastir(request.form.get("tutar"))
        tarih = request.form.get("tarih", "").strip()
        odeme_sekli = request.form.get("odeme_sekli", "").strip()
        odemeyi_yapan = request.form.get("odemeyi_yapan", "").strip()
        aciklama = request.form.get("aciklama", "").strip()

        if tutar:
            cur.execute(
                "INSERT INTO ariza_tahsilat (ariza_id, tarih, tutar, odeme_sekli, odemeyi_yapan, aciklama) VALUES (%s, %s, %s, %s, %s, %s)",
                (ariza_id, tarih, tutar, odeme_sekli, odemeyi_yapan, aciklama),
            )
            cur.execute("UPDATE ariza SET alinan_ucret = alinan_ucret + %s WHERE id = %s", (tutar, ariza_id))
            db.commit()

        cur.close()
        return redirect(url_for("ariza_tahsilat", ariza_id=ariza_id))

    cur.execute("SELECT * FROM ariza WHERE id = %s", (ariza_id,))
    kayit = cur.fetchone()
    if kayit is None:
        cur.close()
        flash("Kayıt bulunamadı.")
        return redirect(url_for("ariza_listesi"))

    cur.execute("SELECT * FROM ariza_tahsilat WHERE ariza_id = %s ORDER BY tarih DESC, id DESC", (ariza_id,))
    tahsilatlar = cur.fetchall()
    cur.close()

    return render_template(
        "ariza_tahsilat.html", kayit=kayit, tahsilatlar=tahsilatlar,
        bugun=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/ariza/<int:ariza_id>/fatura-kes", methods=["GET", "POST"])
@login_required
def ariza_fatura_kes(ariza_id):
    """Arıza kaydındaki Arıza Ücreti üzerinden e-Fatura/e-Arşiv Fatura keser."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM ariza WHERE id = %s", (ariza_id,))
    ariza = cur.fetchone()
    cur.close()
    if ariza is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("ariza_listesi"))

    onerilen_tur = _fatura_turu_belirle(ariza["kimlik_no"])

    if request.method == "POST":
        fatura_turu = request.form.get("fatura_turu")
        if fatura_turu not in ("earsiv", "efatura"):
            flash("Geçerli bir fatura türü seçin.")
            return redirect(url_for("ariza_fatura_kes", ariza_id=ariza_id))
        if not ariza["kimlik_no"] or not ariza["adres"]:
            flash("Fatura kesmeden önce arıza kaydının TC Kimlik No/Vergi No ve Adres bilgilerini doldurun.")
            return redirect(url_for("ariza_duzenle", ariza_id=ariza_id))

        kalemler = _fatura_kalemlerini_formdan_oku(request.form)
        if not kalemler:
            flash("En az bir fatura kalemi girilmeli (açıklama, miktar ve birim fiyat).")
            return redirect(url_for("ariza_fatura_kes", ariza_id=ariza_id))

        fatura_tarihi_form = request.form.get("fatura_tarihi", "").strip()
        try:
            fatura_tarihi_dt = datetime.strptime(fatura_tarihi_form, "%Y-%m-%d")
        except ValueError:
            flash("Geçerli bir fatura tarihi seçin.")
            return redirect(url_for("ariza_fatura_kes", ariza_id=ariza_id))
        if fatura_tarihi_dt.date() > datetime.now().date():
            flash("Fatura tarihi bugünden ileri bir tarih olamaz.")
            return redirect(url_for("ariza_fatura_kes", ariza_id=ariza_id))

        fatura_id = _hizli_fatura_gonder(
            db, "ariza", ariza_id, dict(ariza), kalemler, fatura_turu,
            session.get("kullanici_adi", ""), fatura_tarihi=fatura_tarihi_form,
        )
        return redirect(url_for("fatura_goruntule", fatura_id=fatura_id))

    onizleme = []
    if float(ariza["ariza_ucret"] or 0) > 0:
        taban, _kdv = _hizli_kdv_ayir(float(ariza["ariza_ucret"]))
        onizleme.append({"aciklama": "ARIZA TAMİR ÜCRETİ", "miktar": 1, "birim_fiyat": taban})

    fatura_kes_url = url_for("ariza_fatura_kes", ariza_id=ariza_id)
    return render_template(
        "fatura_kes.html", kaynak=ariza, kaynak_tur="ariza", kaynak_ad=f"{ariza['adi']} {ariza['soyadi']}",
        duzenle_url=url_for("ariza_duzenle", ariza_id=ariza_id) + f"?hedef={_url_quote(fatura_kes_url, safe='')}",
        geri_url=fatura_kes_url,
        onizleme=onizleme,
        onerilen_tur=onerilen_tur, hizli_ayarli_mi=_hizli_ayarli_mi(), geri="",
        bugun=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/ariza/<int:ariza_id>/mesaj-gonder", methods=["GET", "POST"])
@login_required
def ariza_mesaj_gonder(ariza_id):
    """Bir arıza kaydına WhatsApp/SMS/E-posta mesajı gönderir."""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM ariza WHERE id = %s", (ariza_id,))
    ariza = cur.fetchone()
    cur.close()
    if ariza is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("ariza_listesi"))

    ad = f"{ariza['adi']} {ariza['soyadi']}"

    if request.method == "POST":
        kanal = request.form.get("kanal", "whatsapp")
        icerik = request.form.get("icerik", "").strip()
        if not icerik:
            flash("Mesaj içeriği boş olamaz.")
            return redirect(url_for("ariza_mesaj_gonder", ariza_id=ariza_id))
        if kanal == "eposta":
            eposta = ariza["eposta"]
            if not eposta:
                flash("Bu arızaya kayıtlı bir e-posta adresi yok, mesaj gönderilemedi.")
                return redirect(url_for("ariza_duzenle", ariza_id=ariza_id))
            _mesaj_gonder(
                db, "ariza", ariza_id, ad, None, kanal, icerik,
                session.get("kullanici_adi", ""), alici_eposta=eposta,
            )
        else:
            telefon = ariza["telefon"] or ariza["telefon2"]
            if not telefon:
                flash("Bu arızaya kayıtlı bir telefon numarası yok, mesaj gönderilemedi.")
                return redirect(url_for("ariza_duzenle", ariza_id=ariza_id))
            _mesaj_gonder(
                db, "ariza", ariza_id, ad, telefon, kanal, icerik,
                session.get("kullanici_adi", ""),
            )
        flash("Mesaj gönderim isteği oluşturuldu, sonucunu 'Mesajlarım' sayfasından görebilirsiniz.")
        return redirect(url_for("mesaj_listesi"))

    hedefler = [{
        "id": ariza["id"], "ad": ad,
        "telefon": ariza["telefon"] or ariza["telefon2"] or "",
        "eposta": ariza["eposta"] or "",
    }]
    return render_template(
        "mesaj_gonder.html", hedefler=hedefler, tekli=True,
        baslik=f"Mesaj Gönder - {ad}",
        gonder_url=url_for("ariza_mesaj_gonder", ariza_id=ariza_id),
        geri_url=url_for("ariza_duzenle", ariza_id=ariza_id),
        liste_url=url_for("ariza_listesi"), netgsm_ayarli_mi=_netgsm_ayarli_mi(),
        sms_hazir=_netgsm_sms_ayarli_mi(), eposta_hazir=_eposta_ayarli_mi(), geri="",
    )


@app.route("/ariza-tahsilat/<int:tahsilat_id>/sil", methods=["POST"])
@login_required
def ariza_tahsilat_sil(tahsilat_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT ariza_id, tutar FROM ariza_tahsilat WHERE id = %s", (tahsilat_id,))
    kayit = cur.fetchone()
    ariza_id = None
    if kayit:
        ariza_id = kayit["ariza_id"]
        cur.execute("UPDATE ariza SET alinan_ucret = alinan_ucret - %s WHERE id = %s", (kayit["tutar"], ariza_id))
        cur.execute("DELETE FROM ariza_tahsilat WHERE id = %s", (tahsilat_id,))
        db.commit()
    cur.close()
    if ariza_id:
        return redirect(url_for("ariza_tahsilat", ariza_id=ariza_id))
    return redirect(url_for("ariza_listesi"))


@app.route("/ariza-tahsilat/<int:tahsilat_id>/makbuz")
@login_required
def ariza_tahsilat_makbuz(tahsilat_id):
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT t.*, a.adi AS ariza_adi, a.soyadi AS ariza_soyadi, a.seri_no AS ariza_seri_no "
        "FROM ariza_tahsilat t JOIN ariza a ON a.id = t.ariza_id WHERE t.id = %s",
        (tahsilat_id,),
    )
    kayit = cur.fetchone()
    cur.close()
    if kayit is None:
        flash("Kayıt bulunamadı.")
        return redirect(url_for("ariza_listesi"))

    ad_soyad = f"{kayit['ariza_adi']} {kayit['ariza_soyadi']}"
    odemeyi_yapan = kayit["odemeyi_yapan"] or ad_soyad
    not_goster = bool(kayit["odemeyi_yapan"]) and kayit["odemeyi_yapan"].strip().upper() != ad_soyad.strip().upper()

    return render_template(
        "makbuz.html",
        sira_no=str(kayit["id"]).zfill(6),
        tarih=_gg_aa_yyyy(kayit["tarih"]),
        sayin=ad_soyad,
        tutar=kayit["tutar"],
        tutar_yazi=_tutar_yaziya_cevir(kayit["tutar"]),
        odeme_esles=_odeme_sekli_esle(kayit["odeme_sekli"]),
        odeme_sekli_metin=kayit["odeme_sekli"],
        aciklama_basligi="Arıza Ücreti Ödemesi",
        seri_no=kayit["ariza_seri_no"],
        odemeyi_yapan=odemeyi_yapan,
        ad_soyad=ad_soyad,
        not_goster=not_goster,
        geri_url=url_for("ariza_tahsilat", ariza_id=kayit["ariza_id"]),
    )


@app.route("/admin/toplu-abone-yukle", methods=["GET", "POST"])
@login_required
def toplu_abone_yukle():
    veri_dosyasi = os.path.join(os.path.dirname(os.path.abspath(__file__)), "abone_toplu_veri.b64")
    if not os.path.exists(veri_dosyasi):
        return (
            "Veri dosyası (abone_toplu_veri.b64) bulunamadı. "
            "Bu dosyanın app.py ile aynı klasörde (repo kök dizininde) olduğundan emin olun.",
            404,
        )

    with open(veri_dosyasi, "rb") as f:
        b64_veri = f.read()
    csv_metin = gzip.decompress(base64.b64decode(b64_veri)).decode("utf-8-sig")
    satirlar = list(csv.DictReader(io.StringIO(csv_metin)))

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM abone")
    mevcut_sayi = cur.fetchone()["c"]

    onay = request.method == "POST" and request.form.get("onayla") == "1"
    zorla = request.method == "POST" and request.form.get("zorla") == "1"
    _csrf_gizli_alan = f'<input type="hidden" name="csrf_token" value="{generate_csrf()}">'

    if not onay:
        if mevcut_sayi > 50:
            aksiyon = (
                f"<p style='color:#b00;font-weight:bold'>Dikkat: tabloda hâlihazırda {mevcut_sayi} kayıt var. "
                f"Bu işlem mevcut kayıtları SİLMEZ, üzerine {len(satirlar)} yeni kayıt EKLER. "
                f"Bu veriyi daha önce yüklediyseniz tekrar yüklemeyin, kayıtlar çiftlenir.</p>"
                f"<form method='post'>{_csrf_gizli_alan}"
                f"<input type='hidden' name='onayla' value='1'><input type='hidden' name='zorla' value='1'>"
                f"<button type='submit' style='font-size:20px;color:#b00;background:none;border:none;"
                f"text-decoration:underline;cursor:pointer;padding:0;'>"
                f"Yine de devam et ve {len(satirlar)} kaydı ekle</button></form>"
            )
        else:
            aksiyon = (
                f"<form method='post'>{_csrf_gizli_alan}"
                f"<input type='hidden' name='onayla' value='1'>"
                f"<button type='submit' style='font-size:20px;background:none;border:none;"
                f"text-decoration:underline;cursor:pointer;padding:0;'>"
                f"Evet, {len(satirlar)} kaydı içe aktar</button></form>"
            )
        cur.close()
        return f"""
        <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
        <h2>Toplu Abone Yükleme</h2>
        <p>Excel dosyasından hazırlanan <b>{len(satirlar)}</b> abone kaydı,
        veritabanındaki <b>abone</b> tablosuna eklenmeye hazır.</p>
        <p>Şu anda tabloda <b>{mevcut_sayi}</b> kayıt bulunuyor.</p>
        {aksiyon}
        </body></html>
        """

    if mevcut_sayi > 50 and not zorla:
        cur.close()
        return "Güvenlik nedeniyle işlem durduruldu. Lütfen onay sayfasındaki linke tekrar tıklayın.", 400

    kolonlar = [
        "s_no", "koy_adi", "adi", "soyadi", "sayac_no", "senet_tutari", "sayac_tutari",
        "alinan_tutar", "malzeme_tutari", "malzeme_alinan", "senet_no",
        "senet_sahibi_adi", "senet_sahibi_soyadi", "telefon", "baba_adi",
        "montaj_tarihi", "odeme_tarihi", "odeme_sekli", "odemeyi_gonderen",
        "aciklama", "muhtara_odenecek", "muhtara_odenen", "fatura_no",
    ]
    sayisal_alanlar = {
        "senet_tutari", "sayac_tutari", "alinan_tutar", "malzeme_tutari",
        "malzeme_alinan", "muhtara_odenecek", "muhtara_odenen",
    }

    def _sayi(v):
        try:
            return float(v) if v not in (None, "") else 0.0
        except ValueError:
            return 0.0

    def _metin(v):
        return (v or "").strip()

    toplu_degerler = []
    for satir in satirlar:
        degerler = []
        for kolon in kolonlar:
            deger_ham = satir.get(kolon, "")
            if kolon == "s_no":
                degerler.append(deger_ham if deger_ham not in (None, "") else None)
            elif kolon in sayisal_alanlar:
                degerler.append(_sayi(deger_ham))
            else:
                degerler.append(_metin(deger_ham))
        toplu_degerler.append(tuple(degerler))

    kolonlar_sql = ", ".join(kolonlar)
    yer_tutucular = ", ".join(["%s"] * len(kolonlar))
    cur.executemany(
        f"INSERT INTO abone ({kolonlar_sql}) VALUES ({yer_tutucular})",
        toplu_degerler,
    )
    db.commit()
    eklenen = len(toplu_degerler)
    cur.close()

    return f"""
    <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
    <h2 style="color:#0a0">Başarılı</h2>
    <p><b>{eklenen}</b> abone kaydı başarıyla eklendi.</p>
    <p><a href="/abone">Abone listesine git</a></p>
    </body></html>
    """


@app.route("/admin/toplu-ariza-yukle", methods=["GET", "POST"])
@login_required
def toplu_ariza_yukle():
    veri_dosyasi = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ariza_toplu_veri.b64")
    if not os.path.exists(veri_dosyasi):
        return (
            "Veri dosyası (ariza_toplu_veri.b64) bulunamadı. "
            "Bu dosyanın app.py ile aynı klasörde (repo kök dizininde) olduğundan emin olun.",
            404,
        )

    with open(veri_dosyasi, "rb") as f:
        b64_veri = f.read()
    csv_metin = gzip.decompress(base64.b64decode(b64_veri)).decode("utf-8-sig")
    satirlar = list(csv.DictReader(io.StringIO(csv_metin)))

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM ariza")
    mevcut_sayi = cur.fetchone()["c"]

    onay = request.method == "POST" and request.form.get("onayla") == "1"
    zorla = request.method == "POST" and request.form.get("zorla") == "1"
    _csrf_gizli_alan = f'<input type="hidden" name="csrf_token" value="{generate_csrf()}">'

    if not onay:
        if mevcut_sayi > 50:
            aksiyon = (
                f"<p style='color:#b00;font-weight:bold'>Dikkat: tabloda hâlihazırda {mevcut_sayi} kayıt var. "
                f"Bu işlem mevcut kayıtları SİLMEZ, üzerine {len(satirlar)} yeni kayıt EKLER. "
                f"Bu veriyi daha önce yüklediyseniz tekrar yüklemeyin, kayıtlar çiftlenir.</p>"
                f"<form method='post'>{_csrf_gizli_alan}"
                f"<input type='hidden' name='onayla' value='1'><input type='hidden' name='zorla' value='1'>"
                f"<button type='submit' style='font-size:20px;color:#b00;background:none;border:none;"
                f"text-decoration:underline;cursor:pointer;padding:0;'>"
                f"Yine de devam et ve {len(satirlar)} kaydı ekle</button></form>"
            )
        else:
            aksiyon = (
                f"<form method='post'>{_csrf_gizli_alan}"
                f"<input type='hidden' name='onayla' value='1'>"
                f"<button type='submit' style='font-size:20px;background:none;border:none;"
                f"text-decoration:underline;cursor:pointer;padding:0;'>"
                f"Evet, {len(satirlar)} kaydı içe aktar</button></form>"
            )
        cur.close()
        return f"""
        <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
        <h2>Toplu Arıza Yükleme</h2>
        <p>Excel dosyasından hazırlanan <b>{len(satirlar)}</b> arıza kaydı,
        veritabanındaki <b>ariza</b> tablosuna eklenmeye hazır.</p>
        <p>Şu anda tabloda <b>{mevcut_sayi}</b> kayıt bulunuyor.</p>
        {aksiyon}
        </body></html>
        """

    if mevcut_sayi > 50 and not zorla:
        cur.close()
        return "Güvenlik nedeniyle işlem durduruldu. Lütfen onay sayfasındaki linke tekrar tıklayın.", 400

    kolonlar = [
        "s_no", "ozel_s_no", "koy_adi", "yeni_seri_no", "seri_no", "adi", "soyadi",
        "ariza_ucret", "alinan_ucret", "gelis_tarihi", "takilan_tarih",
        "sayac_kredisi", "islem_aciklama",
    ]
    sayisal_alanlar = {"ariza_ucret", "alinan_ucret"}

    def _sayi(v):
        try:
            return float(v) if v not in (None, "") else 0.0
        except ValueError:
            return 0.0

    def _metin(v):
        return (v or "").strip()

    toplu_degerler = []
    for satir in satirlar:
        degerler = []
        for kolon in kolonlar:
            deger_ham = satir.get(kolon, "")
            if kolon == "s_no":
                degerler.append(deger_ham if deger_ham not in (None, "") else None)
            elif kolon in sayisal_alanlar:
                degerler.append(_sayi(deger_ham))
            else:
                degerler.append(_metin(deger_ham))
        toplu_degerler.append(tuple(degerler))

    kolonlar_sql = ", ".join(kolonlar)
    yer_tutucular = ", ".join(["%s"] * len(kolonlar))
    cur.executemany(
        f"INSERT INTO ariza ({kolonlar_sql}) VALUES ({yer_tutucular})",
        toplu_degerler,
    )
    db.commit()
    eklenen = len(toplu_degerler)
    cur.close()

    return f"""
    <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
    <h2 style="color:#0a0">Başarılı</h2>
    <p><b>{eklenen}</b> arıza kaydı başarıyla eklendi.</p>
    <p><a href="/ariza">Arıza listesine git</a></p>
    </body></html>
    """


@app.route("/admin/tarih-formati-duzelt", methods=["GET", "POST"])
@login_required
def tarih_formati_duzelt():
    onay = request.method == "POST" and request.form.get("onayla") == "1"
    db = get_db()
    cur = db.cursor()

    hedefler = [
        ("abone", ["montaj_tarihi", "odeme_tarihi", "odeme_gun_sozu"]),
        ("ariza", ["gelis_tarihi", "takilan_tarih", "teslim_tarihi"]),
    ]

    bulunanlar = []
    for tablo, kolonlar in hedefler:
        for kolon in kolonlar:
            cur.execute(f"SELECT id, {kolon} FROM {tablo} WHERE {kolon} IS NOT NULL AND {kolon} != ''")
            for satir in cur.fetchall():
                iso = _ddmmyyyy_to_iso(satir[kolon])
                if iso:
                    bulunanlar.append((tablo, kolon, satir["id"], satir[kolon], iso))

    if not onay:
        cur.close()
        if not bulunanlar:
            return """
            <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
            <h2>Tarih Formatı Düzeltme</h2>
            <p>Düzeltilmesi gereken GG.AA.YYYY formatında kayıt bulunamadı. Tüm tarihler zaten doğru formatta.</p>
            <p><a href="/abone">Abone listesine git</a></p>
            </body></html>
            """
        return f"""
        <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
        <h2>Tarih Formatı Düzeltme</h2>
        <p><b>{len(bulunanlar)}</b> tarih alanı, aktarım sırasında GG.AA.YYYY metin formatında kaydedilmiş.
        Bu işlem bunları veritabanının beklediği YYYY-AA-GG formatına çevirecek. Görünen tarihler
        (GG.AA.YYYY) değişmeyecek, sadece arka plandaki kayıt biçimi düzelecek.</p>
        <form method="post">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">
        <input type="hidden" name="onayla" value="1">
        <button type="submit" style="font-size:20px;background:none;border:none;text-decoration:underline;cursor:pointer;padding:0;">
        Evet, {len(bulunanlar)} tarihi düzelt</button>
        </form>
        </body></html>
        """

    for tablo, kolon, kayit_id, eski, iso in bulunanlar:
        cur.execute(f"UPDATE {tablo} SET {kolon} = %s WHERE id = %s", (iso, kayit_id))
    db.commit()
    duzeltilen = len(bulunanlar)
    cur.close()

    return f"""
    <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
    <h2 style="color:#0a0">Başarılı</h2>
    <p><b>{duzeltilen}</b> tarih alanı düzeltildi.</p>
    <p><a href="/abone">Abone listesine git</a></p>
    </body></html>
    """


@app.route("/admin/fotograflari-kucult", methods=["GET", "POST"])
@login_required
def fotograflari_kucult():
    """Bu route eklenmeden ÖNCE yüklenmiş, hâlâ orijinal boyutta duran fotoğrafları tek seferlik küçültür."""
    onay = request.method == "POST" and request.form.get("onayla") == "1"
    db = get_db()
    cur = db.cursor()

    ESIK_BAYT = 300 * 1024
    hedefler = [("abone_fotograf", "abone_id"), ("ariza_fotograf", "ariza_id")]

    bulunanlar = []
    toplam_eski_boyut = 0
    for tablo, _sahip_kolonu in hedefler:
        cur.execute(
            f"SELECT id, content_type, octet_length(icerik) AS boyut FROM {tablo} "
            f"WHERE content_type LIKE %s AND octet_length(icerik) > %s",
            ("image/%", ESIK_BAYT),
        )
        for satir in cur.fetchall():
            bulunanlar.append((tablo, satir["id"], satir["boyut"]))
            toplam_eski_boyut += satir["boyut"]

    if not onay:
        cur.close()
        if not bulunanlar:
            return """
            <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
            <h2>Fotoğrafları Küçült</h2>
            <p>Küçültülmesi gereken (300 KB üstü) fotoğraf bulunamadı. Tüm fotoğraflar zaten küçük boyutta.</p>
            <p><a href="/abone">Abone listesine git</a></p>
            </body></html>
            """
        return f"""
        <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
        <h2>Fotoğrafları Küçült</h2>
        <p><b>{len(bulunanlar)}</b> fotoğraf, toplam <b>{toplam_eski_boyut/1024/1024:.1f} MB</b>,
        bu özellik eklenmeden önce yüklendiği için hâlâ orijinal (küçültülmemiş) boyutta duruyor —
        bu yüzden hem kayıt formunda hem de fotoğrafı açarken yavaş yükleniyorlar.</p>
        <p>Bu işlem bu fotoğrafları ekranda göstermek için yeterli ama çok daha küçük boyuta
        küçültecek — fotoğraflar KAYBOLMAZ, sadece dosya boyutu küçülür. Video dosyalarına dokunulmaz.</p>
        <form method="post">
        <input type="hidden" name="csrf_token" value="{generate_csrf()}">
        <input type="hidden" name="onayla" value="1">
        <button type="submit" style="font-size:20px;background:none;border:none;text-decoration:underline;cursor:pointer;padding:0;">
        Evet, {len(bulunanlar)} fotoğrafı küçült</button>
        </form>
        </body></html>
        """

    toplam_yeni_boyut = 0
    basarisiz = 0
    for tablo, foto_id, _eski_boyut in bulunanlar:
        cur.execute(f"SELECT icerik FROM {tablo} WHERE id = %s", (foto_id,))
        kayit = cur.fetchone()
        if not kayit:
            continue
        icerik = bytes(kayit["icerik"])
        yeni_icerik, yeni_tur = _fotografi_kucult(icerik)
        if yeni_tur:
            cur.execute(
                f"UPDATE {tablo} SET icerik = %s, content_type = %s WHERE id = %s",
                (psycopg2.Binary(yeni_icerik), yeni_tur, foto_id),
            )
            toplam_yeni_boyut += len(yeni_icerik)
        else:
            basarisiz += 1
            toplam_yeni_boyut += len(icerik)
    db.commit()
    kucultulen = len(bulunanlar) - basarisiz
    cur.close()

    return f"""
    <html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;line-height:1.5">
    <h2 style="color:#0a0">Başarılı</h2>
    <p><b>{kucultulen}</b> fotoğraf küçültüldü: {toplam_eski_boyut/1024/1024:.1f} MB → {toplam_yeni_boyut/1024/1024:.1f} MB.</p>
    {f'<p style="color:#a13a2e;">{basarisiz} fotoğraf küçültülemedi (bozuk/desteklenmeyen dosya), orijinal haliyle bırakıldı.</p>' if basarisiz else ''}
    <p><a href="/abone">Abone listesine git</a></p>
    </body></html>
    """


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug_modu = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug_modu)
